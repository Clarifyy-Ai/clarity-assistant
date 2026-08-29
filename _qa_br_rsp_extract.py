# -*- coding: utf-8 -*-
"""Extract TC-BR / TC-RSP / TC-PUB and related QA notes from the BB workbook."""
from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook

WB_PATH = Path(r"Clarify_AI_BB_Manual_QA_Workbook (2).xlsx")
OUT_JSON = Path("_qa_br_rsp_extract.json")
IMG_DIR = Path("_qa_br_rsp_images")

ID_PREFIXES = ("TC-BR-", "TC-RSP-", "TC-PUB-")
RESPONSIVE_RE = re.compile(
    r"responsive|viewport|breakpoint|mobile|tablet|layout.?break|overflow|media.?quer",
    re.I,
)
FAILISH_RE = re.compile(r"fail|partial|blocked|broken", re.I)


def cell_hyperlink(cell):
    hl = getattr(cell, "hyperlink", None)
    if not hl:
        return None
    target = getattr(hl, "target", None) or getattr(hl, "location", None)
    return str(target) if target else None


def extract_urls_from_text(text) -> list[str]:
    if not text:
        return []
    return re.findall(r"https?://[^\s\]\)\"'<>]+", str(text))


def row_dict(headers, values):
    d = {}
    for h, v in zip(headers, values):
        if h is None:
            continue
        key = str(h).strip()
        if isinstance(v, str):
            d[key] = v.strip()
        elif v is None:
            d[key] = None
        else:
            d[key] = v
    return d


def summarize_case(rec: dict) -> str:
    tid = rec.get("Test Case ID") or "?"
    pf = rec.get("Pass / Fail") or ""
    notes = rec.get("Notes") or ""
    actual = rec.get("Actual Result") or ""
    urls = rec.get("urls") or []
    evidence = rec.get("evidence_images") or []
    lines = [
        f"ID: {tid}",
        f"Pass/Fail: {pf}",
        f"Notes: {notes}",
        f"Actual: {actual}",
    ]
    if urls:
        lines.append("URLs: " + " | ".join(urls))
    else:
        lines.append("URLs: (none)")
    if evidence:
        bits = [
            f"{e.get('cell')}->{e.get('saved_path')}" for e in evidence
        ]
        lines.append("Evidence images: " + " | ".join(bits))
    return "\n".join(lines)


def extract_richdata_images(wb_path: Path, img_dir: Path):
    """Excel 365 cell images via richData (vm index), not classic drawings."""
    img_dir.mkdir(exist_ok=True)
    images = []
    with zipfile.ZipFile(wb_path) as zf:
        rid_media = {}
        for rel in ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels")):
            rid = rel.attrib.get("Id")
            tgt = rel.attrib.get("Target", "")
            if "media" in tgt:
                rid_media[rid] = "xl/media/" + Path(tgt).name

        value_media = []
        rv_root = ET.fromstring(zf.read("xl/richData/richValueRel.xml"))
        for child in rv_root:
            for attr, val in child.attrib.items():
                if val.startswith("rId") and val in rid_media:
                    value_media.append(rid_media[val])

        for n in zf.namelist():
            if n.startswith("xl/media/"):
                out = img_dir / Path(n).name
                if not out.exists():
                    out.write_bytes(zf.read(n))

        wbxml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = []
        for sh in wbxml.findall(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
        ):
            sheets.append(
                (
                    sh.attrib.get("name"),
                    sh.attrib.get(
                        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                    ),
                )
            )
        wb_rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_sheetpath = {
            rel.attrib.get("Id"): "xl/" + rel.attrib.get("Target", "").lstrip("/")
            for rel in wb_rels
        }

        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for sname, rid in sheets:
            spath = rid_to_sheetpath.get(rid)
            if not spath or spath not in zf.namelist():
                continue
            root = ET.fromstring(zf.read(spath))
            for c in root.findall(".//m:c", ns):
                vm = c.attrib.get("vm")
                if vm is None:
                    continue
                ref = c.attrib.get("r")
                mref = re.match(r"([A-Z]+)(\d+)", ref or "")
                if not mref:
                    continue
                col_letters, row_s = mref.group(1), int(mref.group(2))
                col_num = 0
                for ch in col_letters:
                    col_num = col_num * 26 + (ord(ch) - 64)
                vm_i = int(vm)
                media = None
                # vm appears 1-based into richValueRels in this workbook
                if 1 <= vm_i <= len(value_media):
                    media = value_media[vm_i - 1]
                images.append(
                    {
                        "sheet": sname,
                        "cell": ref,
                        "excel_row": row_s,
                        "excel_col": col_num,
                        "vm_index": vm_i,
                        "media": media,
                        "saved_path": str(img_dir / Path(media).name) if media else None,
                    }
                )
        media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]
        package = {
            "media_files_count": len(media_files),
            "drawing_files": [n for n in zf.namelist() if "drawing" in n.lower()],
            "rich_data": True,
        }
    return images, package


def main():
    print(f"Loading {WB_PATH} ...")
    wb = load_workbook(WB_PATH, data_only=False, keep_links=True)

    sheet_names = list(wb.sheetnames)
    print("\n=== SHEET NAMES ===")
    for s in sheet_names:
        print(f"  - {s}")

    rich_images, package = extract_richdata_images(WB_PATH, IMG_DIR)
    media_note = (
        f"Excel richData cell images (not classic xdr drawings). "
        f"{package['media_files_count']} xl/media files extracted to {IMG_DIR}/. "
        f"{len(rich_images)} cell image refs with anchors (sheet/cell/row/col). "
        "openpyxl ws._images empty; media is embedded via rich values."
    )
    print("\n=== MEDIA ===")
    print(media_note)

    result = {
        "workbook": str(WB_PATH.resolve()),
        "sheet_names": sheet_names,
        "images": rich_images,
        "media_note": media_note,
        "xlsx_package": package,
        "tc_br": [],
        "tc_rsp": [],
        "tc_pub": [],
        "responsive_fail_partial_admin_notes": [],
        "anushka_extra_bug": [],
        "summaries_br_rsp_text": "",
    }

    ws = wb["05 All Test Cases"]
    header_row = 4
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    col_index = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
    note_cols = []
    for name in (
        "Notes",
        "Actual Result",
        "Screenshot / Evidence Requirement",
        "Defect ID",
    ):
        if name in col_index:
            note_cols.append((name, col_index[name]))

    print("\n=== HEADERS (row 4) ===")
    print(headers)

    images_by_row = {}
    for im in rich_images:
        if im.get("sheet") == "05 All Test Cases":
            images_by_row.setdefault(im["excel_row"], []).append(im)

    responsive_admin = []

    for r in range(header_row + 1, ws.max_row + 1):
        tid = ws.cell(r, 1).value
        if tid is None:
            continue
        tid_s = str(tid).strip()
        values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        rec = row_dict(headers, values)

        hyperlinks = {}
        urls = set()
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            hl = cell_hyperlink(cell)
            hname = str(headers[c - 1] or f"col{c}").strip()
            if hl:
                hyperlinks[hname] = hl
                urls.add(hl)
            for u in extract_urls_from_text(cell.value):
                urls.add(u)

        rec["row"] = r
        rec["hyperlinks"] = hyperlinks
        rec["urls"] = sorted(urls)
        rec["evidence_images"] = images_by_row.get(r, [])

        notes = str(rec.get("Notes") or "")
        actual = str(rec.get("Actual Result") or "")
        pf = str(rec.get("Pass / Fail") or "")
        blob = f"{notes}\n{actual}\n{pf}"

        if any(tid_s.startswith(p) for p in ID_PREFIXES):
            if tid_s.startswith("TC-BR-"):
                result["tc_br"].append(rec)
            elif tid_s.startswith("TC-RSP-"):
                result["tc_rsp"].append(rec)
            elif tid_s.startswith("TC-PUB-"):
                result["tc_pub"].append(rec)

        if FAILISH_RE.search(pf) and RESPONSIVE_RE.search(blob):
            if not any(x.get("row") == r for x in responsive_admin):
                responsive_admin.append(rec)

    result["responsive_fail_partial_admin_notes"] = responsive_admin
    result["images_on_br_rsp_pub_rows"] = [
        im
        for im in rich_images
        if im.get("excel_row")
        in {c["row"] for c in result["tc_br"] + result["tc_rsp"] + result["tc_pub"]}
    ]

    print(
        f"\n=== COUNTS === TC-BR={len(result['tc_br'])} "
        f"TC-RSP={len(result['tc_rsp'])} TC-PUB={len(result['tc_pub'])} "
        f"responsive Fail/Partial={len(responsive_admin)}"
    )

    aws = wb["Anushka-Extra bug"]
    a_headers = None
    a_header_row = 1
    for r in range(1, min(15, aws.max_row or 15) + 1):
        vals = [aws.cell(r, c).value for c in range(1, (aws.max_column or 20) + 1)]
        nonempty = [v for v in vals if v is not None and str(v).strip()]
        if len(nonempty) >= 3:
            joined = " ".join(str(v) for v in nonempty).lower()
            if any(
                k in joined
                for k in (
                    "bug",
                    "id",
                    "title",
                    "status",
                    "severity",
                    "notes",
                    "description",
                    "module",
                )
            ):
                a_headers = vals
                a_header_row = r
                break
    if a_headers is None:
        a_headers = [
            aws.cell(1, c).value for c in range(1, (aws.max_column or 20) + 1)
        ]
        a_header_row = 1

    for r in range(a_header_row + 1, (aws.max_row or 0) + 1):
        values = [aws.cell(r, c).value for c in range(1, len(a_headers) + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        rec = row_dict(a_headers, values)
        rec["row"] = r
        urls = set()
        hyperlinks = {}
        for c, h in enumerate(a_headers, 1):
            cell = aws.cell(r, c)
            hl = cell_hyperlink(cell)
            if hl:
                hyperlinks[str(h or f"col{c}")] = hl
                urls.add(hl)
            for u in extract_urls_from_text(cell.value):
                urls.add(u)
        rec["hyperlinks"] = hyperlinks
        rec["urls"] = sorted(urls)
        result["anushka_extra_bug"].append(rec)

    print(f"Anushka-Extra bug rows: {len(result['anushka_extra_bug'])}")

    parts = ["=" * 60, "TC-BR CASES", "=" * 60]
    for rec in result["tc_br"]:
        parts.append(summarize_case(rec))
        parts.append("-" * 40)
    parts += ["", "=" * 60, "TC-RSP CASES", "=" * 60]
    for rec in result["tc_rsp"]:
        parts.append(summarize_case(rec))
        parts.append("-" * 40)
    summary_text = "\n".join(parts)
    result["summaries_br_rsp_text"] = summary_text

    def default(o):
        if hasattr(o, "isoformat"):
            return o.isoformat()
        return str(o)

    OUT_JSON.write_text(
        json.dumps(result, indent=2, ensure_ascii=False, default=default),
        encoding="utf-8",
    )
    print(f"\nWrote {OUT_JSON}")
    print("\n\n########## FULL BR + RSP SUMMARY ##########\n")
    print(summary_text)
    wb.close()


if __name__ == "__main__":
    main()
