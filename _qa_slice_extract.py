"""Extract target QA cases for remediation — not a test file."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

WB = Path(r"C:\Users\TECH-GENIUSES\Downloads\clarity-assistant\Clarify_AI_BB_Manual_QA_Workbook (2).xlsx")
EXTRACT = Path("_qa_cases_extract.json")
OUT = Path("_qa_target_slice.json")

TARGET = {
    "TC-BILL-001",
    "TC-BILL-002",
    "TC-BILL-003",
    "TC-BILL-004",
    "TC-BILL-005",
    "TC-BILL-006",
    "TC-BILL-007",
    "TC-BILL-008",
    "TC-REG-006",
    "TC-REG-013",
    "TC-JRN-004",
    "TC-CR-001",
    "TC-CR-002",
    "TC-CR-003",
    "TC-CR-004",
    "TC-CR-005",
    "TC-CR-006",
    "TC-PREP-007",
    "TC-FB-005",
    "TC-REG-005",
    "TC-SCH-001",
    "TC-SCH-002",
    "TC-SCH-003",
    "TC-SCH-004",
    "TC-SCH-005",
    "TC-INT-001",
    "TC-INT-002",
    "TC-INT-003",
    "TC-INT-004",
    "TC-INT-005",
    "TC-SET-010",
    "TC-SET-001",
    "TC-SET-002",
    "TC-SET-003",
    "TC-SET-004",
    "TC-SET-005",
    "TC-SET-006",
    "TC-SET-007",
    "TC-SET-008",
    "TC-SET-009",
    "TC-SET-011",
    "TC-SET-012",
    "TC-SET-013",
    "TC-SET-014",
    "TC-SET-015",
    "TC-NTF-001",
    "TC-NTF-002",
    "TC-NTF-003",
    "TC-NTF-004",
}


def main():
    data = json.loads(EXTRACT.read_text(encoding="utf-8"))
    cases = data.get("test_cases") or []
    out = [c for c in cases if (c.get("Test Case ID") or "") in TARGET]
    print(f"Matched {len(out)} / {len(TARGET)} from extract")

    wb = load_workbook(WB, data_only=False)
    anushka = []
    ws = wb["Anushka-Extra bug"]
    headers = [c.value for c in ws[1]]
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers)) if headers[i]}
        if any(v is not None and str(v).strip() for v in row.values()):
            anushka.append(row)

    # Defect log: find header row
    ws3 = wb["39 Defect Log"]
    defect_header_row = None
    defect_headers = []
    for r in range(1, 15):
        vals = [ws3.cell(r, c).value for c in range(1, 25)]
        joined = " ".join(str(v) for v in vals if v)
        if "Defect ID" in joined or "Test Case" in joined:
            defect_header_row = r
            defect_headers = vals
            break
    defects = []
    if defect_header_row:
        for r in range(defect_header_row + 1, ws3.max_row + 1):
            row = {}
            for i, h in enumerate(defect_headers):
                if not h:
                    continue
                row[str(h)] = ws3.cell(r, i + 1).value
            if any(v is not None and str(v).strip() for v in row.values()):
                tid = str(row.get("Test Case ID") or row.get("Related TC") or "")
                did = str(row.get("Defect ID") or "")
                blob = json.dumps(row, default=str)
                if any(t in blob for t in TARGET) or did.startswith("DEF-") or "BILL" in blob or "razorpay" in blob.lower():
                    defects.append(row)

    payload = {"cases": out, "anushka": anushka, "defects_related": defects, "headers": data.get("headers")}
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"Wrote {OUT} cases={len(out)} anushka={len(anushka)} defects={len(defects)}")

    # Compact summary
    for c in sorted(out, key=lambda x: x.get("Test Case ID") or ""):
        tid = c.get("Test Case ID")
        keys_of_interest = [
            "Pass/Fail / Status",
            "Status",
            "Result",
            "Actual Result",
            "Expected Result",
            "Tester Comments / Notes",
            "Comments",
            "Notes",
            "Defect ID",
            "Evidence / Screenshots",
            "Evidence Links",
            "Priority",
            "Title",
            "Test Case Title",
            "Module",
        ]
        print(f"\n## {tid}")
        for k in keys_of_interest:
            if k in c and c[k]:
                print(f"  {k}: {str(c[k]).replace(chr(10), ' | ')[:500]}")
        # dump other non-empty interesting fields
        for k, v in c.items():
            if not v or k in keys_of_interest or k in ("Test Case ID", "Exact Steps", "Preconditions"):
                continue
            if any(s in k.lower() for s in ("actual", "comment", "note", "fail", "pass", "status", "evidence", "defect", "obs")):
                print(f"  {k}: {str(v).replace(chr(10), ' | ')[:500]}")

    print("\n=== ANUSHKA ===")
    for a in anushka:
        print(json.dumps(a, default=str, ensure_ascii=False)[:1500])
        print("---")


if __name__ == "__main__":
    main()
