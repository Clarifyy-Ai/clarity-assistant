"""Final production QA workbook — manual-only cases for Venkat & Sultana.

Does not copy passwords. Credentials live in gitignored .env.qa.local.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, Reference

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Clarify_AI_Final_Production_QA_Workbook.xlsx"
SITE = "https://clarify.ai.sltfinanceindia.com"
TODAY = date(2026, 8, 13)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
LABEL = Font(bold=True)
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
P0 = PatternFill("solid", fgColor="FCE4D6")
P1 = PatternFill("solid", fgColor="DDEBF7")
CRED = PatternFill("solid", fgColor="FFF2CC")
WARN = PatternFill("solid", fgColor="FCE4D6")
OK = PatternFill("solid", fgColor="E2EFDA")
VENKAT = PatternFill("solid", fgColor="DDEBF7")
SULTANA = PatternFill("solid", fgColor="E2EFDA")
PASS_F = PatternFill("solid", fgColor="C6EFCE")
FAIL_F = PatternFill("solid", fgColor="FFC7CE")
BLOCK_F = PatternFill("solid", fgColor="FFEB9C")


def case(
    tid: str,
    pri: str,
    module: str,
    scenario: str,
    pre: str,
    data: str,
    steps: str,
    expected: str,
    url: str,
    owner: str,
    why: str,
    covers: str,
    uiux: str,
) -> dict:
    return {
        "id": tid,
        "priority": pri,
        "module": module,
        "scenario": scenario,
        "pre": pre,
        "data": data,
        "steps": steps,
        "expected": expected,
        "url": url,
        "owner": owner,
        "why": why,
        "covers": covers,
        "uiux": uiux,
        "env": "Production (closed beta)",
    }


CASES: list[dict] = [
    # ── Venkat: host / auth / email / billing / security ──────────────
    case(
        "FP-001", "P0", "Production smoke",
        "Production host loads and is not localhost",
        "Incognito Chrome. Logged out. No VPN required.",
        f"URL={SITE}/ | Chrome latest desktop 1440px",
        "1. Open the production URL in Incognito.\n"
        "2. Wait up to 15 seconds for splash/spinner.\n"
        "3. Check the address bar is exactly the production host (no localhost, no 127.0.0.1).\n"
        "4. Confirm landing headline and primary CTA are visible.\n"
        "5. Open DevTools Console; note any red errors (ignore known 3rd-party noise).",
        "Splash ends. Landing shows Career Pilot branding and a real CTA. "
        "Address bar is production. No infinite spinner. Config failure shows a recoverable error, not a blank page.",
        f"{SITE}/", "Venkata",
        "Must be executed against the live frontend host after deploy. Agent cannot publish or hit production DNS.",
        "SMOKE-APP-001, SMOKE-APP-002, PUBLIC-001",
        "First viewport must show product name, headline, and CTA without a blocking overlay.",
    ),
    case(
        "FP-002", "P0", "Production smoke",
        "Legal + Help pages on production",
        "Incognito. Logged out.",
        f"{SITE}/terms | {SITE}/privacy | {SITE}/help",
        "1. Open /terms and confirm operator/legal copy loads (not a 404).\n"
        "2. Open /privacy and confirm privacy copy loads.\n"
        "3. Open /help. If a status-page URL is configured, it is a real link; if not, a mailto fallback is shown — do not invent a status domain.\n"
        "4. From the footer, click Pricing and Gov Exams and confirm they load.",
        "Terms and Privacy render. Help is usable. Footer links work. No leftover localhost links in the page.",
        f"{SITE}/terms", "Venkata",
        "Live host content and mailto/status configuration are ops-owned.",
        "PUBLIC-006, PUBLIC-011, PUBLIC-012, PUBLIC-015",
        "Legal pages must be readable on mobile 375px as well as desktop.",
    ),
    case(
        "FP-003", "P0", "Authentication / Signup",
        "New signup receives a branded verification email",
        "Incognito. Use a mailbox you control (not qa.free). Have access to that inbox.",
        f"{SITE}/signup | unique email you own | password from your vault (do not write it in this sheet)",
        "1. Open /signup.\n"
        "2. Fill Full name, Email, Password, Confirm password. Toggle both visibility eyes.\n"
        "3. Accept Terms. Submit Create account.\n"
        "4. Confirm the app shows Check your email (not onboarding, not dashboard).\n"
        "5. Open the inbox. Wait up to 5 minutes (check spam).\n"
        "6. Confirm From address is branded (not a raw noreply@supabase) if Resend is configured.\n"
        "7. Screenshot the email (redact nothing that is a password; there should be none).",
        "Account is created. User is held on verify-email. A verification message arrives. "
        "If email never arrives, mark Blocked and log Resend/SMTP as the blocker — do not Pass.",
        f"{SITE}/signup", "Venkata",
        "Requires a real mailbox and production send-email / Resend secrets.",
        "AUTH-SIGNUP-006, AUTH-SIGNUP-007, AUTH-VERIFY-001",
        "Password toggles work. Error if Terms unchecked. Duplicate email shows already registered.",
    ),
    case(
        "FP-004", "P0", "Authentication / Verify",
        "Verification link unlocks the app and is not localhost",
        "FP-003 inbox still open. New user still unverified.",
        "Verification email from FP-003",
        "1. Hover the Verify / Confirm link. Confirm the host is production, not http://localhost.\n"
        "2. Click the link.\n"
        "3. After confirm, the app should allow /onboarding (not /app/* yet if onboarding incomplete).\n"
        "4. Try opening /app/dashboard while still unverified in a second Incognito before clicking the link (or with a leftover unverified session): must stay on verify-email.\n"
        "5. After verify, complete onboarding later in FP-009.",
        "Link host is production. Click confirms email. Unverified users cannot use /app/*. "
        "If the link points at localhost, Fail (release blocker) even if the page loads.",
        f"{SITE}/verify-email", "Venkata",
        "Auth Site URL / redirect allowlist is production ops. Agent cannot change live Auth settings.",
        "AUTH-VERIFY-002, AUTH-VERIFY-004",
        "After verify, user is not dumped into a blank spinner.",
    ),
    case(
        "FP-005", "P0", "Authentication / Reset",
        "Forgot-password email and reset link work on production",
        "Logged out. Use qa.pro (password only from .env.qa.local). Inbox for that mailbox, or Admin can read Auth logs.",
        f"{SITE}/forgot-password | qa.pro@clarify.ai.test",
        "1. Open /forgot-password.\n"
        "2. Submit qa.pro@clarify.ai.test. Confirm a neutral success message (no user enumeration).\n"
        "3. Open the reset email. Confirm the link host is production, not localhost.\n"
        "4. Open the link, set a NEW password from the vault (do not type old/new passwords into the sheet).\n"
        "5. Log in with the new password.\n"
        "6. Re-open the same reset link — it must not work a second time.\n"
        "7. Immediately restore the vault password after the test (or note Blocked if you cannot restore).",
        "Neutral submit success. Reset link is production-hosted. Password changes. Reused link fails. "
        "User can sign in with the new password. Never paste the password into Evidence.",
        f"{SITE}/forgot-password", "Venkata",
        "Needs live SMTP/Resend and production Auth redirect URLs.",
        "AUTH-RESET-001, AUTH-RESET-002, AUTH-RESET-003, AUTH-RESET-004",
        "Reset form shows password rules and visibility toggle.",
    ),
    case(
        "FP-006", "P0", "Authentication / Login",
        "Pro login reaches Dashboard on production",
        "Logged out Incognito. Pro account from .env.qa.local.",
        f"{SITE}/login | qa.pro@clarify.ai.test",
        "1. Open /login.\n"
        "2. Toggle password visibility.\n"
        "3. Submit wrong password once — expect invalid credentials, stay on /login.\n"
        "4. Submit correct Pro credentials (from vault).\n"
        "5. Confirm land on /app/dashboard (not /app/live, not a blank spinner).\n"
        "6. Hard-refresh (F5). Session must remain.\n"
        "7. Open /app/dashboard in a second tab, then Log out from Settings. Both tabs must lose /app access.",
        "Wrong password stays on login. Correct login reaches Dashboard. Refresh keeps session. "
        "Logout clears access. No auth.role.load.timed_out infinite spinner.",
        f"{SITE}/login", "Venkata",
        "Live auth against production Supabase; agent mocks only.",
        "AUTH-LOGIN-001, AUTH-LOGIN-002, AUTH-LOGIN-003, AUTH-SESSION-001, AUTH-SESSION-002, SHELL-001",
        "Login heading Welcome back; Sign in button visible; no blocked overlay.",
    ),
    case(
        "FP-007", "P0", "Authentication / OAuth",
        "Google sign-in on production (N/A if provider disabled)",
        "Incognito. Google account you control. VITE_OAUTH_PROVIDERS must include Google in this environment.",
        f"{SITE}/login | Google",
        "1. On /login, confirm which OAuth buttons are actually shown (Google / GitHub / others).\n"
        "2. If Google is absent, mark N/A and screenshot the login card.\n"
        "3. Click Google. Complete consent.\n"
        "4. Confirm callback returns to production (not localhost) and lands on onboarding or dashboard.\n"
        "5. Sign out. Sign in with Google again (returning user) — must not create a duplicate profile.\n"
        "6. Repeat once but Cancel on the Google screen — return to /login with a clean error, not a crash.",
        "Shown providers match the environment allowlist. Success callback is production. "
        "Cancel is clean. No open redirect. No duplicate user. GitHub: run the same if the button is shown (record in Notes).",
        f"{SITE}/login", "Venkata",
        "Live OAuth client IDs / redirect allowlist are ops. Agent cannot complete Google consent.",
        "AUTH-OAUTH-001, AUTH-OAUTH-002, AUTH-OAUTH-005, AUTH-OAUTH-006, AUTH-OAUTH-007",
        "OAuth buttons match what is configured; no dead LinkedIn/Azure buttons if those providers are off.",
    ),
    case(
        "FP-008", "P0", "Authentication / Restricted",
        "Banned and past-due accounts cannot use the app",
        "Need qa.banned and qa.pastdue from .env.qa.local (seed via npm run qa:seed-accounts). Admin access to flip flags if missing.",
        f"{SITE}/login | qa.banned@clarify.ai.test | qa.pastdue@clarify.ai.test",
        "1. Log in as banned user.\n"
        "2. Confirm a suspended / banned screen. Try /app/dashboard and /app/live — must not load the product.\n"
        "3. Sign out.\n"
        "4. Log in as past-due user.\n"
        "5. Confirm billing-recovery / past-due messaging. Paid Edge actions (Prep generate, Live hint) must be blocked.\n"
        "6. As Admin, confirm you did not need to edit the JWT in DevTools for this to work.",
        "Banned: dedicated blocked screen, no /app shell. Past-due: recovery copy, paid functions denied. "
        "If accounts were never seeded, mark Blocked (not Fail) and ask ops to run qa:seed-accounts.",
        f"{SITE}/login", "Venkata",
        "Requires seeded banned/past-due users on the live project.",
        "AUTH-RESTRICT-001, AUTH-RESTRICT-002, AUTH-RESTRICT-003, AUTH-RESTRICT-004",
        "Blocked screens use Access Denied / suspended copy, with a way back to login or billing.",
    ),
    case(
        "FP-009", "P0", "Onboarding",
        "New verified user finishes onboarding on Dashboard (not Live)",
        "Use the FP-003/004 user (verified, onboarding incomplete). Or qa.onboarding if seeded.",
        f"{SITE}/onboarding",
        "1. After verify, confirm redirect to /onboarding (not /app/*).\n"
        "2. Step 1: enter Target role and Experience level. Continue stays disabled until both are set.\n"
        "3. Optional step: Skip microphone, skip resume, set anxiety 1–5, pick mic/speaker if devices exist.\n"
        "4. Refresh mid-flow — draft values must still be there.\n"
        "5. Click Continue to Dashboard / Skip optional setup.\n"
        "6. Confirm URL is /app/dashboard. Confirm a Start Practice CTA exists on Dashboard (do not auto-start Live).\n"
        "7. Revisit /onboarding without ?rerun=1 — must bounce back to Dashboard.",
        "Completion lands on /app/dashboard only. Live/Practice Coach is a separate CTA. "
        "Skip and anxiety persist. Already-onboarded users cannot get stuck in the wizard.",
        f"{SITE}/onboarding", "Venkata",
        "Needs a real unverified→verified production user. Agent only mocked this.",
        "ONBOARD-001, ONBOARD-003, ONBOARD-004, ONBOARD-005, ONBOARD-007, ONBOARD-010, ONBOARD-011",
        "Wizard is usable at 375px. Skip still requires role + level.",
    ),
    case(
        "FP-010", "P0", "Billing / Stripe",
        "Free user Checkout → test payment → Pro + credits once",
        "Logged in as qa.free (vault). Stripe TEST mode. Card 4242 4242 4242 4242 / 12/34 / 123 / ZIP 400001. Do not use a live card.",
        f"{SITE}/app/settings/billing | Stripe test card only",
        "1. Open /app/billing — must redirect to /app/settings/billing (not 404).\n"
        "2. Note current plan (Free) and credit balance.\n"
        "3. Click Upgrade / Pro (monthly or annual). Stripe Checkout must open with a real test Price ID (not a blank session).\n"
        "4. Pay with 4242… test card. Complete Checkout.\n"
        "5. Return to the app. Refresh. Plan must be Pro. Credits must increase once (not double).\n"
        "6. Open Customer Portal from Billing and confirm it loads.\n"
        "7. If Checkout fails with missing price/key, Fail and attach the on-screen error (no secrets).",
        "Billing path canonical. Checkout opens. Test payment succeeds. Server plan is Pro. "
        "Credits granted once. Portal opens. If Stripe live/test keys are missing, Blocked — not Pass.",
        f"{SITE}/app/settings/billing", "Venkata",
        "Live Stripe keys, price IDs, and webhook are ops. Agent cannot charge.",
        "BILL-STRIPE-001, BILL-STRIPE-002, BILL-STRIPE-004, BILL-STRIPE-005, BILL-STRIPE-008, BILL-PLAN-002",
        "Plan names on screen are Free / Pro / Max only (no Elite/Unlimited).",
    ),
    case(
        "FP-011", "P0", "Billing / Free gate",
        "Free user is denied paid Edge Functions with a structured error",
        "Logged in as qa.free (still Free — do this BEFORE FP-010, or use a second Free account).",
        f"{SITE}/app/prep | {SITE}/app/documents",
        "1. Confirm plan badge is Free.\n"
        "2. Attempt a Pro-only action (Company Research generate, or Gap Analysis if Free is gated).\n"
        "3. Confirm the UI shows upgrade / capability required — not a generic 500, not a silent empty success.\n"
        "4. Confirm credits are not deducted for a denied call.\n"
        "5. Open /pricing, switch Annual, confirm Pro yearly = monthly×12×0.8 (20% off) and CTA includes plan+interval.",
        "Paid capability is blocked with honest copy (CAPABILITY_REQUIRED / upgrade). No charge. Pricing math is exact 20%.",
        f"{SITE}/app/settings/billing", "Venkata",
        "Must hit deployed Edge Functions with a real Free JWT.",
        "BILL-PLAN-001, SEC-AUTH-002, GAP-010, COMPANY-002, PUBLIC-003, PUBLIC-004",
        "Upgrade CTA is visible; no false Unlimited copy.",
    ),
    case(
        "FP-012", "P0", "Settings / Deletion",
        "Disposable account deletion completes and cannot sign in after",
        "Use qa.disposable only (never Pro/Admin). Password from .env.qa.local. Confirm ops has a backup seed.",
        f"{SITE}/app/settings | Danger zone",
        "1. Log in as disposable user.\n"
        "2. Open Settings → Danger (or Account deletion).\n"
        "3. Note credits/profile exist.\n"
        "4. Type the required confirmation (DELETE or email). Submit.\n"
        "5. Modal must close. Status toast may say processing / completed — must not name internal function names.\n"
        "6. Wait until login fails for that email (up to 2 minutes). Try login — must fail.\n"
        "7. As Admin, confirm the user is gone or marked deleted. Do not screenshot secrets.",
        "Deletion is requested as a durable operation. User cannot log in afterwards. "
        "UI does not leak Edge Function names. If the migration is not applied remotely, Blocked.",
        f"{SITE}/app/settings", "Venkata",
        "Irreversible live delete; needs disposable fixture and deployed delete-account.",
        "SETTINGS-DANGER-001, SETTINGS-DANGER-003",
        "Confirm dialog is explicit. Logout/login after delete cannot recover the account.",
    ),
    case(
        "FP-013", "P0", "Security / RLS",
        "User B cannot read User A sessions, documents, or answers",
        "qa.user-a and qa.user-b seeded. Each has at least one session or document of their own.",
        f"{SITE}/app/sessions | {SITE}/app/documents | {SITE}/app/answers",
        "1. Log in as User A. Create or open one session, one document, one answer. Copy the detail URLs.\n"
        "2. Sign out. Log in as User B.\n"
        "3. Paste User A session URL, document URL, and answer URL.\n"
        "4. Each must 404 / Access denied / empty — never User A content.\n"
        "5. User B list pages show only User B rows.\n"
        "6. Sign out. As anonymous, open /app/sessions — must go to /login.",
        "No cross-user data. Anonymous cannot open /app/*. If User A/B were never seeded, Blocked.",
        f"{SITE}/app/sessions", "Venkata",
        "Needs two live authenticated users on production RLS.",
        "SEC-RLS-001, SEC-RLS-002, SEC-RLS-003, ANSWER-014, SESSION-006, ANALYTICS-010, SEC-AUTH-001",
        "Denied state uses Access Denied / empty, not another user’s name in the header.",
    ),
    case(
        "FP-014", "P0", "Admin",
        "Non-admin Access Denied; Admin console loads",
        "qa.pro (non-admin) and qa.admin from vault.",
        f"{SITE}/app/admin",
        "1. Log in as Pro (non-admin). Open /app/admin.\n"
        "2. Must show heading Access Denied, text You are not authorized, button Return to Dashboard — not a silent dashboard redirect.\n"
        "3. Sign out. Log in as Admin. Open /app/admin.\n"
        "4. Admin dashboard loads with nav (Users, Billing, Gov, etc.).\n"
        "5. Confirm Admin cannot be faked by editing localStorage/profile JSON in DevTools (refresh still non-admin if you try on Pro).",
        "Non-admin: Access Denied page. Admin: console loads. Client-side role spoof does not grant admin.",
        f"{SITE}/app/admin", "Venkata",
        "Live user_roles on production.",
        "ADMIN-001, ADMIN-002, ADMIN-003, ADMIN-018",
        "Access Denied has a clear heading and Return to Dashboard.",
    ),
    case(
        "FP-015", "P1", "Settings / Security",
        "Password change and MFA (if offered) on production",
        "Use a dedicated QA user you can restore. Not Admin-only.",
        f"{SITE}/app/settings/security",
        "1. Open Settings → Security.\n"
        "2. Change password using vault old/new values (do not record them in the sheet).\n"
        "3. Log out and in with the new password.\n"
        "4. If TOTP/MFA enrollment is shown, enroll with an authenticator app, sign out, sign in with code.\n"
        "5. If MFA is not offered, mark N/A for the MFA steps only and Pass password-change if it worked.",
        "Password change works. MFA either works end-to-end or is honestly absent (N/A), never a broken enroll button.",
        f"{SITE}/app/settings/security", "Venkata",
        "Live Auth factors; MFA is ops policy.",
        "SETTINGS-SECURITY-001",
        "Security page explains what will happen if MFA is enabled.",
    ),

    # ── Sultana: live product / devices / AI / exams ──────────────────
    case(
        "FP-016", "P0", "Practice Coach",
        "Setup wizard + microphone permission on production",
        "Logged in as Pro. Real microphone. Chrome desktop. Allow prompts.",
        f"{SITE}/app/live",
        "1. Open /app/live (or Dashboard → Start Practice).\n"
        "2. Complete the setup wizard: role, mode, acknowledgements. Next/Start stay disabled until required fields are valid.\n"
        "3. Refresh mid-wizard — draft must survive.\n"
        "4. Select a real microphone. Allow the browser permission.\n"
        "5. If permission is denied, confirm honest guidance (Settings / browser), not a hang.\n"
        "6. On a phone (375px) open the same URL and confirm an honest notice that overlay/Practice Coach needs desktop — no fake overlay.",
        "Wizard validates. Mic list is real devices. Denied permission is explained. Mobile is honest about desktop requirement.",
        f"{SITE}/app/live", "Sultana",
        "Needs a physical mic, browser permission, and production host.",
        "LIVE-SETUP-001, LIVE-SETUP-002, LIVE-SETUP-003, LIVE-SETUP-006, LIVE-OVERLAY-013, SHELL-004",
        "Mobile More sheet still reaches Log out. Desktop overlay shortcuts are hidden on mobile.",
    ),
    case(
        "FP-017", "P0", "Practice Coach",
        "Live overlay: speech → transcript → one AI hint → credits once",
        "Pro user with credits. Deepgram configured. Mic working from FP-016. Second person or a recorded question helps.",
        f"{SITE}/app/live/overlay | Note credit balance before start",
        "1. Write down credit balance.\n"
        "2. Start Practice Coach until overlay is visible.\n"
        "3. Speak a clear interview question (e.g. Tell me about a time you led a project).\n"
        "4. Confirm transcript text appears (not stuck on Listening…).\n"
        "5. Trigger one hint (button or Ctrl+Shift+H). Confirm a framework/hint appears.\n"
        "6. Double-click the hint control rapidly — credits must drop once, not twice.\n"
        "7. End session. Post-session summary appears. New balance = start minus expected cost.",
        "Transcript is real speech. One hint charges once. Summary is shown. No stealth/invisible claims. "
        "If Deepgram/AI keys are missing, Blocked with the on-screen error.",
        f"{SITE}/app/live", "Sultana",
        "Live STT + AI provider + credits on deployed functions.",
        "LIVE-OVERLAY-001, LIVE-OVERLAY-003, LIVE-OVERLAY-004, LIVE-OVERLAY-006, LIVE-OVERLAY-014, LIVE-OVERLAY-016, CREDIT-003",
        "Overlay is readable; credit confirmation after chargeable action.",
    ),
    case(
        "FP-018", "P0", "Practice Coach",
        "Overlay remains visible in screen share (honest, not stealth)",
        "Desktop Chrome. Meet/Zoom/Teams or Chrome tab share. Pro session running.",
        "Screen share of the overlay window or the browser tab",
        "1. Start Practice Coach overlay.\n"
        "2. Share that window/tab to a second account or a phone camera pointed at a second screen.\n"
        "3. Confirm the overlay IS visible to the other party (product is interview prep, not stealth).\n"
        "4. Confirm in-product copy does not claim invisible / undetectable overlay.\n"
        "5. Dock/resize overlay; it must stay usable.\n"
        "6. Try Ctrl+Shift+H and Ctrl+Shift+P if focus is on the app.",
        "Overlay is visible in share. No stealth-evasion marketing. Hotkeys work on desktop. Layout survives resize.",
        f"{SITE}/app/live", "Sultana",
        "Requires a real meeting share. Agent cannot screen-share.",
        "LIVE-OVERLAY-002, LIVE-OVERLAY-007, LIVE-OVERLAY-009, HONESTY-003",
        "Always-on-top is opt-in, not forced. Copy stays consumer interview-prep.",
    ),
    case(
        "FP-019", "P1", "Practice Coach",
        "Deepgram disconnect and long silence do not freeze the session",
        "Pro live session. Optionally disable network for 10 seconds or unplug headset.",
        f"{SITE}/app/live/overlay",
        "1. Start listening.\n"
        "2. Stay silent for 30 seconds — UI must not lock; you can still End session.\n"
        "3. Toggle mic off/on or drop network 10s then restore.\n"
        "4. Confirm a recoverable error or reconnect — not a white overlay forever.\n"
        "5. End session cleanly.",
        "Silence is idle, not crash. Disconnect is recoverable or honestly failed. User can always End.",
        f"{SITE}/app/live", "Sultana",
        "Live media/network; agent has no Deepgram socket.",
        "LIVE-OVERLAY-010, LIVE-OVERLAY-011, RELIABILITY-001",
        "Error text is human, not a stack trace.",
    ),
    case(
        "FP-020", "P0", "Mock Interview",
        "Full mock run to scorecard / debrief (or honest not_scored)",
        "Pro user. Credits available. Do not leave the tab during scoring.",
        f"{SITE}/app/mock",
        "1. Open /app/mock. Choose Behavioural, set role/company.\n"
        "2. Start. Confirm one session is created (not two if you double-click Start).\n"
        "3. Answer at least 2 questions (type or speak). Complete the flow.\n"
        "4. On completion, open the scorecard. If scoring fails or you gave no answers, label must be Not scored — never a fake 0.\n"
        "5. Open /app/debriefs (plural). Open the new item. /app/debrief (singular) must redirect to plural.\n"
        "6. Confirm credits charged once for the mock.",
        "One session. Scorecard is honest (numeric only if scored). Debrief lives under /app/debriefs. No double charge.",
        f"{SITE}/app/mock", "Sultana",
        "Live AI scoring/debrief on production.",
        "MOCK-001, MOCK-004, MOCK-005, MOCK-011, DEBRIEF-001, ANALYTICS-004, SESSION-003, SESSION-004",
        "Question text is fully visible (not clamped so you cannot read the fail reason).",
    ),
    case(
        "FP-021", "P1", "Prep Lab",
        "STAR / Rephraser generate once, save to Answer Bank, no invented jobs",
        "Pro user. Use your real resume facts only.",
        f"{SITE}/app/prep/star-builder | {SITE}/app/answers",
        "1. Open Prep Lab. Open STAR builder.\n"
        "2. Enter a real situation from the resume. Generate.\n"
        "3. Confirm output does not invent employers/metrics that were not in the prompt.\n"
        "4. Confirm credits drop once. Double-submit does not double-charge.\n"
        "5. Save to Answer Bank (confirm dialog). Open /app/answers and find it.\n"
        "6. /app/answer-bank must redirect to /app/answers.",
        "Generation is evidence-grounded. One charge. Answer appears in bank. Legacy answer-bank URL redirects.",
        f"{SITE}/app/prep", "Sultana",
        "Live prep-tool provider + credits.",
        "PREP-001, PREP-002, PREP-004, PREP-007, PREP-008, PREP-011, ANSWER-001, ANSWER-003, ANSWER-008",
        "Prep tools use the shared shell. Save confirm is explicit.",
    ),
    case(
        "FP-022", "P0", "Documents / Gap",
        "Upload resume + JD, run gap analysis, no duplicate re-charge",
        "Pro user. Sample PDF resume (<10 MB) and JD file/text (JD limit is 5 MB — show it before the picker).",
        f"{SITE}/app/documents",
        "1. Open Documents. Confirm Resume max 10 MB and JD max 5 MB are shown before file pickers.\n"
        "2. Upload a PDF resume. Wait until parsed.\n"
        "3. Upload the same PDF again — expect duplicate handling (DUPLICATE_DOCUMENT), no second parse charge.\n"
        "4. Add a JD (paste or file).\n"
        "5. Select resume + JD. Run Gap Analysis. Note credits before/after.\n"
        "6. Confirm matched/missing skills cite resume evidence — no invented employers.\n"
        "7. Refresh; analysis still there. Replace resume content; analysis marked stale if that UX exists.\n"
        "8. As Free (or a Free alt), the same Gap action must 403 capability — not a billed success.",
        "Limits visible first. Duplicate does not re-charge. Gap runs once, evidence-grounded, persists. "
        "Free is denied. If gap-analysis is not deployed, Blocked.",
        f"{SITE}/app/documents", "Sultana",
        "Live parse-resume + gap-analysis + storage.",
        "DOC-001, DOC-003, DOC-005, DOC-006, DOC-014, GAP-001, GAP-002, GAP-003, GAP-004, GAP-005, GAP-006, GAP-007, GAP-008, GAP-010",
        "File limits appear before the OS file dialog. Stale badge is understandable.",
    ),
    case(
        "FP-023", "P0", "Gov Exam Mock Tests",
        "Region gate: non-India user is not silently given APPSC",
        "One India profile (region IN or Asia/Kolkata) and one non-India profile (region US or America/New_York). Pro/Max as required.",
        f"{SITE}/app/mock-test | {SITE}/gov-exams",
        "1. Log in as non-India user. Open /app/mock-test.\n"
        "2. Confirm you are NOT quietly given an APPSC/India full paper. Expect redirect, gate, or explicit unavailable copy.\n"
        "3. Log in as India user. Hub opens. Search an exam.\n"
        "4. Try a full mock. If approved question count < configured paper size, generation must fail closed and offer a labeled Custom Practice Set — never a silently short “full” paper.\n"
        "5. If a certified full pack exists, generate once and confirm credits once.",
        "Non-India never receives India papers by locale spoofing alone. Full mock is fail-closed on short banks. Custom path is labeled.",
        f"{SITE}/app/mock-test", "Sultana",
        "Needs two real region profiles and live select-test-questions.",
        "GOV-001, GOV-002, GOV-003, GOV-GEN-001, GOV-GEN-003, GOV-GEN-005",
        "Independent-platform disclaimer visible. No predicted/leaked paper claims.",
    ),
    case(
        "FP-024", "P0", "Gov Exam Mock Tests",
        "In-test timer: manual submit once and expiry auto-submit",
        "India user. A generated (or custom) test with a short duration if ops can set one; otherwise use remaining time and document the wait.",
        f"{SITE}/app/mock-test/session/{{id}}",
        "1. Start a test. Confirm full-screen runner and a server-based timer (refresh does not reset the clock to full time).\n"
        "2. Answer at least one question. Click Submit. Confirm submit happens once (second click is idempotent / already completed).\n"
        "3. Start a second test (or coordinate a short expires_at with ops). Let timer hit 0 without clicking Submit.\n"
        "4. Confirm auto-submit (or next interaction shows completed). Score is honest; unanswered is not fabricated as 0 knowledge if not scored.\n"
        "5. User B must not open User A’s attempt URL.",
        "Timer is server-authoritative. Submit is idempotent. Expiry auto-submits. Attempts are isolated.",
        f"{SITE}/app/mock-test", "Sultana",
        "Live submit-test + timer. Agent cannot wait a real exam clock.",
        "GOV-TEST-001, GOV-TEST-002, GOV-TEST-007, GOV-TEST-008, GOV-TEST-009, GOV-CONTENT-010",
        "Timer remains visible on laptop 1366px. Submit confirmation is clear.",
    ),
    case(
        "FP-025", "P0", "Analytics",
        "Unscored sessions show Not scored, never a fake 0",
        "Pro user with at least one incomplete or failed-score session (use FP-020 if it ended not_scored). Otherwise create a mock and abort before answers.",
        f"{SITE}/app/analytics",
        "1. Open /app/analytics.\n"
        "2. Find a session with no score or failed scoring.\n"
        "3. Confirm the label is Not scored (or equivalent), not 0.\n"
        "4. A truly scored session may show 0 only if the server marked it scored with score 0.\n"
        "5. Charts have an accessible name (aria / title) beyond a blank graphic.",
        "Missing scores are not coerced to zero. Scored zeros remain possible only when status is scored.",
        f"{SITE}/app/analytics", "Sultana",
        "Needs production analytics-dashboard after redeploy.",
        "ANALYTICS-001, ANALYTICS-004",
        "Dashboard is readable; no chart-only-color meaning without text.",
    ),
    case(
        "FP-026", "P0", "Interview Day",
        "Interview Day checklist + Continue in browser launches Coach with context",
        "Pro user. Create an interview for TODAY (company + role) under /app/interviews.",
        f"{SITE}/app/interview-day | {SITE}/app/interviews",
        "1. Create an interview dated today with company and role. Save.\n"
        "2. From the list overflow/row, use Edit and change the role; save.\n"
        "3. Open /app/interview-day. Checklist loads. Toggle one item; refresh — it persists.\n"
        "4. Click Continue in browser (not only Electron).\n"
        "5. Confirm Practice Coach setup is pre-filled with that company/role.\n"
        "6. If desktop installer is unpublished, the page must say so honestly.",
        "Edit works from list. Checklist persists. Web path works with context. No fake desktop download.",
        f"{SITE}/app/interview-day", "Sultana",
        "Calendar/date + live navigation; Electron optional.",
        "INTERVIEW-001, INTERVIEW-002, INTERVIEW-003, DAY-001, DAY-002, DAY-006, LIVE-OVERLAY-017",
        "Continue in browser is visible without scrolling off a modal.",
    ),
    case(
        "FP-027", "P1", "Compatibility",
        "Chrome desktop + real phone: login, dashboard, More Logout",
        "Physical phone (iOS Safari or Android Chrome) plus desktop Chrome. Pro user.",
        f"{SITE}/login | phone 375px-class device",
        "1. Desktop Chrome: login → dashboard → one Prep or Documents click. Note any layout clip.\n"
        "2. Phone: open landing, login, dashboard.\n"
        "3. Open More. Scroll to Log out (must be reachable above the home indicator).\n"
        "4. Log out. Confirm /login.\n"
        "5. Optional: Firefox or Edge desktop smoke of login + dashboard only.",
        "Core path works on Chrome desktop and one real phone. More sheet Logout is reachable. No unusable overlap.",
        f"{SITE}/app/dashboard", "Sultana",
        "Real devices; agent Playwright is desktop-mocked only.",
        "COMPAT-001, MOBILE-001, MOBILE-003, MOBILE-004, MOBILE-005, SHELL-003, SHELL-004, SETTINGS-AUTH-001",
        "Bottom nav does not hide Logout. Tap targets are usable.",
    ),
    case(
        "FP-028", "P1", "Desktop / Electron",
        "Windows desktop installer overlay smoke (N/A if unpublished)",
        "Windows machine. Installer from the production download URL, or mark N/A if ops has not published one.",
        "Electron build from the in-app download / release folder",
        "1. If no installer is published, mark N/A and screenshot the honest unpublished message on Interview Day / Help.\n"
        "2. If published: install, sign in, start Practice Coach.\n"
        "3. Confirm overlay, mic permission, and End session.\n"
        "4. Confirm always-on-top is off by default.\n"
        "5. Quit; confirm shortcuts unregister (no stuck global hotkeys in other apps).",
        "Either honest N/A or a working Windows overlay smoke. Do not Pass a missing installer.",
        f"{SITE}/app/interview-day", "Sultana",
        "Physical Windows + signed build. Agent cannot run Electron device smoke here.",
        "LIVE-OVERLAY-017, Electron smoke checklist",
        "First-run permissions are explained before capture starts.",
    ),
]


HEADERS = [
    "Test Case ID",
    "Priority",
    "Module",
    "Test scenario",
    "Preconditions",
    "Test data",
    "Steps",
    "Expected result",
    "Actual result",
    "Status",
    "Environment",
    "Evidence",
    "Defect ID",
    "Notes",
    "URL",
    "QA Owner",
    "Date",
    "COMMENTS",
    "SCREENSHOT",
    "UI&UX",
    "Why manual (AI cannot)",
    "Replaces / covers IDs from Basic (7)",
]


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def main() -> None:
    wb = Workbook()

    # ── Instructions ────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["A1"] = "Career Pilot v1.0.0 — Final Production QA (Manual only)"
    ws0["A1"].font = TITLE_FONT
    ws0.merge_cells("A1:F1")
    ws0["A2"] = (
        "This is the last production pack for Venkata and Sultana. "
        "It is NOT a re-run of all 259 Basic (7) cases. Agent-verified unit/Playwright/static cases are excluded. "
        "Execute only what a human can do on the live host with real email, Stripe, mic, OAuth, and devices."
    )
    ws0["A2"].alignment = WRAP
    ws0.merge_cells("A2:F2")
    ws0.row_dimensions[2].height = 48

    info = [
        ("Test site (use this only)", f"{SITE}/"),
        ("Login", f"{SITE}/login"),
        ("Signup", f"{SITE}/signup"),
        ("Dashboard", f"{SITE}/app/dashboard"),
        ("Practice Coach", f"{SITE}/app/live"),
        ("Mock Interview", f"{SITE}/app/mock"),
        ("Prep Lab", f"{SITE}/app/prep"),
        ("Documents", f"{SITE}/app/documents"),
        ("Gov Exam Mock Tests", f"{SITE}/app/mock-test"),
        ("Billing (canonical)", f"{SITE}/app/settings/billing"),
        ("Debriefs (canonical)", f"{SITE}/app/debriefs"),
        ("Answers (canonical)", f"{SITE}/app/answers"),
        ("", ""),
        ("How to execute", ""),
        ("1", "Read Preconditions. Use the account in Test data. Passwords only from .env.qa.local — never this workbook."),
        ("2", "Open URL. Follow numbered Steps exactly."),
        ("3", "Compare to Expected result. Fill Actual result, Status, Evidence (screenshot link), Defect ID."),
        ("4", "Finish ALL P0 owned by you before P1."),
        ("5", "If a dependency is missing (Stripe, Resend, Deepgram, seed users), mark Blocked — never Pass."),
        ("6", "Never paste passwords, tokens, or API keys into Evidence / COMMENTS / SCREENSHOT."),
        ("", ""),
        ("Status values", "Not Run | Pass | Fail | Blocked | N/A"),
        ("Priority", "P0 release blocker | P1 high"),
        ("Browser", "Chrome latest (primary). FP-027 adds a real phone. FP-028 is Windows Electron or N/A."),
        ("Stripe TEST card", "4242 4242 4242 4242  |  Exp 12/34  |  CVC 123  |  ZIP 400001  (TEST mode only)"),
        ("", ""),
        ("Out of scope (already covered by agent / do not retest unless a Fail is found)", ""),
        ("A", "Route redirects: /app/rooms → dashboard, /app/debrief → /app/debriefs, /app/answer-bank → /app/answers, /app/billing → settings/billing"),
        ("B", "Unit: pricing 20%, score not_scored, India helpers, deletion states, duplicate hash, palette ranking"),
        ("C", "Playwright (mocked): Access Denied copy, Interview Day web CTA, More-sheet Logout, onboarding Skip → Dashboard"),
        ("D", "Do not revive Practice Rooms / WebRTC. Do not claim certified gov full-sim packs if bank is short."),
    ]
    r = 4
    for k, v in info:
        ws0.cell(r, 1, k).font = LABEL
        ws0.cell(r, 2, v).alignment = WRAP
        ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    cred_row = r + 1
    ws0.cell(cred_row, 1, "QA LOGIN EMAILS (passwords are NOT in this file)").font = Font(
        bold=True, size=12, color="C00000"
    )
    ws0.merge_cells(start_row=cred_row, start_column=1, end_row=cred_row, end_column=6)
    ws0.cell(cred_row + 1, 1, "Passwords: operator vault / .env.qa.local after npm run qa:seed-accounts. Rotate Admin first. Enable MFA.")
    ws0.merge_cells(start_row=cred_row + 1, start_column=1, end_row=cred_row + 1, end_column=6)

    for col, h in enumerate(["Role", "Email", "Password location", "Use for", "Primary owner"], 1):
        cell = ws0.cell(cred_row + 2, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    creds = [
        ("Free", "qa.free@clarify.ai.test", ".env.qa.local → QA_FREE_PASSWORD", "Gates, Checkout (before upgrade)", "Venkata"),
        ("Pro", "qa.pro@clarify.ai.test", ".env.qa.local → QA_PRO_PASSWORD", "Main product (Coach, Mock, Gap, Gov)", "Sultana"),
        ("Max", "qa.max@clarify.ai.test", ".env.qa.local → QA_MAX_PASSWORD", "Max-tier only if Pro is insufficient", "Sultana"),
        ("Admin", "qa.admin@clarify.ai.test", ".env.qa.local → QA_ADMIN_PASSWORD", "Admin console; ban/past-due flags", "Venkata"),
        ("Disposable", "qa.disposable@clarify.ai.test", ".env.qa.local → QA_DISPOSABLE_PASSWORD", "Account deletion only", "Venkata"),
        ("User A", "qa.user-a@clarify.ai.test", ".env.qa.local → QA_USER_A_PASSWORD", "RLS owner", "Venkata"),
        ("User B", "qa.user-b@clarify.ai.test", ".env.qa.local → QA_USER_B_PASSWORD", "RLS peer", "Venkata"),
        ("Banned", "qa.banned@clarify.ai.test", ".env.qa.local → QA_BANNED_PASSWORD", "Suspended screen", "Venkata"),
        ("Past-due", "qa.pastdue@clarify.ai.test", ".env.qa.local → QA_PAST_DUE_PASSWORD", "Billing recovery", "Venkata"),
        ("New mailbox", "(tester-owned address)", "Your inbox — not the sheet", "Signup + verify + reset emails", "Venkata"),
    ]
    for i, row in enumerate(creds):
        for c, val in enumerate(row, 1):
            cell = ws0.cell(cred_row + 3 + i, c, val)
            cell.fill = CRED
            cell.border = THIN
            cell.alignment = WRAP

    note = cred_row + 3 + len(creds) + 1
    ws0.cell(note, 1, "Assignment").font = LABEL
    ws0.cell(
        note + 1,
        1,
        "Venkata: FP-001 … FP-015 (host, auth, email, OAuth, billing, deletion, RLS, admin). "
        "Sultana: FP-016 … FP-028 (Coach, overlay, mock, prep, documents/gap, gov exams, analytics, Interview Day, mobile, Electron).",
    )
    ws0.merge_cells(start_row=note + 1, start_column=1, end_row=note + 1, end_column=6)
    ws0.cell(note + 1, 1).alignment = WRAP
    ws0.row_dimensions[note + 1].height = 40

    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 42
    ws0.column_dimensions["C"].width = 42
    ws0.column_dimensions["D"].width = 42
    ws0.column_dimensions["E"].width = 16
    ws0.column_dimensions["F"].width = 18
    ws0.row_dimensions[1].height = 22

    # ── Assignment ──────────────────────────────────────────────────
    wsa = wb.create_sheet("Assignment")
    wsa["A1"] = "Who does what"
    wsa["A1"].font = TITLE_FONT
    wsa.merge_cells("A1:F1")
    wsa["A2"] = (
        f"Total cases: {len(CASES)}  |  Basic (7) had 259 executed rows. "
        "This pack is the last human production gate."
    )
    wsa.merge_cells("A2:F2")

    for col, h in enumerate(
        ["QA Owner", "Case IDs", "P0 count", "P1 count", "Focus", "Do not start until"], 1
    ):
        cell = wsa.cell(4, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    venkat = [c for c in CASES if c["owner"] == "Venkata"]
    sultana = [c for c in CASES if c["owner"] == "Sultana"]
    assign_rows = [
        (
            "Venkata",
            "FP-001 to FP-015",
            sum(1 for c in venkat if c["priority"] == "P0"),
            sum(1 for c in venkat if c["priority"] == "P1"),
            "Production host, signup/verify/reset email, OAuth, banned/past-due, onboarding destination, Stripe Checkout/Portal, account deletion, RLS User A/B, Admin 403",
            "Resend/SMTP live; Stripe TEST keys + webhook; seed accounts including disposable + A/B + banned/past-due; Site URL is production",
        ),
        (
            "Sultana",
            "FP-016 to FP-028",
            sum(1 for c in sultana if c["priority"] == "P0"),
            sum(1 for c in sultana if c["priority"] == "P1"),
            "Practice Coach mic/STT/overlay/share, mock scorecard, prep+answer bank, documents/gap, gov region+timer, analytics not_scored, Interview Day web path, real phone, Electron N/A-or-smoke",
            "Deepgram + at least one AI provider on Edge; India and non-India profiles; sample PDF resume; optional Windows installer",
        ),
    ]
    fills = [VENKAT, SULTANA]
    for i, row in enumerate(assign_rows):
        for c, val in enumerate(row, 1):
            cell = wsa.cell(5 + i, c, val)
            cell.fill = fills[i]
            cell.border = THIN
            cell.alignment = WRAP
        wsa.row_dimensions[5 + i].height = 72

    wsa["A8"] = "Order inside each owner"
    wsa["A8"].font = LABEL
    wsa["A9"] = (
        "Venkata: FP-001 smoke → FP-003/004/005 email chain (new mailbox) → FP-006 Pro login → "
        "FP-011 Free gate BEFORE upgrading → FP-010 Stripe → remaining auth/security.\n"
        "Sultana: FP-016 setup/mic → FP-017 live hint → FP-020 mock → FP-022 documents/gap → FP-023/024 gov → rest."
    )
    wsa["A9"].alignment = WRAP
    wsa.merge_cells("A9:F9")
    wsa.row_dimensions[9].height = 48

    wsa.column_dimensions["A"].width = 14
    wsa.column_dimensions["B"].width = 22
    wsa.column_dimensions["C"].width = 12
    wsa.column_dimensions["D"].width = 12
    wsa.column_dimensions["E"].width = 55
    wsa.column_dimensions["F"].width = 55

    # ── Checklist ───────────────────────────────────────────────────
    ws = wb.create_sheet("Checklist")
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    for i, c in enumerate(CASES, 2):
        vals = [
            c["id"],
            c["priority"],
            c["module"],
            c["scenario"],
            c["pre"],
            c["data"],
            c["steps"],
            c["expected"],
            "",
            "Not Run",
            c["env"],
            "",
            "",
            "",
            c["url"],
            c["owner"],
            "",
            "",
            "",
            c["uiux"],
            c["why"],
            c["covers"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, val)
            cell.border = THIN
            cell.alignment = WRAP if col not in (1, 2, 10, 16) else CENTER
        ws.cell(i, 2).fill = P0 if c["priority"] == "P0" else P1
        ws.cell(i, 16).fill = VENKAT if c["owner"] == "Venkata" else SULTANA
        ws.row_dimensions[i].height = 96

    last = len(CASES) + 1
    style_header(ws, len(HEADERS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last}"

    dv = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,N/A"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"J2:J{last}")
    dv2 = DataValidation(type="list", formula1='"Venkata,Sultana"', allow_blank=False)
    ws.add_data_validation(dv2)
    dv2.add(f"P2:P{last}")

    ws.conditional_formatting.add(
        f"J2:J{last}",
        CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_F),
    )
    ws.conditional_formatting.add(
        f"J2:J{last}",
        CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_F),
    )
    ws.conditional_formatting.add(
        f"J2:J{last}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_F),
    )

    widths = {
        "A": 12, "B": 10, "C": 26, "D": 42, "E": 36, "F": 36, "G": 58, "H": 48,
        "I": 28, "J": 12, "K": 22, "L": 22, "M": 14, "N": 18, "O": 42, "P": 12,
        "Q": 12, "R": 24, "S": 16, "T": 36, "U": 40, "V": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Accounts ────────────────────────────────────────────────────
    wsc = wb.create_sheet("Accounts _ Config")
    wsc["A1"] = "Required test accounts (Ready? before the run)"
    wsc["A1"].font = TITLE_FONT
    wsc.merge_cells("A1:D1")
    for col, h in enumerate(["Account", "Purpose", "Ready? (Yes/No)", "Notes"], 1):
        cell = wsc.cell(3, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    accounts = [
        ("New mailbox (tester-owned)", "FP-003/004/005 signup, verify, reset", "", "Must receive email"),
        ("qa.free", "FP-011 then FP-010 Checkout", "", "Do not upgrade until FP-011 is done"),
        ("qa.pro", "Sultana product cases", "", "Main account"),
        ("qa.max", "Only if Pro credits insufficient", "", "Optional"),
        ("qa.admin", "FP-008 flags, FP-014 admin", "", "MFA recommended"),
        ("qa.disposable", "FP-012 deletion", "", "Will be destroyed"),
        ("qa.user-a / qa.user-b", "FP-013 RLS", "", "Each needs own session/doc"),
        ("qa.banned / qa.pastdue", "FP-008", "", "Seed if missing"),
        ("India-region user", "FP-023/024", "", "profile.region=IN or Kolkata TZ"),
        ("Non-India user", "FP-023", "", "region=US — must not get APPSC"),
    ]
    for i, row in enumerate(accounts, 4):
        for c, val in enumerate(row, 1):
            cell = wsc.cell(i, c, val)
            cell.border = THIN
            cell.alignment = WRAP
    dv_yes = DataValidation(type="list", formula1='"Yes,No,Blocked"', allow_blank=True)
    wsc.add_data_validation(dv_yes)
    dv_yes.add("C4:C13")

    wsc["A15"] = "Required configuration (ops — Blocked if No)"
    wsc["A15"].font = TITLE_FONT
    for col, h in enumerate(["Item", "Needed for", "Ready? (Yes/No)", "Notes"], 1):
        cell = wsc.cell(16, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    config = [
        ("Production URL is not localhost", "FP-001, FP-004, FP-005", "", "Auth Site URL + redirect allowlist"),
        ("Resend / send-email secret", "FP-003, FP-004, FP-005", "", ""),
        ("Stripe TEST keys + price IDs + webhook", "FP-010", "", "Never use a live card"),
        ("OAuth Google (and GitHub if shown)", "FP-007", "", "Redirect allowlist includes production"),
        ("Deepgram", "FP-017, FP-019", "", ""),
        ("AI provider (Gemini/OpenAI/Anthropic)", "FP-017, FP-020, FP-021, FP-022", "", ""),
        ("Migration 20260813100000 applied", "FP-012 gap/deletion tables", "", "account_deletion_operations, gap_analyses"),
        ("Edge Functions redeployed", "All AI/billing/delete/gap/gov", "", "See RUNBOOK Aug 13 remote apply"),
        ("Sample PDF resume < 10 MB + JD < 5 MB", "FP-022", "", ""),
        ("Windows Electron installer", "FP-028", "", "Mark N/A if unpublished"),
        ("Frontend host deployed this commit", "Entire pack", "", "Code-only local verify is not enough"),
    ]
    for i, row in enumerate(config, 17):
        for c, val in enumerate(row, 1):
            cell = wsc.cell(i, c, val)
            cell.border = THIN
            cell.alignment = WRAP
            if c == 1:
                cell.fill = WARN
    dv_yes2 = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    wsc.add_data_validation(dv_yes2)
    dv_yes2.add("C17:C27")
    for col, w in enumerate([42, 40, 16, 48], 1):
        wsc.column_dimensions[get_column_letter(col)].width = w

    # ── Fail Log ────────────────────────────────────────────────────
    wsf = wb.create_sheet("Fail Log")
    fail_h = [
        "ID", "Owner", "Section", "URL", "Steps (what you did)", "Expected",
        "Actual", "Screenshot", "Severity", "Defect ID", "Blocked by (ops)",
    ]
    for col, h in enumerate(fail_h, 1):
        cell = wsf.cell(1, col, h)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = HEADER_FONT
    for r in range(2, 42):
        for c in range(1, 12):
            wsf.cell(r, c, "").border = THIN
    sev = DataValidation(type="list", formula1='"P0,P1"', allow_blank=True)
    wsf.add_data_validation(sev)
    sev.add("I2:I41")
    wsf.freeze_panes = "A2"
    for col, w in enumerate([10, 12, 22, 40, 30, 28, 28, 20, 12, 14, 28], 1):
        wsf.column_dimensions[get_column_letter(col)].width = w

    # ── Sign-off ────────────────────────────────────────────────────
    wss = wb.create_sheet("Sign-off")
    wss["A1"] = "Final production QA sign-off"
    wss["A1"].font = TITLE_FONT
    wss.merge_cells("A1:C1")
    sign = [
        ("Site tested", SITE),
        ("Workbook", OUT.name),
        ("Total items", len(CASES)),
        ("Venkata cases", len(venkat)),
        ("Sultana cases", len(sultana)),
        ("P0 count", sum(1 for c in CASES if c["priority"] == "P0")),
        ("P1 count", sum(1 for c in CASES if c["priority"] == "P1")),
        ("", ""),
        ("Passed (COUNTIF Checklist!J:J Pass)", f'=COUNTIF(Checklist!J:J,"Pass")'),
        ("Failed", f'=COUNTIF(Checklist!J:J,"Fail")'),
        ("Blocked", f'=COUNTIF(Checklist!J:J,"Blocked")'),
        ("N/A", f'=COUNTIF(Checklist!J:J,"N/A")'),
        ("Not Run", f'=COUNTIF(Checklist!J:J,"Not Run")'),
        ("", ""),
        ("Release recommendation", "Go / Conditional Go closed beta / No-go"),
        ("Known blockers (from plan)", "Stripe live/test secrets, Resend, Auth Site URL, EF redeploy, FE host, MFA, gov bank certification (0 full-sim packs)"),
        ("Venkata name / date", ""),
        ("Sultana name / date", ""),
        ("Reviewer name / date", ""),
        ("Recommendation notes", "Do not Go if any P0 is Fail. Blocked P0 keeps CONDITIONAL_GO / NO_GO."),
    ]
    for i, (k, v) in enumerate(sign, start=3):
        wss.cell(i, 1, k).font = LABEL
        cell = wss.cell(i, 2, v)
        cell.border = THIN
        cell.alignment = WRAP
    wss.column_dimensions["A"].width = 42
    wss.column_dimensions["B"].width = 72
    wss.row_dimensions[18].height = 36
    wss.row_dimensions[22].height = 36

    # pie data
    wss["D3"] = "Status"
    wss["E3"] = "Count"
    wss["D3"].font = HEADER_FONT
    wss["E3"].font = HEADER_FONT
    wss["D3"].fill = HEADER_FILL
    wss["E3"].fill = HEADER_FILL
    wss["D4"] = "Pass"
    wss["E4"] = f'=COUNTIF(Checklist!J:J,"Pass")'
    wss["D5"] = "Fail"
    wss["E5"] = f'=COUNTIF(Checklist!J:J,"Fail")'
    wss["D6"] = "Blocked"
    wss["E6"] = f'=COUNTIF(Checklist!J:J,"Blocked")'
    wss["D7"] = "Not Run"
    wss["E7"] = f'=COUNTIF(Checklist!J:J,"Not Run")'
    pie = PieChart()
    pie.title = "Final pack status"
    labels = Reference(wss, min_col=4, min_row=4, max_row=7)
    data = Reference(wss, min_col=5, min_row=3, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 12
    pie.height = 8
    wss.add_chart(pie, "D9")

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Cases: {len(CASES)}  Venkata={len(venkat)}  Sultana={len(sultana)}")
    print(
        "P0",
        sum(1 for c in CASES if c["priority"] == "P0"),
        "P1",
        sum(1 for c in CASES if c["priority"] == "P1"),
    )


if __name__ == "__main__":
    main()
