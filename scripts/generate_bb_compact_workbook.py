"""
Clarify AI — compact black-box QA workbook (same layout as
Clarify_AI_BB_Manual_QA_Workbook (5).xlsx).

Sheets:
  00 Cover & Instructions
  00b Section Completion Gate
  01 Application Inventory
  02 User Roles & Permissions
  03 Test Accounts
  05 All Test Cases          ← every case, peach cols next to ID, filter by Tester
  39 Defect Log
  Anushka-Extra bug
  Sultana-Extra bug
  Venkat-Extra bug

Team: Anushka, Sultana, Venkat. Window: complete within 2 days.
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from blackbox_qa.common import (  # noqa: E402
    ALT_F,
    BODY_FONT,
    CENTER,
    FAIL_F,
    HEADER_FILL,
    HEADER_FONT,
    P0_F,
    P1_F,
    PASS_F,
    SECTION_FONT,
    SITE,
    SKIP_F,
    SUB_FONT,
    THIN,
    TITLE_FONT,
    WARN_F,
    WRAP,
    write_title_block,
)
from blackbox_qa.staffing import (  # noqa: E402
    MODULE_OWNER,
    TESTER_FOCUS,
    TESTER_LIST,
    TESTERS,
    WINDOW_LABEL,
    SECTION_STATUS_LIST,
    assign_tester,
    load_qa_env,
)
from blackbox_qa.cases_a import (  # noqa: E402
    auth_cases,
    dashboard_cases,
    module_index_cases,
    onboarding_cases,
    public_cases,
)
from blackbox_qa.cases_b import (  # noqa: E402
    ai_coach_cases,
    documents_cases,
    gov_exam_cases,
    gov_exam_live_cases,
    live_copilot_cases,
    mock_interview_cases,
    practice_coach_cases,
    prep_lab_cases,
    resume_jd_cases,
)
from blackbox_qa.cases_c import (  # noqa: E402
    accessibility_cases,
    admin_cases,
    ai_fallback_cases,
    analytics_cases,
    answer_bank_cases,
    api_network_cases,
    billing_cases,
    coding_lab_cases,
    community_cases,
    credits_cases,
    integrations_cases,
    journey_cases,
    learning_cases,
    notifications_cases,
    regression_cases,
    reports_cases,
    responsive_cases,
    scheduler_cases,
    security_cases,
    sessions_cases,
    settings_cases,
)
from blackbox_qa.retest_wave import DEFECT_SEEDS  # noqa: E402

import generate_blackbox_manual_qa_workbook as gen  # noqa: E402

OUT = ROOT / "Clarify_AI_BB_Manual_QA_Workbook.xlsx"
PEACH = PatternFill("solid", fgColor="FFF7ED")
PEACH_HDR = PatternFill("solid", fgColor="C2410C")
BLOCK_F = PatternFill("solid", fgColor="FDE047")
NA_F = PatternFill("solid", fgColor="E2E8F0")

ALL_HEADERS = [
    "Test Case ID",
    "Actual Result",
    "Pass / Fail",
    "Defect ID",
    "Execution Date",
    "Notes",
    "Module",
    "Sub-Module",
    "Feature",
    "Priority",
    "Severity",
    "Test Type",
    "User Type",
    "Role",
    "Account ID",
    "Tester",
    "Preconditions",
    "Test Data",
    "Exact Steps",
    "Expected Result (per step)",
    "Final Expected Result",
    "Screenshot / Evidence Requirement",
]

COL_WIDTHS = [16, 30, 12, 14, 16, 26, 20, 18, 28, 12, 13, 14, 13, 12, 18, 14, 36, 32, 48, 44, 36, 30]

CASE_SOURCES = [
    module_index_cases,
    public_cases,
    auth_cases,
    onboarding_cases,
    dashboard_cases,
    practice_coach_cases,
    live_copilot_cases,
    mock_interview_cases,
    gov_exam_cases,
    gov_exam_live_cases,
    ai_coach_cases,
    prep_lab_cases,
    documents_cases,
    resume_jd_cases,
    answer_bank_cases,
    scheduler_cases,
    sessions_cases,
    reports_cases,
    analytics_cases,
    billing_cases,
    credits_cases,
    settings_cases,
    notifications_cases,
    integrations_cases,
    learning_cases,
    community_cases,
    coding_lab_cases,
    admin_cases,
    security_cases,
    accessibility_cases,
    responsive_cases,
    api_network_cases,
    ai_fallback_cases,
    regression_cases,
    journey_cases,
]

GATE_MODULES = [
    "Module Smoke",
    "Public Pages",
    "Authentication",
    "Onboarding",
    "Dashboard",
    "Practice Coach",
    "Live Copilot",
    "Mock Interview",
    "Government Exams",
    "AI Coach / Chatbot",
    "Prep Lab",
    "Documents",
    "Resume / JD Parsing",
    "Answer Bank",
    "Interview Scheduler",
    "Sessions",
    "Reports",
    "Analytics",
    "Billing",
    "Credits",
    "Settings",
    "Notifications",
    "Integrations",
    "Learning Hub",
    "Community",
    "Coding Lab",
    "Admin Portal",
    "Security",
    "Accessibility",
    "Responsive / Cross-Browser",
    "API / Network Observation",
    "AI / Fallback",
    "Regression",
    "Cross-Module Journeys",
]

EXTRA_BUG_HEADERS = [
    "Defect ID",
    "Title",
    "Module",
    "Test Case ID",
    "Severity (P0-P4)",
    "Business Impact",
    "Reproducibility",
    "Environment",
    "Account ID",
    "Steps to Reproduce",
    "Expected",
    "Actual",
    "Evidence Links",
    "Status",
    "Date",
]


def collect_cases() -> list[dict]:
    rows: list[dict] = []
    for fn in CASE_SOURCES:
        for case in fn():
            rows.append(assign_tester(dict(case)))
    return rows


def sheet_cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "00 Cover & Instructions"
    start = write_title_block(
        ws,
        "Clarify AI — Black-Box Manual QA Workbook",
        "CREDENTIALS PRE-FILLED | TESTERS PRE-ASSIGNED | ALL TEST CASES IN ONE SHEET | "
        "COMPLETE WITHIN 2 DAYS | UPDATE PEACH COLUMNS ONLY",
        cols=6,
    )
    lines = [
        "⚠ CONFIDENTIAL — This workbook contains QA test account passwords for closed-beta testers only. Do not forward outside the QA team. Do not commit to a public repo.",
        "",
        "CYCLE",
        WINDOW_LABEL,
        f"APP URL: {SITE}",
        "Engineering status: NO_GO / RELEASE BLOCKED until one Custom or Full Mock paper completes on the live URL.",
        "",
        "TEAM",
        "1. Anushka — Government Exams + live proof (TC-GOV-LIVE-*), Practice Coach, Live Overlay, Mock Interview, Documents, Resume/JD, AI Chat, Module smoke, Journey 2, Admin gov (TC-ADM-019–024), AI Fallback",
        "2. Sultana — Billing, Credits, Dashboard, Prep Lab, Sessions, Reports, Analytics, Answer Bank, Learning/Community/Coding, Admin (non-gov), API, Regression, Journeys 1 and 4–5",
        "3. Venkat — Authentication, Onboarding, Security, Public pages, Settings, Integrations, Scheduler, Accessibility, Responsive / Cross-Browser",
        "",
        "WHAT IS ALREADY FILLED FOR YOU",
        "- Test Accounts (email + password from .env.qa.local)",
        "- Every test case: steps, expected results, Account ID, Tester, Priority",
        "",
        "HOW TO USE '05 All Test Cases'",
        "All module tabs are merged into one sheet. Use the filter arrows on the header row to show just your name in Tester (or one Module). The 5 peach columns sit right next to Test Case ID. Header rows stay frozen as you scroll.",
        "Three columns that never varied (Workflow, Test Environment, Validation Points) were removed — defaults are in the note at the top of that sheet.",
        "",
        "WHAT YOU UPDATE (peach / orange columns only)",
        "- Actual Result",
        "- Pass / Fail (dropdown)",
        "- Defect ID (if Fail)",
        "- Execution Date",
        "- Notes (optional)",
        "",
        "SECTION COMPLETION GATE (sheet 00b)",
        "1. Filter 05 All Test Cases to your Tester. Execute P0/P1 first.",
        "2. Ready? becomes YES only when every case for that module has Pass/Fail ≠ Not Run AND Actual Result is filled.",
        "3. ONLY then set Section Status = ALL FILLED — READY TO CLOSE.",
        "",
        "BLOCKER CODES (use on Blocked rows, not Fail)",
        "BLK-MIG = migrations not on live DB | BLK-EDGE = Edge 404 | BLK-PY = Python worker not shipped | BLK-CFG = OAuth/Calendar/Resend/payments | BLK-BANK = Official/PYQ empty | BLK-CRED = missing fixture",
        "",
        "PASS RULE: Opening a page/button alone is NOT Pass.",
        "DEAD CTA RULE: A visible broken button is Fail. Hidden or labeled Not configured is Pass.",
        "EXTRA BUGS: log new defects on your Extra bug sheet AND on 39 Defect Log.",
    ]
    headings = {
        "CYCLE",
        "TEAM",
        "WHAT IS ALREADY FILLED FOR YOU",
        "HOW TO USE '05 All Test Cases'",
        "WHAT YOU UPDATE (peach / orange columns only)",
        "SECTION COMPLETION GATE (sheet 00b)",
        "BLOCKER CODES (use on Blocked rows, not Fail)",
    }
    r = start
    for line in lines:
        cell = ws.cell(row=r, column=1, value=line)
        cell.alignment = WRAP
        if line.startswith("⚠"):
            cell.fill = FAIL_F
            cell.font = Font(bold=True, color="7F1D1D")
        elif line in headings:
            cell.font = SECTION_FONT
        else:
            cell.font = BODY_FONT
        ws.row_dimensions[r].height = 32 if len(line) > 120 else (18 if line else 10)
        r += 1
    ws.column_dimensions["A"].width = 130
    ws.sheet_properties.tabColor = "0F766E"


def sheet_all_cases(wb: Workbook, cases: list[dict]) -> tuple[int, int]:
    ws = wb.create_sheet("05 All Test Cases")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=22)
    title = ws.cell(row=1, column=1, value=" ")
    title.font = Font(name="Calibri", bold=True, size=11, color="0F172A")
    ws.row_dimensions[1].height = 24
    subtitle = (
        "PRE-FILLED for you. Only the peach columns (Actual Result | Pass/Fail | Defect ID | "
        "Execution Date | Notes) are yours to fill in. Every module's test cases are here in one "
        "sheet — use the filter arrows on row 4 to show just your module (Module column) or just "
        "your name (Tester column). Mark section ALL FILLED on '00b Section Completion Gate' ONLY "
        "when every case for that module has Pass/Fail ≠ Not Run AND Actual Result is filled. "
        "Black-box only — no source/DB/server access. Environment: Closed Beta / QA Target. "
        f"{WINDOW_LABEL}. "
        "Default validation: observable UI state matches expected results; no unexplained spinner; "
        "no silent failure; persistence verified via refresh where applicable. Default evidence: "
        "screenshot of final state (on fail, also capture URL, visible error, and Network status "
        "if DevTools allowed)."
    )
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = SUB_FONT
    cell.alignment = WRAP
    ws.row_dimensions[2].height = 68

    header_row = 4
    for c, h in enumerate(ALL_HEADERS, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = PEACH_HDR if c in (2, 3, 4, 5, 6) else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = "A5"

    first = header_row + 1
    peach_idx = {2, 3, 4, 5, 6}
    for i, case in enumerate(cases):
        r = first + i
        vals = [case.get(h, "") for h in ALL_HEADERS]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c in peach_idx:
                cell.fill = PEACH
            elif ALL_HEADERS[c - 1] == "Priority":
                if v == "P0":
                    cell.fill = P0_F
                elif v == "P1":
                    cell.fill = P1_F
            elif r % 2 == 0:
                cell.fill = ALT_F
        ws.row_dimensions[r].height = 90
    last = first + len(cases) - 1
    ws.auto_filter.ref = f"A{header_row}:V{last}"

    dv_pf = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not Run,N/A,Partial"', allow_blank=True)
    dv_pri = DataValidation(type="list", formula1='"P0,P1,P2,P3,P4"', allow_blank=True)
    dv_t = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_pf)
    ws.add_data_validation(dv_pri)
    ws.add_data_validation(dv_t)
    dv_pf.add(f"C{first}:C{last}")
    dv_pri.add(f"J{first}:J{last}")
    dv_t.add(f"P{first}:P{last}")

    rng = f"C{first}:C{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Run"'], fill=SKIP_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N/A"'], fill=NA_F))

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    return first, last


def sheet_gate(wb: Workbook, first: int, last: int, cases: list[dict]) -> None:
    ws = wb.create_sheet("00b Section Completion Gate", 1)
    start = write_title_block(
        ws,
        "00b — Section Completion Gate",
        "RULE: Set Section Status to 'ALL FILLED — READY TO CLOSE' ONLY when Ready? = YES. "
        "Ready? becomes YES when (a) every test case for that module has Pass/Fail not 'Not Run' and "
        "(b) every Actual Result cell for that module is non-empty. All test cases live in "
        "'05 All Test Cases' — these counts filter that sheet by the Module column. Do not close early.",
        cols=10,
    )
    headers = [
        "Module",
        "Assigned Tester",
        "Total Cases",
        "Not Run (count)",
        "Blank Actual (count)",
        "Ready?",
        "Section Status (select)",
        "Closed By",
        "Closed Date",
        "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    ws.freeze_panes = f"A{start + 1}"
    dv_status = DataValidation(type="list", formula1=SECTION_STATUS_LIST, allow_blank=True)
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_tester)

    ref = "05 All Test Cases"
    for i, module in enumerate(GATE_MODULES):
        r = start + 1 + i
        owner = MODULE_OWNER.get(module, "Venkat")
        if module == "Admin Portal":
            owner = "Sultana"
        ws.cell(row=r, column=1, value=module).border = THIN
        cell_t = ws.cell(row=r, column=2, value=owner)
        cell_t.border = THIN
        dv_tester.add(cell_t)
        g = f"'{ref}'!$G${first}:$G${last}"
        ccol = f"'{ref}'!$C${first}:$C${last}"
        bcol = f"'{ref}'!$B${first}:$B${last}"
        ws.cell(row=r, column=3, value=f'=COUNTIF({g},A{r})').border = THIN
        ws.cell(row=r, column=4, value=f'=COUNTIFS({g},A{r},{ccol},"Not Run")').border = THIN
        ws.cell(row=r, column=5, value=f'=COUNTIFS({g},A{r},{bcol},"")').border = THIN
        cell_ready = ws.cell(row=r, column=6, value=f'=IF(AND(D{r}=0,E{r}=0),"YES","NO")')
        cell_ready.border = THIN
        cell_ready.alignment = CENTER
        cell_st = ws.cell(row=r, column=7, value="Not Started")
        cell_st.border = THIN
        cell_st.fill = PEACH
        dv_status.add(cell_st)
        for c in (8, 9, 10):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN
            cell.fill = PEACH
        ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal", formula=['"YES"'], fill=PASS_F))
        ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal", formula=['"NO"'], fill=FAIL_F))

    last_data = start + len(GATE_MODULES)
    r = last_data + 2
    ws.cell(row=r, column=1, value="TEAM ROLL-UP").font = SECTION_FONT
    ws.cell(row=r + 1, column=1, value="Sections Ready (YES count)").border = THIN
    ws.cell(row=r + 1, column=2, value=f'=COUNTIF(F{start + 1}:F{last_data},"YES")').border = THIN
    for i, w in enumerate([28, 16, 12, 16, 18, 10, 28, 14, 12, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = "DC2626"


def sheet_defect_log(wb: Workbook) -> None:
    ws = wb.create_sheet("39 Defect Log")
    start = write_title_block(
        ws,
        "39 — Defect Log",
        "Every Fail must have a defect row with evidence. No passwords/secrets in attachments. "
        "Open items from the prior cycle are pre-seeded — retest, do not mark Pass without evidence.",
        cols=15,
    )
    headers = [
        "Defect ID",
        "Title",
        "Module",
        "Test Case ID",
        "Severity (P0-P4)",
        "Assigned Tester",
        "Business Impact",
        "Reproducibility",
        "Environment",
        "Account ID",
        "Steps to Reproduce",
        "Expected",
        "Actual",
        "Evidence Links",
        "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    ws.freeze_panes = f"A{start + 1}"
    seeds = list(DEFECT_SEEDS)
    total = max(40, len(seeds))
    dv_sev = DataValidation(type="list", formula1='"P0,P1,P2,P3,P4"', allow_blank=True)
    dv_st = DataValidation(
        type="list",
        formula1='"New,Open,In Progress,Fixed,Verified,Wont Fix,Duplicate"',
        allow_blank=True,
    )
    dv_t = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_sev)
    ws.add_data_validation(dv_st)
    ws.add_data_validation(dv_t)
    for i in range(total):
        r = start + 1 + i
        seed = seeds[i] if i < len(seeds) else {}
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=seed.get(h, ""))
            cell.border = THIN
            cell.alignment = WRAP
            if h in {"Title", "Steps to Reproduce", "Expected", "Actual", "Status", "Assigned Tester"}:
                cell.fill = PEACH
            elif i % 2 == 0:
                cell.fill = ALT_F
        if seed:
            ws.row_dimensions[r].height = 48
    last = start + total
    dv_sev.add(f"E{start + 1}:E{last}")
    dv_t.add(f"F{start + 1}:F{last}")
    dv_st.add(f"O{start + 1}:O{last}")
    for i, w in enumerate([12, 36, 18, 18, 12, 14, 16, 14, 16, 14, 40, 28, 28, 20, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_extra_bugs(wb: Workbook, tester: str, rows: list[dict]) -> None:
    title = f"{tester}-Extra bug"
    ws = wb.create_sheet(title[:31])
    for c, h in enumerate(EXTRA_BUG_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER
    ws.freeze_panes = "A2"
    dv_sev = DataValidation(type="list", formula1='"P0,P1,P2,P3,P4"', allow_blank=True)
    dv_st = DataValidation(
        type="list",
        formula1='"New,Open,In Progress,Fixed,Verified,Wont Fix,Duplicate"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_sev)
    ws.add_data_validation(dv_st)
    n = max(20, len(rows))
    for i in range(n):
        r = 2 + i
        seed = rows[i] if i < len(rows) else {}
        for c, h in enumerate(EXTRA_BUG_HEADERS, 1):
            val = seed.get(h, "")
            if h == "Date" and not val:
                val = seed.get("Date", "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN
            cell.alignment = WRAP
            if h in {"Title", "Steps to Reproduce", "Expected", "Actual", "Status"}:
                cell.fill = PEACH
            elif i % 2 == 0:
                cell.fill = ALT_F
        if seed:
            ws.row_dimensions[r].height = 56
    last = 1 + n
    dv_sev.add(f"E2:E{last}")
    dv_st.add(f"N2:N{last}")
    for i, w in enumerate([12, 40, 22, 16, 12, 16, 14, 18, 14, 40, 28, 28, 16, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:O{last}"
    ws.sheet_properties.tabColor = "C2410C"


def extra_bugs_for(tester: str) -> list[dict]:
    rows = []
    for d in DEFECT_SEEDS:
        if d.get("Assigned Tester") == tester:
            rows.append({
                "Defect ID": d.get("Defect ID", ""),
                "Title": d.get("Title", ""),
                "Module": d.get("Module", ""),
                "Test Case ID": d.get("Test Case ID", ""),
                "Severity (P0-P4)": d.get("Severity (P0-P4)", ""),
                "Business Impact": d.get("Business Impact", ""),
                "Reproducibility": d.get("Reproducibility", "Always"),
                "Environment": d.get("Environment", "Closed beta Chrome"),
                "Account ID": d.get("Account ID", ""),
                "Steps to Reproduce": d.get("Steps to Reproduce", ""),
                "Expected": d.get("Expected", ""),
                "Actual": d.get("Actual", ""),
                "Evidence Links": "",
                "Status": d.get("Status", "Open"),
                "Date": "2026-08-31",
            })
    return rows


def build() -> Path:
    gen.QA_ENV = load_qa_env(ROOT)
    if not gen.QA_ENV.get("QA_FREE_EMAIL"):
        print("WARNING: .env.qa.local missing or incomplete — run npm run qa:seed-accounts")

    cases = collect_cases()
    wb = Workbook()
    sheet_cover(wb)
    first, last = sheet_all_cases(wb, cases)
    sheet_gate(wb, first, last, cases)
    gen.sheet_inventory(wb)
    gen.sheet_roles(wb)
    gen.sheet_accounts(wb)
    sheet_defect_log(wb)
    for tester in TESTERS:
        sheet_extra_bugs(wb, tester, extra_bugs_for(tester))

    # Order: Cover, Gate, Inventory, Roles, Accounts, All TCs, Defect, extras
    def _move(name: str, idx: int) -> None:
        cur = wb.sheetnames.index(name)
        if cur != idx:
            wb.move_sheet(name, offset=idx - cur)

    _move("00b Section Completion Gate", 1)
    _move("01 Application Inventory", 2)
    _move("02 User Roles & Permissions", 3)
    _move("03 Test Accounts", 4)
    _move("05 All Test Cases", 5)

    saved = []
    for dest in (OUT, ROOT / "Clarify_AI_BB_Manual_QA_Workbook_ASSIGNED.xlsx"):
        try:
            wb.save(dest)
            saved.append(dest)
        except PermissionError:
            print(f"NOTE: {dest.name} locked; skipped")
    if not saved:
        fallback = ROOT / "Clarify_AI_BB_Manual_QA_Workbook_2026-08-31.xlsx"
        wb.save(fallback)
        saved.append(fallback)
        print(f"NOTE: saved fallback {fallback.name}")
    return saved[0]


if __name__ == "__main__":
    path = build()
    wb = load_workbook(path, read_only=True, data_only=False)
    print(f"Wrote: {path}")
    print(f"Sheets ({len(wb.sheetnames)}):")
    for name in wb.sheetnames:
        print(f"  - {name}")
    ws = wb["05 All Test Cases"]
    by = Counter()
    p0 = Counter()
    p1 = Counter()
    n = 0
    for row in ws.iter_rows(min_row=5, max_col=16, values_only=True):
        if isinstance(row[0], str) and row[0].startswith("TC-"):
            n += 1
            t = row[15] or ""
            by[str(t)] += 1
            pri = row[9]
            if pri == "P0":
                p0[str(t)] += 1
            if pri == "P1":
                p1[str(t)] += 1
    print(f"Total cases: {n}")
    print("Assignment:")
    for t in TESTERS:
        print(f"  {t}: {by.get(t, 0)}  (P0={p0.get(t, 0)} P1={p1.get(t, 0)})")
    unexpected = [t for t in by if t not in TESTERS]
    print("Unexpected testers:", unexpected or "none")
    blob = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(max_col=4, values_only=True):
            for c in row:
                if isinstance(c, str) and "Raj Balani" in c:
                    blob.append(name)
    print("Raj Balani sheets:", blob or "none")
    wb.close()
