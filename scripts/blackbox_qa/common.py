"""Shared styles and helpers for the black-box manual QA workbook."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SITE = "https://trycareerpilot.com"
LOCAL = "http://localhost:5173"
ENV_NAME = "Closed Beta / QA Target"

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="0F172A")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="0F766E")
SUB_FONT = Font(name="Calibri", size=10, color="64748B", italic=True)
BODY_FONT = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

PASS_F = PatternFill("solid", fgColor="86EFAC")
FAIL_F = PatternFill("solid", fgColor="FCA5A5")
BLOCK_F = PatternFill("solid", fgColor="FDE047")
SKIP_F = PatternFill("solid", fgColor="CBD5E1")
NA_F = PatternFill("solid", fgColor="E2E8F0")
P0_F = PatternFill("solid", fgColor="FECACA")
P1_F = PatternFill("solid", fgColor="FED7AA")
P2_F = PatternFill("solid", fgColor="FEF08A")
P3_F = PatternFill("solid", fgColor="BFDBFE")
P4_F = PatternFill("solid", fgColor="E2E8F0")
ALT_F = PatternFill("solid", fgColor="F8FAFC")
WARN_F = PatternFill("solid", fgColor="FEF3C7")
OK_F = PatternFill("solid", fgColor="D1FAE5")
CARD_F = PatternFill("solid", fgColor="F1F5F9")

# Standard test-case columns (exact order required by workbook design)
TC_HEADERS = [
    "Test Case ID",
    "Module",
    "Sub-Module",
    "Feature",
    "Workflow",
    "Priority",
    "Severity",
    "Test Type",
    "User Type",
    "Role",
    "Account ID",
    "Preconditions",
    "Test Environment",
    "Test Data",
    "Exact Steps",
    "Expected Result (per step)",
    "Final Expected Result",
    "Validation Points",
    "Screenshot / Evidence Requirement",
    "Actual Result",
    "Pass / Fail",
    "Defect ID",
    "Tester",
    "Execution Date",
    "Notes",
]

PASS_FAIL_LIST = '"Pass,Fail,Blocked,Not Run,N/A,Partial"'
PRIORITY_LIST = '"P0,P1,P2,P3,P4"'
SEVERITY_LIST = '"Critical,Major,Minor,Cosmetic"'
TEST_TYPE_LIST = (
    '"Positive,Negative,Boundary,Security,Accessibility,'
    'Responsive,Cross-Browser,Regression,E2E Journey,API Observation,Persistence"'
)
STATUS_CLASS_LIST = (
    '"Fully Working,Partially Working,Requires Configuration,Unavailable,'
    'Coming Soon,Retired,UI-only / Non-functional,Blocked by Environment,'
    'Not Accessible for This Role,PENDING MANUAL CLASSIFICATION"'
)


def tc(
    tid: str,
    module: str,
    feature: str,
    steps: str,
    expected: str,
    final: str,
    *,
    sub: str = "",
    workflow: str = "",
    priority: str = "P1",
    severity: str = "Major",
    test_type: str = "Positive",
    user_type: str = "Authenticated",
    role: str = "user",
    account: str = "FREE_USER_01",
    pre: str = "Logged in. App URL available. Browser DevTools permitted.",
    env: str = ENV_NAME,
    data: str = "Use approved test credentials from Test Accounts sheet.",
    validation: str = "",
    evidence: str = "Screenshot of final state. On fail: URL, visible error, Network status if DevTools allowed.",
    notes: str = "",
) -> dict:
    """Build one atomic black-box test case dict."""
    if not validation:
        validation = (
            "Observable UI state matches expected results; no unexplained spinner; "
            "no silent failure; persistence verified via refresh where applicable."
        )
    return {
        "Test Case ID": tid,
        "Module": module,
        "Sub-Module": sub or module,
        "Feature": feature,
        "Workflow": workflow or feature,
        "Priority": priority,
        "Severity": severity,
        "Test Type": test_type,
        "User Type": user_type,
        "Role": role,
        "Account ID": account,
        "Preconditions": pre,
        "Test Environment": env,
        "Test Data": data,
        "Exact Steps": steps,
        "Expected Result (per step)": expected,
        "Final Expected Result": final,
        "Validation Points": validation,
        "Screenshot / Evidence Requirement": evidence,
        "Actual Result": "",
        "Pass / Fail": "Not Run",
        "Defect ID": "",
        "Tester": "",
        "Execution Date": "",
        "Notes": notes,
    }


def style_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def autosize(ws, headers: list[str], min_w: int = 12, max_w: int = 48) -> None:
    widths = {
        "Test Case ID": 16,
        "Module": 18,
        "Sub-Module": 18,
        "Feature": 28,
        "Workflow": 22,
        "Priority": 10,
        "Severity": 12,
        "Test Type": 14,
        "User Type": 14,
        "Role": 12,
        "Account ID": 18,
        "Preconditions": 36,
        "Test Environment": 18,
        "Test Data": 32,
        "Exact Steps": 48,
        "Expected Result (per step)": 48,
        "Final Expected Result": 36,
        "Validation Points": 32,
        "Screenshot / Evidence Requirement": 32,
        "Actual Result": 28,
        "Pass / Fail": 12,
        "Defect ID": 14,
        "Tester": 14,
        "Execution Date": 14,
        "Notes": 24,
        "Feature / Surface": 36,
        "Route / URL": 40,
        "Classification": 28,
        "Observable Entry Point": 36,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(
            max_w, max(min_w, widths.get(h, 18))
        )


EDITABLE_HEADERS = {
    "Actual Result",
    "Pass / Fail",
    "Defect ID",
    "Execution Date",
    "Notes",
}
EDIT_FILL = PatternFill("solid", fgColor="FFF7ED")  # peach — tester edits only these
LOCKED_HINT_FILL = PatternFill("solid", fgColor="EEF2FF")


def write_rows(ws, headers: list[str], rows: list[dict], start_row: int = 2) -> int:
    for r_idx, row in enumerate(rows):
        excel_row = start_row + r_idx
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if h in EDITABLE_HEADERS:
                cell.fill = EDIT_FILL
            elif excel_row % 2 == 0:
                cell.fill = ALT_F
            if h == "Priority":
                fills = {"P0": P0_F, "P1": P1_F, "P2": P2_F, "P3": P3_F, "P4": P4_F}
                if val in fills:
                    cell.fill = fills[val]
            if h == "Pass / Fail":
                cell.alignment = CENTER
            if h == "Tester":
                cell.alignment = CENTER
                cell.fill = LOCKED_HINT_FILL
        ws.row_dimensions[excel_row].height = 90
    return start_row + len(rows) - 1


def add_tester_validation(ws, col_letter: str, max_row: int, tester_list: str) -> None:
    dv = DataValidation(type="list", formula1=tester_list, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max(max_row, 2)}")


def add_pass_fail_validation(ws, col_letter: str, max_row: int, start_row: int = 2) -> None:
    dv = DataValidation(type="list", formula1=PASS_FAIL_LIST, allow_blank=True)
    dv.error = "Select Pass, Fail, Blocked, Not Run, N/A, or Partial"
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{max(max_row, start_row)}")


def add_priority_validation(ws, col_letter: str, max_row: int, start_row: int = 2) -> None:
    dv = DataValidation(type="list", formula1=PRIORITY_LIST, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{max(max_row, start_row)}")


def apply_pass_fail_cf(ws, col_letter: str, max_row: int, start_row: int = 2) -> None:
    rng = f"{col_letter}{start_row}:{col_letter}{max(max_row, start_row)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Run"'], fill=SKIP_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N/A"'], fill=NA_F))


def write_title_block(ws, title: str, subtitle: str, cols: int = 8) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = SUB_FONT
    s.alignment = WRAP
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 40
    return 4


def write_kv_table(ws, start_row: int, pairs: list[tuple[str, str]]) -> int:
    ws.cell(row=start_row, column=1, value="Field").font = HEADER_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.cell(row=start_row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=start_row, column=2).fill = HEADER_FILL
    r = start_row + 1
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=1).border = THIN
        ws.cell(row=r, column=1).fill = CARD_F
        cell = ws.cell(row=r, column=2, value=v)
        cell.alignment = WRAP
        cell.border = THIN
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 72
    return r


def write_tc_sheet(
    ws,
    title: str,
    cases: list[dict],
    na_note: str | None = None,
    tester_list: str = '"Anushka,Sultana,Venkat"',
) -> dict:
    """Write a standard test-case sheet. Returns {header_row, first_data, last, count}."""
    subtitle = (
        "PRE-FILLED for you. ONLY update peach columns: Actual Result | Pass/Fail | Defect ID | "
        "Execution Date | Notes. Tester is pre-assigned (Anushka / Sultana / Venkat). "
        "Complete assigned work within 2 days. Mark section ALL FILLED on '00b Section Completion Gate' "
        "ONLY when every case has Pass/Fail ≠ Not Run AND Actual Result filled. "
        "Black-box only — no source/DB/server access. Opening a page is NOT Pass."
    )
    if na_note:
        subtitle = f"NOT APPLICABLE — {na_note}\n{subtitle}"
        ws.sheet_properties.tabColor = "94A3B8"

    start = write_title_block(ws, title, subtitle, cols=len(TC_HEADERS))
    header_row = start
    for col, h in enumerate(TC_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        if h in EDITABLE_HEADERS:
            cell.fill = PatternFill("solid", fgColor="C2410C")
        else:
            cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = f"A{header_row + 1}"
    meta = {"header_row": header_row, "first_data": header_row + 1, "last": header_row, "count": 0}
    if cases:
        last = write_rows(ws, TC_HEADERS, cases, start_row=header_row + 1)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(TC_HEADERS))}{last}"
        add_pass_fail_validation(ws, "U", last, start_row=header_row + 1)
        add_priority_validation(ws, "F", last, start_row=header_row + 1)
        apply_pass_fail_cf(ws, "U", last, start_row=header_row + 1)
        # Tester col W = 23; apply from first_data
        dv = DataValidation(type="list", formula1=tester_list, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"W{header_row + 1}:W{last}")
        meta.update({"last": last, "count": len(cases)})
    else:
        na = tc(
            "N/A-001",
            title,
            "Sheet marked N/A",
            "1. Confirm with QA lead that this module is out of scope for this release.",
            "1. Confirmation recorded.",
            "Sheet remains N/A until scope changes.",
            priority="P4",
            severity="Cosmetic",
            test_type="Positive",
            notes=na_note or "No cases generated.",
        )
        last = write_rows(ws, TC_HEADERS, [na], start_row=header_row + 1)
        add_pass_fail_validation(ws, "U", last, start_row=header_row + 1)
        meta.update({"last": last, "count": 1})
    autosize(ws, TC_HEADERS)
    ws.row_dimensions[header_row].height = 30
    ws.sheet_view.showGridLines = False
    return meta
