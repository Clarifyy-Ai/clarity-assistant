"""
Clarify AI — Black-Box Manual QA Workbook generator

Independent workbook for human manual testers who have ONLY:
  - application URL
  - approved test credentials (stored outside the workbook)
  - approved fixtures
  - browser + permitted DevTools

No source/DB/Supabase/server access required to execute cases.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Clarify_AI_BlackBox_Manual_QA_Workbook.xlsx"
OUT_ALT = ROOT / "Clarify_AI_BlackBox_Manual_QA_Workbook_ASSIGNED.xlsx"
OUT_DATED = ROOT / "Clarify_AI_BlackBox_Manual_QA_Workbook_ASSIGNED_2026-08-31.xlsx"

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from blackbox_qa.common import (  # noqa: E402
    ALT_F,
    BODY_FONT,
    FAIL_F,
    HEADER_FILL,
    HEADER_FONT,
    OK_F,
    PASS_F,
    P0_F,
    P1_F,
    SECTION_FONT,
    SITE,
    STATUS_CLASS_LIST,
    SUB_FONT,
    TC_HEADERS,
    THIN,
    TITLE_FONT,
    WARN_F,
    WRAP,
    CENTER,
    write_kv_table,
    write_tc_sheet,
    write_title_block,
)
from blackbox_qa.inventory_data import (  # noqa: E402
    ACCOUNTS,
    ENV_FIELDS,
    INVENTORY,
    PENDING,
    ROLES,
)
from blackbox_qa.cases_a import (  # noqa: E402
    auth_cases,
    dashboard_cases,
    module_index_cases,
    onboarding_cases,
    public_cases,
)
from blackbox_qa.cases_b import (  # noqa: E402
    ai_coach_cases,
    documents_cases,
    gov_exam_cases,
    gov_exam_live_cases,
    live_copilot_cases,
    mock_interview_cases,
    practice_coach_cases,
    prep_lab_cases,
    resume_jd_cases,
)
from blackbox_qa.cases_c import (  # noqa: E402
    accessibility_cases,
    admin_cases,
    ai_fallback_cases,
    analytics_cases,
    answer_bank_cases,
    api_network_cases,
    billing_cases,
    coding_lab_cases,
    community_cases,
    credits_cases,
    integrations_cases,
    journey_cases,
    learning_cases,
    notifications_cases,
    regression_cases,
    reports_cases,
    responsive_cases,
    scheduler_cases,
    security_cases,
    sessions_cases,
    settings_cases,
)


from openpyxl.worksheet.datavalidation import DataValidation

from blackbox_qa.staffing import (  # noqa: E402
    SHEET_OWNER,
    TESTER_LIST,
    TESTERS,
    SECTION_STATUS_LIST,
    WINDOW_LABEL,
    WINDOW_STATUS_LIST,
    assign_all,
    build_credential_rows,
    execution_window_rows,
    load_qa_env,
    stripe_sandbox_note,
)
from blackbox_qa.retest_wave import DEFECT_SEEDS, RETEST_WAVE  # noqa: E402

# Populated during build() for Section Completion Gate formulas
SHEET_META: dict[str, dict] = {}
QA_ENV: dict[str, str] = {}


def sheet_cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "00 Cover & Instructions"
    start = write_title_block(
        ws,
        "Clarify AI — Black-Box Manual QA Workbook",
        "CYCLE 31 AUG 2026 | COMPLETE WITHIN 2 DAYS | 3 TESTERS PRE-ASSIGNED | UPDATE PEACH COLUMNS ONLY",
        cols=6,
    )
    lines = [
        "⚠ CONFIDENTIAL — This workbook contains QA test account passwords for closed-beta testers only. Do not forward outside the QA team. Do not commit to a public repo.",
        "",
        "CYCLE",
        "Wave 2 — Retest after 30-08 remediation + live Gov exam proof.",
        "Window: complete within 2 days (31 Aug – 1 Sep 2026, IST). Testers pick their own hours.",
        "App: https://clarify.ai.sltfinanceindia.com",
        "",
        "TEAM (3 testers — every case is pre-assigned)",
        "1. Anushka — Government Exams + live proof, AI Fallback, Practice Coach, Live Overlay, Mock Interview, Documents, Resume/JD, AI Chat, Module smoke, Journey 2, Admin gov ingest/review (TC-ADM-019–024)",
        "2. Sultana — Billing, Credits, Dashboard, Prep Lab, Sessions, Reports, Analytics, Answer Bank, Learning/Community/Coding, Admin (non-gov), API, Regression, Journeys 1 and 4–5",
        "3. Venkat — Authentication, Onboarding, Security, Public pages, Settings, Integrations, Scheduler, Accessibility, Responsive / Cross-Browser",
        "",
        "ENGINEERING STATUS (as of 31 Aug 2026) — NOT LIVE-PROVEN",
        "- 30-08 remediation pack = NO_GO (unit/contract only).",
        "- Government exam stack = RELEASE_BLOCKED / PARTIALLY_COMPLETE.",
        "- Migrations not applied; Edge deploy failed; Python/Render paper factory not shipped.",
        "- No live SEARCH → GENERATE → SIT → SUBMIT paper yet.",
        "Execute everything user-visible TODAY. If a case needs undeployed backend, mark Blocked (not Fail) with a blocker code.",
        "",
        "BLOCKER CODES",
        "BLK-MIG = migrations not on live DB",
        "BLK-EDGE = Edge function 404 / not redeployed",
        "BLK-PY = Python/Render worker not shipped",
        "BLK-CFG = OAuth / Calendar / Resend / Stripe-Razorpay config",
        "BLK-BANK = Official/PYQ bank empty (accepted only AFTER one Custom/Full Mock paper exists)",
        "BLK-CRED = fixture account missing / wrong credit balance",
        "",
        "WHAT IS ALREADY FILLED FOR YOU",
        "- Test Accounts (email + password from .env.qa.local)",
        "- Test Environment URL, browsers, resolutions, sandbox card",
        "- Every test case: steps, expected results, Account ID, Tester, Priority",
        "- Sheet 00a: one row per tester, 2-day window (no day-by-day calendar)",
        "- Sheet 00c: remediations to retest",
        "- Sheet 00d: live Gov exam proof (Anushka)",
        "",
        "WHAT YOU UPDATE (peach / orange columns only)",
        "- Actual Result",
        "- Pass / Fail (dropdown)",
        "- Defect ID (if Fail)",
        "- Execution Date",
        "- Notes (optional)",
        "",
        "SECTION COMPLETION GATE (sheet 00b)",
        "1. Execute all cases on your assigned sheets.",
        "2. Watch Ready? column — it becomes YES only when every case has Pass/Fail ≠ Not Run AND Actual Result is filled.",
        "3. ONLY then set Section Status = ALL FILLED — READY TO CLOSE.",
        "4. Do not mark ALL FILLED early — the Ready? formula will still show NO.",
        "",
        "EXECUTION WINDOW (sheet 00a)",
        "Three rows only (one per tester). Status: Not Started / In Progress / Done / Blocked.",
        "P0/P1 first; P2+ if time remains inside the same 2-day window.",
        "",
        f"APP URL: {SITE}",
        "PASS RULE: Opening a page/button alone is NOT Pass.",
        "DEAD CTA RULE: A visible broken button is Fail. Hidden or labeled “Not configured” is Pass.",
        "RELEASE: Stay RELEASE BLOCKED until 00d TC-GOV-LIVE-06 (or Full Mock equivalent) produces a real completed paper on the live URL.",
    ]
    r = start
    headings = {
        "CYCLE",
        "TEAM (3 testers — every case is pre-assigned)",
        "ENGINEERING STATUS (as of 31 Aug 2026) — NOT LIVE-PROVEN",
        "BLOCKER CODES",
        "WHAT IS ALREADY FILLED FOR YOU",
        "WHAT YOU UPDATE (peach / orange columns only)",
        "SECTION COMPLETION GATE (sheet 00b)",
        "EXECUTION WINDOW (sheet 00a)",
    }
    for line in lines:
        cell = ws.cell(row=r, column=1, value=line)
        cell.alignment = WRAP
        if line.startswith("⚠"):
            cell.fill = FAIL_F
            cell.font = Font(bold=True, color="7F1D1D")
        elif line in headings:
            cell.font = SECTION_FONT
        else:
            cell.font = BODY_FONT
        ws.row_dimensions[r].height = 18 if line else 10
        r += 1
    ws.column_dimensions["A"].width = 130
    ws.sheet_properties.tabColor = "0F766E"


def _inventory_tester(area: str) -> str:
    area_l = area.lower()
    if area_l in {"government exams"}:
        return "Anushka"
    if area_l in {"practice coach", "mock interview", "documents"}:
        return "Anushka"
    if area_l in {"admin"}:
        return "Sultana"
    if area_l in {"billing"}:
        return "Sultana"
    if area_l in {"dashboard", "prep lab", "sessions", "reports", "analytics", "answer bank",
                  "learning hub", "community", "coding lab", "practice", "company research", "assessments"}:
        return "Sultana"
    if area_l in {"auth", "onboarding", "security", "public", "settings", "scheduler", "retired", "guide"}:
        return "Venkat"
    return "Venkat"


def sheet_inventory(wb: Workbook) -> None:
    ws = wb.create_sheet("01 Application Inventory")
    start = write_title_block(
        ws,
        "01 — Application Inventory (user-observable)",
        "Classify ONLY after manual execution. Opening a route ≠ Fully Working. "
        f"All rows start as {PENDING}. Tester pre-assigned (Anushka / Sultana / Venkat).",
        cols=10,
    )
    headers = [
        "Area",
        "Feature / Surface",
        "Route / URL",
        "Observable Entry Point",
        "Primary Role",
        "Classification",
        "Evidence / Notes",
        "Tester",
        "Date",
        "Defect ID",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = f"A{start + 1}"
    dv = DataValidation(type="list", formula1=STATUS_CLASS_LIST, allow_blank=True)
    ws.add_data_validation(dv)
    dv_t = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_t)
    for i, (area, feature, route, entry, role, note) in enumerate(INVENTORY):
        r = start + 1 + i
        route_display = f"{SITE}{route}" if route.startswith("/") else route
        vals = [
            area, feature, route_display, entry, role, PENDING, note,
            _inventory_tester(area), "", "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = ALT_F
            if c == 6:
                cell.fill = WARN_F
            if c == 8:
                cell.fill = PatternFill("solid", fgColor="EEF2FF")
        dv.add(ws.cell(row=r, column=6))
        dv_t.add(ws.cell(row=r, column=8))
        ws.row_dimensions[r].height = 36
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{start + len(INVENTORY)}"
    widths = [16, 36, 48, 28, 14, 28, 36, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_roles(wb: Workbook) -> None:
    ws = wb.create_sheet("02 User Roles & Permissions")
    start = write_title_block(
        ws,
        "02 — User Roles & Permissions (black-box)",
        "Expected permissions as observed in the product UI/plan gates — not from source inspection.",
        cols=5,
    )
    headers = ["Role", "Description", "Can Access", "Cannot Access", "Primary Test Focus"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    for i, row in enumerate(ROLES):
        r = start + 1 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        ws.row_dimensions[r].height = 40
    for i, w in enumerate([22, 28, 40, 36, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_accounts(wb: Workbook) -> None:
    ws = wb.create_sheet("03 Test Accounts")
    start = write_title_block(
        ws,
        "03 — Test Accounts (EMAIL + PASSWORD PRE-FILLED)",
        "Loaded from .env.qa.local for closed-beta QA. CONFIDENTIAL — QA team only. "
        "Use Account ID referenced in each test case.",
        cols=12,
    )
    headers = [
        "Test Account ID",
        "Role",
        "Plan",
        "Purpose",
        "Email / Username",
        "Password",
        "Credits (seed)",
        "Credential Location",
        "Environment",
        "Restrictions",
        "Expected Permissions",
        "Assigned Testers (shared)",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER
        if h in ("Email / Username", "Password"):
            cell.fill = PatternFill("solid", fgColor="B91C1C")
    ws.freeze_panes = f"A{start + 1}"
    rows = build_credential_rows(QA_ENV)
    for i, row in enumerate(rows):
        r = start + 1 + i
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.border = THIN
            cell.alignment = WRAP
            if h == "Password":
                cell.fill = WARN_F
            elif h == "Email / Username":
                cell.fill = OK_F
        ws.row_dimensions[r].height = 40
    # Sandbox payment block
    r = start + len(rows) + 3
    ws.cell(row=r, column=1, value="SANDBOX PAYMENT (pre-filled)").font = SECTION_FONT
    for i, (k, v) in enumerate(stripe_sandbox_note(QA_ENV)):
        ws.cell(row=r + 1 + i, column=1, value=k).border = THIN
        ws.cell(row=r + 1 + i, column=2, value=v).border = THIN
        ws.cell(row=r + 1 + i, column=2).fill = WARN_F
    for i, w in enumerate([18, 22, 12, 28, 32, 28, 12, 28, 18, 28, 22, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = "B91C1C"


def sheet_environment(wb: Workbook) -> None:
    ws = wb.create_sheet("04 Test Environment")
    start = write_title_block(
        ws,
        "04 — Test Environment (PRE-FILLED)",
        "Core fields pre-filled. Testers only update Tester name / Test date / Browser version when executing.",
        cols=2,
    )
    filled = list(ENV_FIELDS)
    # Override with concrete values
    overrides = {
        "Web application URL": QA_ENV.get("QA_BASE_URL_PROD", SITE),
        "Admin URL": QA_ENV.get("QA_BASE_URL_PROD", SITE).rstrip("/") + "/app/admin",
        "Local URL (optional)": QA_ENV.get("QA_BASE_URL_LOCAL", "http://localhost:5173"),
        "Test environment name": "Closed Beta / QA Target",
        "Browser (primary)": "Chrome (latest stable)",
        "Browser version": "(tester updates at execution)",
        "Desktop resolution": "1920 × 1080 (also 1440 × 900, 1366 × 768)",
        "Mobile resolution": "360 × 800, 375 × 812, 414 × 896",
        "Tablet resolution": "768 × 1024",
        "Supported browsers": "Chrome, Edge, Firefox, Safari (where available)",
        "Required extensions": "None required. Disable ad blockers for QA.",
        "Required permissions": "Microphone; notifications optional; system/tab audio for Practice Coach",
        "Test email system": "QA disposable inboxes for NEW_USER / verify flows (coordinate with QA lead)",
        "Payment sandbox": "Stripe/Razorpay TEST mode — card on Test Accounts sheet",
        "Stripe test card (sandbox)": QA_ENV.get("QA_STRIPE_TEST_CARD", "4242424242424242"),
        "File upload test folder": "Use approved QA fixtures (PDF/DOCX/TXT/corrupt/oversized) from QA lead share",
        "Supported languages": "English (primary); exam languages as shown in Gov Exam UI",
        "Timezone": "Asia/Kolkata (IST)",
        "Test date": "2026-08-31 (cycle start; complete by 2026-09-01)",
        "Tester name": "Anushka | Sultana | Venkat (see 00a Execution Window + case Tester column)",
        "Credential location": "Sheet 03 Test Accounts (from .env.qa.local)",
        "DevTools policy": "Network + Console for evidence; do not forge auth",
        "Secrets policy": "Do not screenshot full passwords in evidence; redact in defect attachments",
    }
    pairs = [(k, overrides.get(k, v)) for k, v in filled]
    # Ensure all overrides present
    seen = {k for k, _ in pairs}
    for k, v in overrides.items():
        if k not in seen:
            pairs.append((k, v))
    write_kv_table(ws, start, pairs)
    r = start + len(pairs) + 3
    ws.cell(row=r, column=1, value="Tester acknowledgements (tick when starting)").font = SECTION_FONT
    for i, text in enumerate(
        [
            "I will only edit peach columns on test-case sheets.",
            "I will mark ALL FILLED on Section Gate only when Ready?=YES.",
            "I will use sandbox payments only.",
            "I will not share this workbook outside the QA team.",
        ]
    ):
        ws.cell(row=r + 1 + i, column=1, value="☐").border = THIN
        ws.cell(row=r + 1 + i, column=2, value=text).alignment = WRAP


def sheet_execution_window(wb: Workbook) -> None:
    ws = wb.create_sheet("00a Execution Window", 1)
    start = write_title_block(
        ws,
        "00a — Execution Window (2 days)",
        f"{WINDOW_LABEL}. One row per tester. No day-by-day calendar and no time slots. "
        "Update Status as you go. Testers choose their own hours.",
        cols=11,
    )
    headers = [
        "Tester",
        "Assigned sheets / work",
        "Window",
        "Cycle start",
        "Cycle end",
        "Location / URL",
        "Status",
        "Section Gate Link",
        "Notes / blockers",
        "Updated By",
        "Updated At",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER
    ws.freeze_panes = f"A{start + 1}"
    rows = execution_window_rows()
    dv_status = DataValidation(type="list", formula1=WINDOW_STATUS_LIST, allow_blank=True)
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_tester)
    edit_cols = {7, 9, 10, 11}
    peach = PatternFill("solid", fgColor="FFF7ED")
    for i, row in enumerate(rows):
        r = start + 1 + i
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.border = THIN
            cell.alignment = WRAP
            if c in edit_cols:
                cell.fill = peach
        dv_tester.add(ws.cell(row=r, column=1))
        dv_status.add(ws.cell(row=r, column=7))
        ws.row_dimensions[r].height = 72
    r = start + len(rows) + 3
    ws.cell(row=r, column=1, value="Testers").font = SECTION_FONT
    for i, name in enumerate(TESTERS):
        ws.cell(row=r + 1 + i, column=1, value=f"{i + 1}. {name}").border = THIN
    ws.cell(row=r, column=3, value="Rule").font = SECTION_FONT
    ws.cell(row=r + 1, column=3, value="P0/P1 first. P2+ if time remains in the same 2-day window.").alignment = WRAP
    for i, w in enumerate([14, 72, 42, 14, 14, 42, 14, 28, 28, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = "2563EB"


def sheet_retest_wave(wb: Workbook) -> None:
    ws = wb.create_sheet("00c Retest Wave")
    start = write_title_block(
        ws,
        "00c — Retest Wave (31 Aug 2026)",
        "Retest engineering remediations from qa-evidence/30-08-2026-remediation.md. "
        "Opening a page is NOT Pass. Undeployed backend = Blocked + blocker code, not Fail.",
        cols=14,
    )
    headers = [
        "Test Case ID",
        "Module",
        "Tester",
        "Priority",
        "Previous result (if known)",
        "Fix claimed by eng",
        "Live proof required?",
        "Runnable now? (Yes / Blocked)",
        "Blocker",
        "Pass / Fail",
        "Actual Result",
        "Defect ID",
        "Evidence",
        "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL if h not in {"Pass / Fail", "Actual Result", "Defect ID", "Evidence", "Notes"} else PatternFill("solid", fgColor="C2410C")
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER
    ws.freeze_panes = f"A{start + 1}"
    peach = PatternFill("solid", fgColor="FFF7ED")
    edit_headers = {"Pass / Fail", "Actual Result", "Defect ID", "Evidence", "Notes", "Blocker", "Runnable now? (Yes / Blocked)"}
    dv_pf = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not Run,N/A,Partial"', allow_blank=True)
    dv_run = DataValidation(type="list", formula1='"Yes,Blocked"', allow_blank=True)
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    dv_pri = DataValidation(type="list", formula1='"P0,P1,P2,P3,P4"', allow_blank=True)
    ws.add_data_validation(dv_pf)
    ws.add_data_validation(dv_run)
    ws.add_data_validation(dv_tester)
    ws.add_data_validation(dv_pri)
    last = start
    for i, row in enumerate(RETEST_WAVE):
        r = start + 1 + i
        last = r
        vals = [
            row.get("Test Case ID", ""),
            row.get("Module", ""),
            row.get("Tester", ""),
            row.get("Priority", ""),
            row.get("Previous result", ""),
            row.get("Fix claimed by eng", ""),
            row.get("Live proof required?", ""),
            row.get("Runnable now?", ""),
            row.get("Blocker", ""),
            "Not Run",
            "",
            "",
            "",
            row.get("Notes", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if headers[c - 1] in edit_headers:
                cell.fill = peach
            elif headers[c - 1] == "Priority":
                if v == "P0":
                    cell.fill = P0_F
                elif v == "P1":
                    cell.fill = P1_F
            elif r % 2 == 0:
                cell.fill = ALT_F
        dv_tester.add(ws.cell(row=r, column=3))
        dv_pri.add(ws.cell(row=r, column=4))
        dv_run.add(ws.cell(row=r, column=8))
        dv_pf.add(ws.cell(row=r, column=10))
        ws.row_dimensions[r].height = 48
    ws.conditional_formatting.add(
        f"J{start + 1}:J{last}",
        CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_F),
    )
    ws.conditional_formatting.add(
        f"J{start + 1}:J{last}",
        CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_F),
    )
    ws.conditional_formatting.add(
        f"J{start + 1}:J{last}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=WARN_F),
    )
    for i, w in enumerate([22, 22, 12, 10, 28, 36, 18, 16, 28, 12, 24, 14, 20, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{last}"
    ws.sheet_properties.tabColor = "C2410C"


def sheet_section_gate(wb: Workbook) -> None:
    """Gate: ALL FILLED allowed only when Ready?=YES (formulas)."""
    ws = wb.create_sheet("00b Section Completion Gate", 2)
    start = write_title_block(
        ws,
        "00b — Section Completion Gate",
        "RULE: Set Section Status to 'ALL FILLED — READY TO CLOSE' ONLY when Ready? = YES. "
        "Ready? becomes YES when (a) every test case Pass/Fail is not 'Not Run' and "
        "(b) every Actual Result cell is non-empty. Do not close early.",
        cols=10,
    )
    headers = [
        "Sheet",
        "Assigned Tester",
        "Total Cases",
        "Not Run (count)",
        "Blank Actual (count)",
        "Ready?",
        "Section Status (select)",
        "Closed By",
        "Closed Date",
        "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    ws.freeze_panes = f"A{start + 1}"

    dv_status = DataValidation(type="list", formula1=SECTION_STATUS_LIST, allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_tester)

    # Sheet names in workbook are truncated to 31 chars — use exact titles we create
    order = [
        "00d Live Gov Exam Proof",
        "05 Module Test Cases",
        "06 Public Pages",
        "07 Authentication",
        "08 Onboarding",
        "09 Dashboard",
        "10 Practice Coach",
        "11 Live Copilot",
        "12 Mock Interview",
        "13 Government Exams",
        "14 AI Coach Chatbot",
        "15 Prep Lab",
        "16 Documents",
        "17 Resume JD Parsing",
        "18 Answer Bank",
        "19 Interview Scheduler",
        "20 Sessions",
        "21 Reports",
        "22 Analytics",
        "23 Billing",
        "24 Credits",
        "25 Settings",
        "26 Notifications",
        "27 Integrations",
        "28 Learning Hub",
        "29 Community",
        "30 Coding Lab",
        "31 Admin Portal",
        "32 Security",
        "33 Accessibility",
        "34 Responsive Cross-Browser",
        "35 API Network Observation",
        "36 AI Fallback",
        "37 Regression",
        "38 Cross-Module Journeys",
    ]

    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    for i, sheet_name in enumerate(order):
        r = start + 1 + i
        meta = SHEET_META.get(sheet_name, {})
        first = meta.get("first_data", 5)
        last = meta.get("last", 5)
        count = meta.get("count", 0)
        owner = SHEET_OWNER.get(sheet_name, "Venkat")
        # Excel sheet title may be truncated
        ref = sheet_name[:31]
        ws.cell(row=r, column=1, value=sheet_name).border = THIN
        ws.cell(row=r, column=2, value=owner).border = THIN
        ws.cell(row=r, column=3, value=count).border = THIN
        # Pass/Fail is column U; Actual Result is column T
        not_run = f"COUNTIF('{ref}'!U{first}:U{last},\"Not Run\")"
        blank_actual = f"COUNTBLANK('{ref}'!T{first}:T{last})"
        ws.cell(row=r, column=4, value=f"={not_run}").border = THIN
        ws.cell(row=r, column=5, value=f"={blank_actual}").border = THIN
        ready = f'=IF(AND(D{r}=0,E{r}=0),"YES","NO")'
        cell_ready = ws.cell(row=r, column=6, value=ready)
        cell_ready.border = THIN
        cell_ready.alignment = CENTER
        cell_status = ws.cell(row=r, column=7, value="Not Started")
        cell_status.border = THIN
        cell_status.fill = PatternFill("solid", fgColor="FFF7ED")
        dv_status.add(cell_status)
        dv_tester.add(ws.cell(row=r, column=2))
        for c in (8, 9, 10):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN
            cell.fill = PatternFill("solid", fgColor="FFF7ED")
        ws.conditional_formatting.add(
            f"F{r}",
            CellIsRule(operator="equal", formula=['"YES"'], fill=PASS_F),
        )
        ws.conditional_formatting.add(
            f"F{r}",
            CellIsRule(operator="equal", formula=['"NO"'], fill=FAIL_F),
        )
        ws.conditional_formatting.add(
            f"G{r}",
            FormulaRule(
                formula=[f'AND(G{r}="ALL FILLED — READY TO CLOSE",F{r}="NO")'],
                fill=FAIL_F,
            ),
        )
        ws.conditional_formatting.add(
            f"G{r}",
            FormulaRule(
                formula=[f'AND(G{r}="ALL FILLED — READY TO CLOSE",F{r}="YES")'],
                fill=PASS_F,
            ),
        )

    # Summary
    last_data = start + len(order)
    r = last_data + 2
    ws.cell(row=r, column=1, value="TEAM ROLL-UP").font = SECTION_FONT
    ws.cell(row=r + 1, column=1, value="Sections Ready (YES count)").border = THIN
    ws.cell(row=r + 1, column=2, value=f'=COUNTIF(F{start + 1}:F{last_data},"YES")').border = THIN
    ws.cell(row=r + 2, column=1, value="Sections marked ALL FILLED").border = THIN
    ws.cell(
        row=r + 2,
        column=2,
        value=f'=COUNTIF(G{start + 1}:G{last_data},"ALL FILLED — READY TO CLOSE")',
    ).border = THIN
    ws.cell(row=r + 3, column=1, value="Invalid ALL FILLED (Ready=NO)").border = THIN
    ws.cell(
        row=r + 3,
        column=2,
        value=f'=COUNTIFS(G{start + 1}:G{last_data},"ALL FILLED — READY TO CLOSE",F{start + 1}:F{last_data},"NO")',
    ).border = THIN
    ws.cell(row=r + 3, column=2).fill = WARN_F

    for i, w in enumerate([28, 14, 12, 14, 16, 10, 28, 14, 12, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = "DC2626"


def add_cases(wb: Workbook, title: str, cases: list[dict], na: str | None = None) -> None:
    ws = wb.create_sheet(title[:31])
    assigned = assign_all(cases)
    meta = write_tc_sheet(ws, title, assigned, na_note=na, tester_list=TESTER_LIST)
    SHEET_META[title] = meta


def sheet_defect_log(wb: Workbook) -> None:
    ws = wb.create_sheet("39 Defect Log")
    start = write_title_block(
        ws,
        "39 — Defect Log",
        "Every Fail must have a defect row with evidence. No passwords/secrets in attachments descriptions. "
        "Open items from the prior cycle are pre-seeded — retest, do not mark Pass without evidence.",
        cols=15,
    )
    headers = [
        "Defect ID",
        "Title",
        "Module",
        "Test Case ID",
        "Severity (P0-P4)",
        "Assigned Tester",
        "Business Impact",
        "Reproducibility",
        "Environment",
        "Account ID",
        "Steps to Reproduce",
        "Expected",
        "Actual",
        "Evidence Links",
        "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    ws.freeze_panes = f"A{start + 1}"
    seeds = list(DEFECT_SEEDS)
    total_rows = max(50, len(seeds))
    peach = PatternFill("solid", fgColor="FFF7ED")
    for i in range(total_rows):
        r = start + 1 + i
        seed = seeds[i] if i < len(seeds) else {}
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=seed.get(h, ""))
            cell.border = THIN
            cell.alignment = WRAP
            if h in {"Title", "Steps to Reproduce", "Expected", "Actual", "Evidence Links", "Status", "Assigned Tester"}:
                cell.fill = peach
            elif i % 2 == 0:
                cell.fill = ALT_F
        if seed:
            ws.row_dimensions[r].height = 48
    from openpyxl.worksheet.datavalidation import DataValidation

    dv_sev = DataValidation(type="list", formula1='"P0,P1,P2,P3,P4"', allow_blank=True)
    dv_rep = DataValidation(type="list", formula1='"Always,Intermittent,Once,Unknown"', allow_blank=True)
    dv_st = DataValidation(
        type="list",
        formula1='"New,Open,In Progress,Fixed,Verified,Wont Fix,Duplicate"',
        allow_blank=True,
    )
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    ws.add_data_validation(dv_sev)
    ws.add_data_validation(dv_rep)
    ws.add_data_validation(dv_st)
    ws.add_data_validation(dv_tester)
    last = start + total_rows
    dv_sev.add(f"E{start + 1}:E{last}")
    dv_tester.add(f"F{start + 1}:F{last}")
    dv_rep.add(f"H{start + 1}:H{last}")
    dv_st.add(f"O{start + 1}:O{last}")
    for i, w in enumerate([12, 36, 18, 18, 12, 14, 18, 14, 16, 14, 40, 28, 28, 20, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_execution_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("40 Execution Summary")
    start = write_title_block(
        ws,
        "40 — Execution Summary",
        "Update counts as testing progresses. Release readiness depends on critical journeys, not page-open smoke.",
        cols=6,
    )
    headers = ["Sheet / Module", "Total Cases", "Pass", "Fail", "Blocked", "Not Run"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    modules = [
        "00d Live Gov Exam Proof",
        "05 Module Test Cases",
        "06 Public Pages",
        "07 Authentication",
        "08 Onboarding",
        "09 Dashboard",
        "10 Practice Coach",
        "11 Live Copilot",
        "12 Mock Interview",
        "13 Government Exams",
        "14 AI Coach / Chatbot",
        "15 Prep Lab",
        "16 Documents",
        "17 Resume / JD Parsing",
        "18 Answer Bank",
        "19 Interview Scheduler",
        "20 Sessions",
        "21 Reports",
        "22 Analytics",
        "23 Billing",
        "24 Credits",
        "25 Settings",
        "26 Notifications",
        "27 Integrations",
        "28 Learning Hub",
        "29 Community",
        "30 Coding Lab",
        "31 Admin Portal",
        "32 Security",
        "33 Accessibility",
        "34 Responsive / Cross-Browser",
        "35 API / Network Observation",
        "36 AI / Fallback",
        "37 Regression",
        "38 Cross-Module Journeys",
    ]
    for i, m in enumerate(modules):
        r = start + 1 + i
        ws.cell(row=r, column=1, value=m).border = THIN
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c, value=0 if c > 1 else "")
            cell.border = THIN
            cell.alignment = CENTER
    r = start + len(modules) + 2
    ws.cell(row=r, column=1, value="Notes / blockers").font = SECTION_FONT
    ws.cell(row=r + 1, column=1, value="").border = THIN
    for i, w in enumerate([36, 12, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_traceability(wb: Workbook) -> None:
    ws = wb.create_sheet("41 Traceability Matrix")
    start = write_title_block(
        ws,
        "41 — Traceability Matrix",
        "Maps user-facing modules to test case ID prefixes for coverage tracking.",
        cols=5,
    )
    headers = ["Module", "Inventory Area", "Primary Sheets", "Case ID Prefixes", "Critical Journey"]
    rows = [
        ("Public marketing", "Public", "06 Public Pages", "TC-PUB-", "—"),
        ("Authentication", "Auth", "07 Authentication", "TC-AUTH-", "Journey 1"),
        ("Onboarding", "Onboarding", "08 Onboarding", "TC-ONB-", "Journey 1"),
        ("Dashboard", "Dashboard", "09 Dashboard", "TC-DASH-", "Journey 1"),
        ("Practice Coach", "Practice Coach", "10 Practice Coach", "TC-PC-", "Journey 1/2"),
        ("Live Overlay", "Practice Coach", "11 Live Copilot", "TC-LIVE-", "Journey 1"),
        ("Mock Interview", "Mock Interview", "12 Mock Interview", "TC-MOCK-", "—"),
        ("Government Exams", "Government Exams", "13 Government Exams", "TC-GOV-", "Journey 3"),
        ("AI Chat", "AI Coach", "14 AI Coach / Chatbot", "TC-AI-", "Journey 2"),
        ("Prep Lab", "Prep Lab", "15 Prep Lab", "TC-PREP-", "—"),
        ("Documents", "Documents", "16 Documents + 17 Resume/JD", "TC-DOC-/TC-RJ-", "Journey 2"),
        ("Answer Bank", "Answer Bank", "18 Answer Bank", "TC-ANS-", "—"),
        ("Scheduler", "Scheduler", "19 Interview Scheduler", "TC-SCH-", "—"),
        ("Sessions/Reports", "Sessions/Reports", "20 Sessions + 21 Reports", "TC-SES-/TC-REP-", "Journey 1"),
        ("Analytics", "Analytics", "22 Analytics", "TC-AN-", "—"),
        ("Billing/Credits", "Billing", "23 Billing + 24 Credits", "TC-BILL-/TC-CR-", "Journey 4"),
        ("Settings", "Settings", "25 Settings", "TC-SET-", "—"),
        ("Learning/Community/Coding", "Learning/Community/Coding", "28–30", "TC-LRN-/TC-COM-/TC-COD-", "Journey 5"),
        ("Admin", "Admin", "31 Admin Portal", "TC-ADM-", "Journey 5"),
        ("Security/A11y/Responsive", "Security+", "32–34", "TC-SEC-/TC-A11Y-/TC-RSP-", "—"),
        ("API/Fallback/Regression", "Cross-cutting", "35–37", "TC-API-/TC-FB-/TC-REG-", "—"),
        ("E2E Journeys", "All", "38 Cross-Module Journeys", "TC-JRN-", "All critical"),
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    for i, row in enumerate(rows):
        r = start + 1 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    for i, w in enumerate([28, 18, 28, 28, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_release(wb: Workbook) -> None:
    ws = wb.create_sheet("42 Release Checklist")
    start = write_title_block(
        ws,
        "42 — Release Checklist",
        "Do NOT mark RELEASE READY merely because pages open. Critical journeys must Pass.",
        cols=3,
    )
    items = [
        ("P0 failures count", "0 required for RELEASE READY"),
        ("P1 failures count", "Record; justify if any"),
        ("P2/P3 failures count", "Record"),
        ("Blocked tests", "List environment blockers"),
        ("Missing credentials", "List"),
        ("Unavailable integrations", "List + Requires Configuration"),
        ("Known limitations", "List"),
        ("Regression result", "All TC-REG-* Pass?"),
        ("Browser/device coverage", "Chrome/Edge/Firefox/Safari + viewports"),
        ("Security result", "TC-SEC-* + Admin gate"),
        ("Accessibility result", "TC-A11Y-*"),
        ("Payment/sandbox result", "TC-BILL-*"),
        ("Government Exam result", "Journey 3 + TC-GOV P0 + 00d live proof"),
        ("Admin result", "TC-ADM-* + non-admin block"),
        ("Journey 1 Signup→Reports", "Pass/Fail"),
        ("Journey 2 Resume→Coach", "Pass/Fail"),
        ("Journey 3 Gov Exam lifecycle", "Pass/Fail"),
        ("Journey 4 Purchase→Consume", "Pass/Fail"),
        ("Journey 5 Admin→User content", "Pass/Fail"),
    ]
    headers = ["Checklist Item", "Result / Notes", "Owner"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    for i, (item, note) in enumerate(items):
        r = start + 1 + i
        ws.cell(row=r, column=1, value=item).border = THIN
        ws.cell(row=r, column=2, value=note).border = THIN
        ws.cell(row=r, column=3, value="").border = THIN
    r = start + len(items) + 3
    ws.cell(row=r, column=1, value="FINAL RECOMMENDATION").font = SECTION_FONT
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(
        type="list",
        formula1='"RELEASE READY,RELEASE READY WITH KNOWN LIMITATIONS,RELEASE BLOCKED"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    cell = ws.cell(row=r + 1, column=1, value="RELEASE BLOCKED")
    cell.fill = FAIL_F
    cell.font = Font(bold=True, size=14)
    cell.border = THIN
    dv.add(cell)
    ws.cell(row=r + 2, column=1, value="Justification / known limitations:").font = BODY_FONT
    ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 6, end_column=3)
    ws.cell(row=r + 3, column=1, value=(
        "NO_GO / RELEASE BLOCKED until a real Custom or Full Mock paper completes on "
        "https://clarify.ai.sltfinanceindia.com (00d TC-GOV-LIVE-06). Official-empty is an accepted "
        "limitation only after that paper exists. Migrations + Edge + Python worker were not live "
        "as of 31 Aug 2026."
    )).alignment = WRAP
    for i, w in enumerate([40, 48, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = "DC2626"


def build() -> Path:
    global QA_ENV
    SHEET_META.clear()
    QA_ENV = load_qa_env(ROOT)
    if not QA_ENV.get("QA_FREE_EMAIL"):
        print("WARNING: .env.qa.local missing or incomplete — passwords may be blank. Run: npm run qa:seed-accounts")

    wb = Workbook()
    sheet_cover(wb)
    sheet_execution_window(wb)
    # Section gate created AFTER cases so formulas know row ranges.
    sheet_inventory(wb)
    sheet_roles(wb)
    sheet_accounts(wb)
    sheet_environment(wb)

    add_cases(wb, "05 Module Test Cases", module_index_cases())
    add_cases(wb, "06 Public Pages", public_cases())
    add_cases(wb, "07 Authentication", auth_cases())
    add_cases(wb, "08 Onboarding", onboarding_cases())
    add_cases(wb, "09 Dashboard", dashboard_cases())
    add_cases(wb, "10 Practice Coach", practice_coach_cases())
    add_cases(wb, "11 Live Copilot", live_copilot_cases())
    add_cases(wb, "12 Mock Interview", mock_interview_cases())
    add_cases(wb, "13 Government Exams", gov_exam_cases())
    add_cases(wb, "14 AI Coach Chatbot", ai_coach_cases())
    add_cases(wb, "15 Prep Lab", prep_lab_cases())
    add_cases(wb, "16 Documents", documents_cases())
    add_cases(wb, "17 Resume JD Parsing", resume_jd_cases())
    add_cases(wb, "18 Answer Bank", answer_bank_cases())
    add_cases(wb, "19 Interview Scheduler", scheduler_cases())
    add_cases(wb, "20 Sessions", sessions_cases())
    add_cases(wb, "21 Reports", reports_cases())
    add_cases(wb, "22 Analytics", analytics_cases())
    add_cases(wb, "23 Billing", billing_cases())
    add_cases(wb, "24 Credits", credits_cases())
    add_cases(wb, "25 Settings", settings_cases())
    add_cases(wb, "26 Notifications", notifications_cases())
    add_cases(wb, "27 Integrations", integrations_cases())
    add_cases(wb, "28 Learning Hub", learning_cases())
    add_cases(wb, "29 Community", community_cases())
    add_cases(wb, "30 Coding Lab", coding_lab_cases())
    add_cases(wb, "31 Admin Portal", admin_cases())
    add_cases(wb, "32 Security", security_cases())
    add_cases(wb, "33 Accessibility", accessibility_cases())
    add_cases(wb, "34 Responsive Cross-Browser", responsive_cases())
    add_cases(wb, "35 API Network Observation", api_network_cases())
    add_cases(wb, "36 AI Fallback", ai_fallback_cases())
    add_cases(wb, "37 Regression", regression_cases())
    add_cases(wb, "38 Cross-Module Journeys", journey_cases())
    add_cases(wb, "00d Live Gov Exam Proof", gov_exam_live_cases())
    sheet_retest_wave(wb)

    sheet_section_gate(wb)

    def _move_to(name: str, idx: int) -> None:
        cur = wb.sheetnames.index(name)
        if cur != idx:
            wb.move_sheet(name, offset=idx - cur)

    _move_to("00b Section Completion Gate", 2)
    _move_to("00c Retest Wave", 3)
    _move_to("00d Live Gov Exam Proof", 4)

    sheet_defect_log(wb)
    sheet_execution_summary(wb)
    sheet_traceability(wb)
    sheet_release(wb)

    saved: list[Path] = []
    for dest in (OUT_DATED, OUT_ALT, OUT):
        try:
            wb.save(dest)
            saved.append(dest)
        except PermissionError:
            print(f"NOTE: {dest.name} locked; skipped")
    if not saved:
        raise PermissionError("Could not write any workbook copy — close Excel and retry")
    return saved[0]


if __name__ == "__main__":
    path = build()
    from openpyxl import load_workbook
    from collections import Counter

    wb2 = load_workbook(path, read_only=True, data_only=False)
    total = 0
    by_tester: Counter[str] = Counter()
    print(f"Wrote: {path}")
    print(f"Sheets ({len(wb2.sheetnames)}):")
    for name in wb2.sheetnames:
        ws = wb2[name]
        n = 0
        # Tester is column 23 (W)
        for row in ws.iter_rows(min_col=1, max_col=23, values_only=True):
            v = row[0]
            if isinstance(v, str) and v.startswith("TC-"):
                n += 1
                tester = row[22] if len(row) > 22 else ""
                if tester:
                    by_tester[str(tester)] += 1
        total += n
        print(f"  - {name}" + (f" ({n} cases)" if n else ""))
    print(f"Total atomic test cases: {total}")
    print("Assignment counts:")
    for t in TESTERS:
        print(f"  {t}: {by_tester.get(t, 0)}")
    unexpected = [t for t in by_tester if t not in TESTERS]
    if unexpected:
        print(f"UNEXPECTED TESTERS: {unexpected}")
    # 00c tester is column C
    ws_retest = wb2["00c Retest Wave"]
    retest_by: Counter[str] = Counter()
    for row in ws_retest.iter_rows(min_row=5, max_col=3, values_only=True):
        if row[0] and row[2] in TESTERS:
            retest_by[str(row[2])] += 1
    print("00c Retest Wave rows:")
    for t in TESTERS:
        print(f"  {t}: {retest_by.get(t, 0)}")
    cover = "\n".join(
        str(c[0] or "") for c in wb2["00 Cover & Instructions"].iter_rows(max_col=1, values_only=True)
    )
    if "Raj Balani" in cover or "Raj Balani" in "".join(wb2.sheetnames):
        print("ERROR: Raj Balani still present")
    else:
        print("Raj Balani: not present on cover or sheet names")
    # Spot-check credentials sheet has emails
    ws_acc = wb2["03 Test Accounts"]
    emails = []
    for row in ws_acc.iter_rows(min_row=5, max_col=6, values_only=True):
        if row[0] and row[4] and "@" in str(row[4]):
            emails.append(str(row[4]))
    print(f"Credential emails loaded: {len(emails)}")
    print(f"Has passwords column populated: {any(r[5] for r in ws_acc.iter_rows(min_row=5, max_col=6, values_only=True) if r[0] and r[5] and r[5] != 'N/A')}")
