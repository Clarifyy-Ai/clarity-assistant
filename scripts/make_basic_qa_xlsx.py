"""Generate a basic QA Excel checklist from QA_CHECKLIST_COMPLETE.txt."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
TXT = ROOT / "QA_CHECKLIST_COMPLETE.txt"
OUT = ROOT / "Clarify_AI_QA_Checklist_Basic.xlsx"


def parse_rows(text: str) -> list[dict]:
    section = ""
    subsection = ""
    rows: list[dict] = []
    item_no = 0

    for raw in text.splitlines():
        line = raw.rstrip()
        if not line or line.startswith("=") or line.startswith("---"):
            continue

        m_sec = re.match(r"^(\d+)\.\s+(.+)$", line)
        if m_sec and not line.startswith("["):
            section = f"{m_sec.group(1)}. {m_sec.group(2).strip()}"
            section = re.sub(r"\s*\(site:.*\)$", "", section)
            subsection = ""
            continue

        m_sub = re.match(r"^(\d+\.\d+)\s+(.+)$", line)
        if m_sub and not line.startswith("["):
            subsection = f"{m_sub.group(1)} {m_sub.group(2).strip()}"
            continue

        m_item = re.match(r"^\[\s*\]\s+(P[012])\s+(.+)$", line)
        if m_item:
            item_no += 1
            rows.append(
                {
                    "id": f"QA-{item_no:03d}",
                    "section": section or "General",
                    "subsection": subsection,
                    "priority": m_item.group(1),
                    "test": m_item.group(2).strip(),
                }
            )
    return rows


def main() -> None:
    rows = parse_rows(TXT.read_text(encoding="utf-8"))

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True)
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    cred_fill = PatternFill("solid", fgColor="FFF2CC")
    p0_fill = PatternFill("solid", fgColor="FCE4D6")
    p1_fill = PatternFill("solid", fgColor="DDEBF7")
    p2_fill = PatternFill("solid", fgColor="E2EFDA")

    # --- Instructions ---
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["A1"] = "Clarify AI — QA Checklist (Basic)"
    ws0["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws0.merge_cells("A1:D1")

    info = [
        ("", ""),
        ("Test Site (use this only)", "https://clarify.ai.sltfinanceindia.com/"),
        ("Login", "https://clarify.ai.sltfinanceindia.com/login"),
        ("Signup", "https://clarify.ai.sltfinanceindia.com/signup"),
        ("App Dashboard", "https://clarify.ai.sltfinanceindia.com/app/dashboard"),
        ("", ""),
        ("Tester Name", ""),
        ("Date", ""),
        ("Browser / OS", "Chrome (primary)"),
        ("", ""),
        ("How to mark Status", "Not Run | Pass | Fail | Blocked | N/A"),
        ("Priority order", "Finish ALL P0 first, then P1, then P2"),
        ("Account order", "1) Pro  2) Free  3) Max  4) Admin"),
        ("", ""),
        ("Stripe test card", "4242 4242 4242 4242"),
        ("Expiry / CVC / ZIP", "12/34  |  123  |  400001"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws0.cell(i, 1, k).font = label_font
        ws0.cell(i, 2, v)

    ws0["A20"] = "QA LOGIN CREDENTIALS (internal — do not post publicly)"
    ws0["A20"].font = Font(bold=True, size=12, color="C00000")
    ws0.merge_cells("A20:D20")

    for col, h in enumerate(["Role", "Email", "Password location", "Use for"], 1):
        cell = ws0.cell(21, col, h)
        cell.fill = header_fill
        cell.font = header_font

    creds = [
        ("Free", "qa.free@clarify.ai.test", ".env.qa.local → QA_FREE_PASSWORD", "Plan limits / upgrade gates"),
        ("Pro", "qa.pro@clarify.ai.test", ".env.qa.local → QA_PRO_PASSWORD", "Main feature coverage (start here)"),
        ("Max", "qa.max@clarify.ai.test", ".env.qa.local → QA_MAX_PASSWORD", "Max-tier / high credits"),
        ("Admin", "qa.admin@clarify.ai.test", ".env.qa.local → QA_ADMIN_PASSWORD", "Admin portal /app/admin only"),
        ("User A", "qa.user-a@clarify.ai.test", ".env.qa.local → QA_USER_A_PASSWORD", "RLS isolation owner"),
        ("User B", "qa.user-b@clarify.ai.test", ".env.qa.local → QA_USER_B_PASSWORD", "RLS isolation peer"),
    ]
    for r, row in enumerate(creds, 22):
        for c, val in enumerate(row, 1):
            cell = ws0.cell(r, c, val)
            cell.fill = cred_fill
            cell.border = thin

    ws0["A27"] = "Same-day tip: Mark Pass/Fail in Checklist sheet. Log failures on Fail Log sheet."
    ws0["A27"].font = Font(italic=True)
    ws0.merge_cells("A27:D27")

    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 36
    ws0.column_dimensions["C"].width = 24
    ws0.column_dimensions["D"].width = 40

    # --- Checklist ---
    ws = wb.create_sheet("Checklist")
    headers = [
        "ID",
        "Section",
        "Subsection",
        "Priority",
        "Test Case",
        "Status",
        "Account Used",
        "Notes / Bug link",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row["id"]).border = thin
        ws.cell(i, 2, row["section"]).border = thin
        ws.cell(i, 3, row["subsection"]).border = thin
        pcell = ws.cell(i, 4, row["priority"])
        pcell.border = thin
        pcell.alignment = Alignment(horizontal="center")
        if row["priority"] == "P0":
            pcell.fill = p0_fill
        elif row["priority"] == "P1":
            pcell.fill = p1_fill
        else:
            pcell.fill = p2_fill
        tcell = ws.cell(i, 5, row["test"])
        tcell.border = thin
        tcell.alignment = Alignment(wrap_text=True, vertical="top")
        scell = ws.cell(i, 6, "Not Run")
        scell.border = thin
        scell.alignment = Alignment(horizontal="center")
        ws.cell(i, 7, "").border = thin
        ws.cell(i, 8, "").border = thin

    dv = DataValidation(
        type="list",
        formula1='"Not Run,Pass,Fail,Blocked,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(rows) + 1}")

    dv2 = DataValidation(
        type="list",
        formula1='"Pro,Free,Max,Admin,New signup,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(dv2)
    dv2.add(f"G2:G{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 36
    ws.row_dimensions[1].height = 22

    # --- Fail Log ---
    wsf = wb.create_sheet("Fail Log")
    fail_headers = [
        "ID",
        "Section",
        "Account",
        "URL",
        "Steps",
        "Expected",
        "Actual",
        "Screenshot",
        "Severity",
    ]
    for col, h in enumerate(fail_headers, 1):
        cell = wsf.cell(1, col, h)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = header_font
    for r in range(2, 32):
        for c in range(1, 10):
            wsf.cell(r, c, "").border = thin
    for col, w in enumerate([10, 22, 12, 40, 30, 25, 25, 20, 12], 1):
        wsf.column_dimensions[get_column_letter(col)].width = w
    wsf.freeze_panes = "A2"
    sev = DataValidation(type="list", formula1='"P0,P1,P2"', allow_blank=True)
    wsf.add_data_validation(sev)
    sev.add("I2:I31")

    # --- Sign-off ---
    wss = wb.create_sheet("Sign-off")
    wss["A1"] = "QA Sign-off"
    wss["A1"].font = Font(bold=True, size=14, color="1F4E79")
    sign_rows = [
        ("Site tested", "https://clarify.ai.sltfinanceindia.com/"),
        ("Total items", len(rows)),
        ("Passed", ""),
        ("Failed", ""),
        ("Blocked", ""),
        ("N/A", ""),
        ("Not Run", ""),
        ("", ""),
        ("Release recommendation", "Go / Go with known issues / No-go"),
        ("Tester name", ""),
        ("Tester signature / date", ""),
        ("Reviewer name", ""),
        ("Reviewer signature / date", ""),
    ]
    for i, (k, v) in enumerate(sign_rows, start=3):
        wss.cell(i, 1, k).font = label_font
        wss.cell(i, 2, v).border = thin
    wss.column_dimensions["A"].width = 28
    wss.column_dimensions["B"].width = 50

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Checklist rows: {len(rows)}")


if __name__ == "__main__":
    main()
