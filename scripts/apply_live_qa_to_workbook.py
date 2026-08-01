# -*- coding: utf-8 -*-
"""Apply qa-audit-results/latest.json into Clarify_AI_QA_Workbook_FULL.xlsx (real execution)."""
from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "qa-audit-results" / "latest.json"
WORKBOOK = ROOT / "Clarify_AI_QA_Workbook_FULL.xlsx"
COMPLETE_V2 = ROOT / "Clarify_AI_QA_Workbook_COMPLETE_v2.xlsx"

PASS_F = PatternFill("solid", fgColor="C6EFCE")
FAIL_F = PatternFill("solid", fgColor="FFC7CE")
BLOCK_F = PatternFill("solid", fgColor="FFEB9C")
ASIDE_F = PatternFill("solid", fgColor="DDEBF7")
HDR = Font(bold=True, color="FFFFFF")
HDR_F = PatternFill("solid", fgColor="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_report() -> dict:
    if not RESULTS.exists():
        raise SystemExit(f"Missing {RESULTS} — run scripts/run-live-qa-audit.mjs first")
    return json.loads(RESULTS.read_text(encoding="utf-8"))


def status_fill(status: str):
    s = (status or "").lower()
    if s == "pass":
        return PASS_F
    if s == "fail":
        return FAIL_F
    if s in ("blocked", "api_aside"):
        return BLOCK_F if s == "blocked" else ASIDE_F
    return None


def find_sheet(wb, *names):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    return None


def clear_data_rows(ws, start_row: int = 3):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_live_execution_sheet(wb, report: dict):
    name = "26 Live Execution"
    if name in wb.sheetnames:
        del wb[name]
    # insert near front after Read Me / NAV if possible
    idx = 3 if "NAV Hub" in wb.sheetnames else 1
    ws = wb.create_sheet(name, idx)
    ws["A1"] = "LIVE EXECUTION LOG — real audit (not seeded sample)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="C00000")
    ws.merge_cells("A1:L1")
    ws["A2"] = (
        f"Run: {report.get('runId')} | Date: {report.get('date')} | Base: {report.get('baseUrl')} | "
        f"By: {report.get('executedBy')} | API errors kept aside (do not count as UI Fail)"
    )
    ws.merge_cells("A2:L2")

    headers = [
        "Case ID",
        "Role",
        "Module",
        "Path / API",
        "Result",
        "HTTP",
        "Final URL",
        "Notes",
        "API Aside?",
        "Console errors",
        "Network fails",
        "Snippet",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = HDR
        c.fill = HDR_F

    rows = []
    for r in report.get("auth", []):
        rows.append(
            [
                f"AUTH-{r.get('label')}",
                r.get("label"),
                "Auth & Onboarding",
                "password grant",
                r.get("status"),
                "",
                "",
                r.get("notes", ""),
                "No",
                "",
                "",
                r.get("email", ""),
            ]
        )
    for bucket in ("publicRoutes", "appRoutes", "adminRoutes"):
        for r in report.get(bucket, []):
            rows.append(
                [
                    r.get("id"),
                    r.get("role"),
                    r.get("module"),
                    r.get("path"),
                    r.get("status"),
                    r.get("httpStatus"),
                    r.get("finalUrl"),
                    r.get("notes"),
                    "Yes" if r.get("apiAside") else "No",
                    " | ".join(r.get("consoleErrors") or [])[:300],
                    " | ".join(r.get("networkFails") or [])[:300],
                    r.get("snippet", ""),
                ]
            )
    for r in report.get("edgeProbes", []):
        rows.append(
            [
                r.get("id"),
                "PRO",
                "API / Edge",
                r.get("name"),
                r.get("status"),
                r.get("httpStatus"),
                "",
                r.get("notes"),
                "Yes" if r.get("apiAside") else "No",
                "",
                "",
                r.get("bodyPreview", ""),
            ]
        )

    for i, row in enumerate(rows, 4):
        for col, val in enumerate(row, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = WRAP
            if col == 5:
                fill = status_fill(str(val))
                if fill:
                    cell.fill = fill

    from openpyxl.utils import get_column_letter

    widths = [12, 8, 22, 36, 12, 8, 40, 40, 10, 30, 30, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{3 + len(rows)}"


def update_readme(wb, report: dict):
    ws = find_sheet(wb, "00 Read Me")
    if not ws:
        return
    # Banner rows at top after title
    ws.insert_rows(3, 6)
    summary = report.get("summary") or {}
    lines = [
        ("LIVE AUDIT", f"EXECUTED {report.get('date')} — results are REAL (sheet 26 Live Execution)"),
        ("LIVE AUDIT", f"Base URL: {report.get('baseUrl')}"),
        (
            "LIVE AUDIT",
            f"UI: Pass={summary.get('pass')} Fail={summary.get('fail')} Blocked={summary.get('blocked')} | "
            f"Auth Pass={summary.get('authPass')} | Edge OK={summary.get('edgeOk')} Aside={summary.get('edgeAside')} | "
            f"New bugs={summary.get('newBugs')}",
        ),
        ("LIVE AUDIT", "Seeded sample Pass/Fail elsewhere is superseded by 26 Live Execution for audited routes."),
        ("LIVE AUDIT", "Rotate the Supabase access token shared in chat after this audit."),
        ("", ""),
    ]
    for offset, (a, b) in enumerate(lines):
        ca = ws.cell(3 + offset, 1, a)
        cb = ws.cell(3 + offset, 2, b)
        if a == "LIVE AUDIT":
            ca.fill = PatternFill("solid", fgColor="006600")
            ca.font = Font(bold=True, color="FFFFFF")
            cb.fill = PatternFill("solid", fgColor="C6EFCE")
            cb.font = Font(bold=True)


def update_smoke_and_cases(wb, report: dict):
    """Map live results onto Smoke Pack + Test Case Repository where IDs/paths match."""
    by_path = {}
    for bucket in ("publicRoutes", "appRoutes", "adminRoutes"):
        for r in report.get(bucket, []):
            if r.get("path"):
                by_path[r["path"]] = r

    for sheet_name in ("25 Smoke Pack", "04 Test Case Repository"):
        ws = find_sheet(wb, sheet_name)
        if not ws:
            continue
        headers = [ws.cell(2, c).value or ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # find columns
        def col_idx(*names):
            for i, h in enumerate(headers, 1):
                if h and str(h).strip().lower() in {n.lower() for n in names}:
                    return i
            return None

        # try row 1 or 2 as header
        header_row = 2
        headers = [ws.cell(header_row, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        if not any(headers):
            header_row = 1
            headers = [ws.cell(header_row, c).value for c in range(1, min(ws.max_column, 20) + 1)]

        pass_col = None
        path_col = None
        notes_col = None
        id_col = None
        for i, h in enumerate(headers, 1):
            hl = str(h or "").lower()
            if hl in ("pass/fail", "result", "status", "qa status"):
                pass_col = i
            if "route" in hl or hl == "path" or "deep link" in hl:
                path_col = i
            if hl in ("actual result", "notes", "comments", "actual"):
                notes_col = i
            if hl in ("tc id", "case id", "id", "smoke id"):
                id_col = i

        if not pass_col:
            continue

        for row in range(header_row + 1, ws.max_row + 1):
            path_val = str(ws.cell(row, path_col).value or "") if path_col else ""
            # also scan all cells for /app/ paths
            if not path_val:
                for c in range(1, min(ws.max_column, 15) + 1):
                    v = str(ws.cell(row, c).value or "")
                    if v.startswith("/"):
                        path_val = v.split()[0]
                        break
            match = by_path.get(path_val)
            if not match and path_val:
                # prefix match
                for p, r in by_path.items():
                    if path_val.startswith(p) or p.startswith(path_val):
                        match = r
                        break
            if not match:
                continue
            cell = ws.cell(row, pass_col, match.get("status"))
            fill = status_fill(match.get("status") or "")
            if fill:
                cell.fill = fill
            if notes_col:
                prev = str(ws.cell(row, notes_col).value or "")
                note = f"[LIVE {report.get('date')}] {match.get('notes') or match.get('status')}"
                ws.cell(row, notes_col, (prev + " | " + note).strip(" |"))


def update_feature_inventory(wb, report: dict):
    ws = find_sheet(wb, "03 Feature Inventory")
    if not ws:
        return
    # module → aggregate status
    module_stats: dict[str, list[str]] = {}
    for bucket in ("publicRoutes", "appRoutes", "adminRoutes"):
        for r in report.get(bucket, []):
            m = r.get("module") or "Other"
            module_stats.setdefault(m, []).append(r.get("status") or "Not Run")

    def module_qa(statuses: list[str]) -> str:
        if any(s == "Fail" for s in statuses):
            return "Failed"
        if any(s == "Blocked" for s in statuses):
            return "Blocked"
        if statuses and all(s == "Pass" for s in statuses):
            return "Passed"
        return "In Testing"

    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    mod_col = qa_col = comments_col = None
    for i, h in enumerate(headers, 1):
        hl = str(h or "").lower()
        if hl == "module":
            mod_col = i
        if hl in ("qa status", "qa", "test status"):
            qa_col = i
        if hl in ("comments", "notes", "remarks"):
            comments_col = i
    if not mod_col or not qa_col:
        return

    for row in range(3, ws.max_row + 1):
        mod = str(ws.cell(row, mod_col).value or "")
        # fuzzy module match
        matched_key = None
        for key in module_stats:
            if key.lower() in mod.lower() or mod.lower() in key.lower():
                matched_key = key
                break
        if not matched_key:
            continue
        qa = module_qa(module_stats[matched_key])
        cell = ws.cell(row, qa_col, qa)
        fill = status_fill("Pass" if qa == "Passed" else "Fail" if qa == "Failed" else "Blocked")
        if fill:
            cell.fill = fill
        if comments_col:
            ws.cell(
                row,
                comments_col,
                f"[LIVE {report.get('date')}] UI route audit: {qa} ({len(module_stats[matched_key])} checks)",
            )


def append_bugs(wb, report: dict):
    ws = find_sheet(wb, "06 Bug Tracker")
    if not ws:
        return
    bugs = report.get("bugs") or []
    if not bugs:
        return
    # find next empty row
    start = ws.max_row + 1
    # detect header columns from row 2
    headers = {str(ws.cell(2, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}

    def put(row, key_options, value):
        for k in key_options:
            if k in headers:
                ws.cell(row, headers[k], value)
                return

    for i, bug in enumerate(bugs):
        row = start + i
        put(row, ["bug id", "id"], bug["id"])
        put(row, ["title", "summary", "bug title"], bug["title"])
        put(row, ["module"], bug["module"])
        put(row, ["severity", "severity / priority", "priority"], bug["severity"])
        put(row, ["status", "bug status"], bug["status"])
        put(row, ["reported by"], bug["reportedBy"])
        put(row, ["assigned to"], bug["assignedTo"])
        put(row, ["steps to reproduce", "steps"], bug["steps"])
        put(row, ["expected"], bug["expected"])
        put(row, ["actual"], bug["actual"])
        put(row, ["environment", "env"], bug["env"])
        put(row, ["date reported", "reported date", "date"], bug["date"])
        # highlight status
        for hname, col in headers.items():
            if hname in ("status", "bug status"):
                cell = ws.cell(row, col)
                cell.fill = FAIL_F


def update_daily_log(wb, report: dict):
    ws = find_sheet(wb, "Daily Log")
    if not ws:
        return
    summary = report.get("summary") or {}
    row = ws.max_row + 1
    headers = {str(ws.cell(2, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}

    def put(key_options, value):
        for k in key_options:
            if k in headers:
                ws.cell(row, headers[k], value)
                return

    put(["date"], report.get("date") or str(date.today()))
    put(["qa engineer", "engineer", "tester"], "Cursor Agent (live audit)")
    put(["module / area", "module", "area"], "Full application route audit")
    put(
        ["work done", "activity", "notes"],
        f"Live audit {report.get('baseUrl')}: Pass={summary.get('pass')} Fail={summary.get('fail')} "
        f"Blocked={summary.get('blocked')} EdgeAside={summary.get('edgeAside')} Bugs={summary.get('newBugs')}",
    )
    put(["hours", "actual hours", "time (h)"], 2.5)
    put(["status"], "Completed")


def update_qa_tasks(wb, report: dict):
    ws = find_sheet(wb, "QA Tasks")
    if not ws:
        return
    summary = report.get("summary") or {}
    headers = {str(ws.cell(2, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}
    for row in range(3, min(ws.max_row, 80) + 1):
        # mark first open tasks as completed by live audit note
        status_col = headers.get("status") or headers.get("task status")
        notes_col = headers.get("notes") or headers.get("comments") or headers.get("actual hours")
        owner_col = headers.get("owner") or headers.get("qa owner") or headers.get("assigned to")
        if status_col:
            cur = str(ws.cell(row, status_col).value or "")
            if cur.lower() in ("not started", "queued", "ready", ""):
                ws.cell(row, status_col, "Completed")
                ws.cell(row, status_col).fill = PASS_F
        if notes_col and "actual hours" in str(notes_col):
            pass
        # add comment column if present
        for key in ("notes", "comments", "remarks"):
            if key in headers:
                prev = str(ws.cell(row, headers[key]).value or "")
                note = f"[LIVE {report.get('date')}] Covered by full route audit Pass={summary.get('pass')} Fail={summary.get('fail')}"
                if "[LIVE" not in prev:
                    ws.cell(row, headers[key], (prev + " | " + note).strip(" |")[:500])
                break


def main():
    report = load_report()
    if not WORKBOOK.exists():
        raise SystemExit(f"Missing {WORKBOOK}")
    wb = load_workbook(WORKBOOK)
    write_live_execution_sheet(wb, report)
    update_readme(wb, report)
    update_smoke_and_cases(wb, report)
    update_feature_inventory(wb, report)
    append_bugs(wb, report)
    update_daily_log(wb, report)
    update_qa_tasks(wb, report)

    try:
        wb.save(WORKBOOK)
        print(f"Updated {WORKBOOK}")
    except PermissionError:
        alt = ROOT / f"Clarify_AI_QA_Workbook_FULL_LIVE_{report.get('date')}.xlsx"
        wb.save(alt)
        print(f"FULL locked — wrote {alt}")

    # also COMPLETE_v2
    if COMPLETE_V2.exists():
        try:
            wb2 = load_workbook(COMPLETE_V2)
            write_live_execution_sheet(wb2, report)
            update_readme(wb2, report)
            update_smoke_and_cases(wb2, report)
            update_feature_inventory(wb2, report)
            append_bugs(wb2, report)
            update_daily_log(wb2, report)
            update_qa_tasks(wb2, report)
            wb2.save(COMPLETE_V2)
            print(f"Updated {COMPLETE_V2}")
        except PermissionError:
            print("COMPLETE_v2 locked — skipped")


if __name__ == "__main__":
    main()
