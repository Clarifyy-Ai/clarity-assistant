#!/usr/bin/env python3
"""Merge docs/qa/audits/p0-p1-live-results.json into Clarify_AI_QA_Checklist_Basic.xlsx."""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Clarify_AI_QA_Checklist_Basic.xlsx"
OUT_XLSX = ROOT / "Clarify_AI_QA_Checklist_Basic.filled.xlsx"
RESULTS = ROOT / "docs/qa/audits/p0-p1-live-results.json"

STATUS_FILL = {
    "Pass": PatternFill("solid", fgColor="C6EFCE"),
    "Fail": PatternFill("solid", fgColor="FFC7CE"),
    "Blocked": PatternFill("solid", fgColor="FFEB9C"),
    "Not Run": PatternFill("solid", fgColor="D9D9D9"),
    "N/A": PatternFill("solid", fgColor="BDD7EE"),
}


def main() -> None:
    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    results = data["results"]

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Checklist"]
    updated = 0
    counts = {"Pass": 0, "Fail": 0, "Blocked": 0, "N/A": 0, "Not Run": 0}
    fails = []
    p0p1_total = 0
    p0p1_pass = 0
    p0p1_fail = 0
    p0p1_blocked = 0
    p0p1_notrun = 0

    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row, 1).value
        if not cid:
            continue
        pri = ws.cell(row, 2).value
        is_p0p1 = pri in ("P0", "P1")
        if is_p0p1:
            p0p1_total += 1

        if cid in results:
            r = results[cid]
            ws.cell(row, 9, r.get("actual", ""))
            ws.cell(row, 10, r.get("status", "Not Run"))
            ws.cell(row, 12, r.get("evidence", ""))
            note = r.get("notes", "")
            if note:
                prev = ws.cell(row, 14).value or ""
                ws.cell(row, 14, f"{note} | {prev}".strip(" |") if prev else note)
            fill = STATUS_FILL.get(r.get("status", ""), None)
            if fill:
                ws.cell(row, 10).fill = fill
            updated += 1
            if r.get("status") == "Fail":
                fails.append(
                    {
                        "id": cid,
                        "module": ws.cell(row, 3).value,
                        "account": "see run",
                        "url": ws.cell(row, 15).value,
                        "steps": (ws.cell(row, 7).value or "")[:200],
                        "expected": (ws.cell(row, 8).value or "")[:200],
                        "actual": r.get("actual", ""),
                        "evidence": r.get("evidence", ""),
                        "severity": pri,
                    }
                )
        elif is_p0p1 and (ws.cell(row, 10).value in (None, "", "Not Run")):
            ws.cell(row, 10, "Not Run")
            ws.cell(row, 10).fill = STATUS_FILL["Not Run"]
            prev = ws.cell(row, 14).value or ""
            tag = "Awaiting manual/hardware run — not auto-passed"
            if tag not in str(prev):
                ws.cell(row, 14, f"{tag} | {prev}".strip(" |") if prev else tag)

        status = ws.cell(row, 10).value or "Not Run"
        if status in counts:
            counts[status] += 1
        else:
            counts["Not Run"] += 1

        if is_p0p1:
            if status == "Pass":
                p0p1_pass += 1
            elif status == "Fail":
                p0p1_fail += 1
            elif status == "Blocked":
                p0p1_blocked += 1
            else:
                p0p1_notrun += 1

    fl = wb["Fail Log"]
    if fl.max_row > 1:
        fl.delete_rows(2, fl.max_row - 1)
    for i, f in enumerate(fails, start=2):
        fl.cell(i, 1, f["id"])
        fl.cell(i, 2, f["module"])
        fl.cell(i, 3, f["account"])
        fl.cell(i, 4, f["url"])
        fl.cell(i, 5, f["steps"])
        fl.cell(i, 6, f["expected"])
        fl.cell(i, 7, f["actual"])
        fl.cell(i, 8, f["evidence"])
        fl.cell(i, 9, f["severity"])

    so = wb["Sign-off"]
    so["B4"] = ws.max_row - 1
    so["B5"] = counts.get("Pass", 0)
    so["B6"] = counts.get("Fail", 0)
    so["B7"] = counts.get("Blocked", 0)
    so["B8"] = counts.get("N/A", 0)
    so["B9"] = counts.get("Not Run", 0)
    so["B11"] = dt.date.today().isoformat()
    so["B12"] = data.get("ts", "")
    so["A16"] = "P0/P1 Pass"
    so["B16"] = p0p1_pass
    so["A17"] = "P0/P1 Fail"
    so["B17"] = p0p1_fail
    so["A18"] = "P0/P1 Blocked"
    so["B18"] = p0p1_blocked
    so["A19"] = "P0/P1 Not Run"
    so["B19"] = p0p1_notrun
    so["A20"] = "P0/P1 Total"
    so["B20"] = p0p1_total
    if p0p1_fail:
        so["B13"] = "FAIL_NO_GO — fix Fail Log (esp. CookieConsent deploy) before release claim"
    elif p0p1_notrun > 50:
        so["B13"] = "CONDITIONAL_PASS — automated P0/P1 subset verified; remaining need manual/hardware"
    else:
        so["B13"] = "CONDITIONAL_PASS — continue manual Blocked items"
    so["B14"] = "Cursor Agent P0/P1 production runner"

    try:
        wb.save(XLSX)
        saved = XLSX
    except PermissionError:
        wb.save(OUT_XLSX)
        saved = OUT_XLSX
        print(f"NOTE: {XLSX.name} is locked (close Excel). Wrote {OUT_XLSX.name} instead.")
    print(f"Updated {updated} rows in {saved.name}")
    print("All counts:", counts)
    print(
        f"P0/P1 only: pass={p0p1_pass} fail={p0p1_fail} blocked={p0p1_blocked} not_run={p0p1_notrun} total={p0p1_total}"
    )
    print("Fails:", len(fails))


if __name__ == "__main__":
    main()
