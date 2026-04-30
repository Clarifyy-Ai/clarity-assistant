#!/usr/bin/env python3
"""Parse the QA checklist xlsx → catalog.json (T-ID indexed)."""
import json, sys
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path("/mnt/documents/clarify-ai-qa-checklist.xlsx")
OUT  = Path("src/test/_generated/catalog.json")

wb = load_workbook(XLSX, data_only=True)
ws = wb["All Tests"]
items = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    items.append({
        "id":          row[0],
        "part":        row[1],
        "section":     row[2],
        "subsection":  row[3],
        "test":        row[4],
        "priority":    row[5],
        "status":      row[6],
    })

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(items, indent=2))
print(f"Wrote {len(items)} items to {OUT}")
