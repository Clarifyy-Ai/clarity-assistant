#!/usr/bin/env python3
"""
Clarify AI — Enterprise QA Testing Workbook generator
Google Sheets–compatible .xlsx with 20 work sheets + Lists + Daily Log.

Domain sheets 14–16 are Clarify-specific (not Marketplace/Checkout/Payment):
  14 Live Coaching Testing
  15 Gov Exam Mock Testing
  16 Billing Credits Testing
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from qa_workbook_inventory import (
    BUGS,
    FEATURES,
    OPS_SECRETS_CHECKLIST,
    feature_counts_by_module,
    feature_qa_columns,
)

OUT = Path(__file__).resolve().parents[1] / "Clarify_AI_Enterprise_QA_Testing_Workbook.xlsx"
OUT_FALLBACK = Path(__file__).resolve().parents[1] / "Clarify_AI_Enterprise_QA_Testing_Workbook_v2.xlsx"
OUT_COMPLETE = Path(__file__).resolve().parents[1] / "Clarify_AI_QA_Workbook_COMPLETE.xlsx"
OUT_COMPLETE_V2 = Path(__file__).resolve().parents[1] / "Clarify_AI_QA_Workbook_COMPLETE_v2.xlsx"
OUT_FULL = Path(__file__).resolve().parents[1] / "Clarify_AI_QA_Workbook_FULL.xlsx"

# ── Tokens ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="0F172A")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="0F766E")
SUB_FONT = Font(name="Calibri", size=10, color="64748B", italic=True)
KPI_FILL = PatternFill("solid", fgColor="0F766E")
KPI_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
KPI_LABEL = Font(name="Calibri", size=9, color="334155")
CARD = PatternFill("solid", fgColor="F8FAFC")
ALT = PatternFill("solid", fgColor="F1F5F9")
PASS_F = PatternFill("solid", fgColor="86EFAC")
FAIL_F = PatternFill("solid", fgColor="FCA5A5")
PROG_F = PatternFill("solid", fgColor="FDE047")
GRAY_F = PatternFill("solid", fgColor="CBD5E1")
BLUE_F = PatternFill("solid", fgColor="93C5FD")
CRIT_F = PatternFill("solid", fgColor="F87171")
HIGH_F = PatternFill("solid", fgColor="FDBA74")
MED_F = PatternFill("solid", fgColor="FDE68A")
LOW_F = PatternFill("solid", fgColor="BFDBFE")
THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

TODAY = date(2026, 7, 29)

NAV_FILL = PatternFill("solid", fgColor="0F766E")
NAV_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
LINK_FONT = Font(name="Calibri", color="0F766E", underline="single", size=10)
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
BLOCK_FILL = PatternFill("solid", fgColor="FEE2E2")
OK_FILL = PatternFill("solid", fgColor="D1FAE5")

MODULES = [
    ("MOD-01", "Auth & Onboarding", "Auth Portal", "Email/OAuth login, MFA, 5-step onboarding", "Auth Lead", "Critical", "In Progress"),
    ("MOD-02", "Dashboard", "Candidate App", "Home KPIs, sessions, credits, upcoming interviews", "Product", "High", "Completed"),
    ("MOD-03", "Live Practice Coach", "Candidate App", "Deepgram STT + AI hints/answers + overlay", "Live Coach", "Critical", "Completed"),
    ("MOD-04", "Mock Interview", "Candidate App", "Timed mock sessions with AI scoring", "Live Coach", "High", "In Progress"),
    ("MOD-05", "Sessions & Debriefs", "Candidate App", "History, detail, AI debrief, scorecard share", "Sessions", "High", "Completed"),
    ("MOD-06", "Analytics & Usage", "Candidate App", "Trends, WPM, fillers, credit usage", "Analytics", "Medium", "Completed"),
    ("MOD-07", "Prep Lab", "Candidate App", "STAR, rephraser, polish, company Qs", "Prep", "High", "Completed"),
    ("MOD-08", "Documents & Resumes", "Candidate App", "Upload, parse-resume, gap analysis, OCR", "Docs", "High", "Completed"),
    ("MOD-09", "Answer Bank", "Candidate App", "STAR library + AI generate via prep-tool", "Prep", "Medium", "In Progress"),
    ("MOD-10", "Interviews & Calendar", "Candidate App", "Manual interviews + Google Calendar sync", "Scheduling", "High", "In Progress"),
    ("MOD-11", "Company Research", "Candidate App", "AI company brief + per-user cache", "Prep", "Medium", "Completed"),
    ("MOD-12", "Notifications & Profile", "Candidate App", "Realtime notifications, XP/streak/badges", "Growth", "Medium", "In Progress"),
    ("MOD-13", "Referrals & Guide", "Candidate App", "Referral credits + in-app guide", "Growth", "Low", "Completed"),
    ("MOD-14", "Gov Exam Mock Tests", "Candidate App", "India gov/competitive MCQ engine", "GovExams", "Critical", "In Progress"),
    ("MOD-15", "Settings", "Candidate App", "Billing, audio, models, security, integrations", "Platform", "High", "In Progress"),
    ("MOD-16", "Admin Portal", "Admin Portal", "Users, revenue, questions, flags, support", "Admin", "Critical", "In Progress"),
    ("MOD-17", "Marketing Site", "Marketing Portal", "Landing, pricing, blog, help, legal", "Growth", "Medium", "Completed"),
    ("MOD-18", "Billing & Credits", "Candidate App", "Stripe/Razorpay, packs, ledger, plan gates", "Billing", "Critical", "Completed"),
    ("MOD-19", "Electron Desktop Overlay", "Electron Overlay", "Always-on-top overlay, hotkeys, window state", "Desktop", "Critical", "In Progress"),
    ("MOD-20", "Security & Platform", "All Portals", "RLS, webhooks, rate limits, CI, observability", "Platform", "Critical", "In Progress"),
]

FEATURE_COUNTS = feature_counts_by_module()
# Ensure every MODULES entry has a count even if inventory gaps
for mid, name, *_rest in MODULES:
    FEATURE_COUNTS.setdefault(name, (1, 0))
# Map by module name (inventory uses names, Module Master uses names in col B)
FEATURE_COUNTS_BY_ID = {}
for mid, name, portal, desc, owner, pri, status in MODULES:
    t, d = FEATURE_COUNTS.get(name, (3, 1))
    FEATURE_COUNTS_BY_ID[mid] = (t, d)

PORTALS = ["Marketing Portal", "Auth Portal", "Candidate App", "Admin Portal", "Electron Overlay", "All Portals"]
ROLES = ["Guest", "Free User", "Pro User", "Max User", "Admin"]
OWNERS = ["Auth Lead", "Product", "Live Coach", "Sessions", "Analytics", "Prep", "Docs", "Scheduling", "Growth", "GovExams", "Platform", "Admin", "Billing", "Desktop", "QA Lead", "Dev Lead"]

# Production QA roster — share with these engineers
QA_TEAM = [
    # name, role, email, capacity hrs/week, skills
    ("Shreya Patil", "QA Engineer", "shreya.patil@clarify.ai", 40,
     "Practice Coach, Overlay, Mock Interview, Sessions/Debrief, Documents, UI/UX, Electron"),
    ("Raj Balani", "QA Engineer", "raj.balani@clarify.ai", 40,
     "Auth/Onboarding, Billing/Credits, Gov Exam Mock Tests, Admin, Prep Lab, Security, Regression"),
]
TESTERS = [t[0] for t in QA_TEAM]

# Developer roster — bugs Assigned To + Dev Tasks board
DEV_TEAM = [
    # name, role, email, capacity hrs/week, focus
    ("Dev Lead", "Engineering Lead", "dev.lead@clarify.ai", 40, "Triage, release, blockers"),
    ("Live Coach Dev", "Frontend Engineer", "live.coach@clarify.ai", 40, "Practice Coach, Overlay, Mock"),
    ("Platform Dev", "Full-Stack Engineer", "platform@clarify.ai", 40, "Auth, Billing, Admin, Edge"),
    ("GovExams Dev", "Full-Stack Engineer", "govexams@clarify.ai", 40, "Mock tests, question bank"),
    ("Desktop Dev", "Electron Engineer", "desktop@clarify.ai", 40, "Electron overlay, hotkeys"),
]
DEVELOPERS = [d[0] for d in DEV_TEAM]

# Module → default QA owner for task seed rows
QA_MODULE_OWNER = {
    "Auth & Onboarding": "Raj Balani",
    "Dashboard": "Shreya Patil",
    "Live Practice Coach": "Shreya Patil",
    "Mock Interview": "Shreya Patil",
    "Sessions & Debriefs": "Shreya Patil",
    "Analytics & Usage": "Raj Balani",
    "Prep Lab": "Raj Balani",
    "Documents & Resumes": "Shreya Patil",
    "Answer Bank": "Raj Balani",
    "Interviews & Calendar": "Shreya Patil",
    "Company Research": "Raj Balani",
    "Notifications & Profile": "Raj Balani",
    "Referrals & Guide": "Raj Balani",
    "Gov Exam Mock Tests": "Raj Balani",
    "Settings": "Raj Balani",
    "Admin Portal": "Raj Balani",
    "Marketing Site": "Shreya Patil",
    "Billing & Credits": "Raj Balani",
    "Electron Desktop Overlay": "Shreya Patil",
    "Security & Platform": "Raj Balani",
}


def lr(wb, name: str) -> str:
    return wb._list_ranges[name]


def progress_bar_formula(cell_ref: str) -> str:
    """Safe REPT progress bar — never negative length, never #VALUE!."""
    return (
        f'=IFERROR(REPT("#",MAX(0,MIN(10,ROUND(N({cell_ref})*10,0))))'
        f'&REPT("-",10-MAX(0,MIN(10,ROUND(N({cell_ref})*10,0)))),REPT("-",10))'
    )


def build_qa_team(wb: Workbook):
    ws = wb.create_sheet("QA Team")
    headers = [
        "QA Person", "Role", "Email", "Capacity Hrs/Week", "Skills Focus",
        "Assigned Cases", "Est Hours Assigned", "Actual Hours Logged", "Utilization %", "Status", "Remarks",
    ]
    rows = []
    for name, role, email, cap, skills in QA_TEAM:
        rows.append([
            name, role, email, cap, skills,
            f'=COUNTIF(\'04 Test Case Repository\'!S:S,A{{ROW}})',  # placeholder fixed below
            None, None, None, "Active", "",
        ])
    # Write with per-row formulas after
    write_table(ws, "QA Team Roster — Clarify AI", "", headers, [[r[0], r[1], r[2], r[3], r[4], None, None, None, None, r[9], r[10]] for r in rows])
    for i, (name, role, email, cap, skills) in enumerate(QA_TEAM):
        r = 3 + i
        ws.cell(r, 6, f"=COUNTIF('04 Test Case Repository'!S$3:S$5000,A{r})")
        ws.cell(r, 7, f"=SUMIF('04 Test Case Repository'!S$3:S$5000,A{r},'04 Test Case Repository'!T$3:T$5000)")
        ws.cell(r, 8, f"=SUMIF('04 Test Case Repository'!S$3:S$5000,A{r},'04 Test Case Repository'!U$3:U$5000)")
        ws.cell(r, 9, f"=IFERROR(H{r}/D{r},0)")
        ws.cell(r, 9).number_format = "0.0%"
        for c in (6, 7, 8, 9):
            ws.cell(r, c).protection = LOCKED
    add_dv(ws, lr(wb, "Status"), "J3:J50")
    # Totals row
    tot = 3 + len(QA_TEAM)
    ws.cell(tot, 1, "TOTAL")
    ws.cell(tot, 1).font = Font(bold=True)
    ws.cell(tot, 4, f"=SUM(D3:D{tot-1})")
    ws.cell(tot, 6, f"=SUM(F3:F{tot-1})")
    ws.cell(tot, 7, f"=SUM(G3:G{tot-1})")
    ws.cell(tot, 8, f"=SUM(H3:H{tot-1})")
    ws.cell(tot, 9, f"=IFERROR(H{tot}/D{tot},0)")
    ws.cell(tot, 9).number_format = "0.0%"
    for c in range(1, 12):
        ws.cell(tot, c).fill = KPI_FILL
        ws.cell(tot, c).font = Font(bold=True, color="FFFFFF")
    finish_sheet(ws, 2, len(rows) + 1, len(headers))
    return ws


def style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def zebra(ws, start: int, end: int, ncols: int):
    for r in range(start, end + 1):
        if (r - start) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                if cell.fill.fgColor is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = ALT
                elif getattr(cell.fill.fgColor, "rgb", None) in ("00000000", "00FFFFFF", "FFFFFF"):
                    cell.fill = ALT


def apply_borders(ws, start: int, end: int, ncols: int):
    for r in range(start, end + 1):
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
            ws.cell(r, c).alignment = WRAP
            ws.cell(r, c).protection = UNLOCKED


def autosize(ws, header_row=2, max_w=36):
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        best = 12
        for r in range(header_row, min(ws.max_row, header_row + 25) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                best = max(best, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = best


def add_dv(ws, formula: str, cells: str):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = "Pick a value from the list"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(cells)


def cf_verdict(ws, col: str, start=3, end=500):
    rng = f"{col}{start}:{col}{end}"
    for val, fill in [
        ("Pass", PASS_F), ("PASS", PASS_F), ("Yes", PASS_F),
        ("Fail", FAIL_F), ("FAIL", FAIL_F), ("No", FAIL_F),
        ("In Progress", PROG_F), ("Blocked", PROG_F), ("Partial", PROG_F),
        ("Not Started", GRAY_F), ("Not Run", GRAY_F), ("Not Tested", GRAY_F), ("N/A", GRAY_F),
        ("Completed", BLUE_F), ("Complete", BLUE_F), ("Passed", BLUE_F), ("Ready", BLUE_F),
    ]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill))


def cf_severity(ws, col: str, start=3, end=500):
    rng = f"{col}{start}:{col}{end}"
    for val, fill in [("Critical", CRIT_F), ("High", HIGH_F), ("Medium", MED_F), ("Low", LOW_F)]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill))


def finish_sheet(ws, header_row: int, data_rows: int, ncols: int, filter=True, freeze=True):
    end = header_row + data_rows
    apply_borders(ws, header_row, end, ncols)
    zebra(ws, header_row + 1, end, ncols)
    if freeze:
        ws.freeze_panes = f"A{header_row + 1}"
    if filter and data_rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{end}"
    autosize(ws, header_row)
    ws.sheet_view.showGridLines = False


def add_nav_bar(ws, current: str):
    """Power BI–style clickable navigation strip (row 1, from col J)."""
    hub = [
        ("01 Dashboard", "Dashboard"),
        ("NAV Hub", "NAV"),
        ("03 Feature Inventory", "Features"),
        ("06 Bug Tracker", "Bugs"),
        ("24 Launch Status", "Launch"),
        ("25 Smoke Pack", "Smoke"),
        ("21 Test Credentials", "Creds"),
        ("22 Environments", "Env"),
        ("23 Module Playbooks", "Playbooks"),
        ("04 Test Case Repository", "Cases"),
        ("20 Release Sign-Off", "Sign-Off"),
    ]
    col = 10
    for sheet, label in hub:
        if sheet == current:
            cell = ws.cell(1, col, f"● {label}")
            cell.fill = NAV_FILL
            cell.font = NAV_FONT
        else:
            cell = ws.cell(1, col, label)
            cell.hyperlink = f"#'{sheet}'!A1"
            cell.font = LINK_FONT
        cell.alignment = CENTER
        col += 1


def bug_status_counts():
    from collections import Counter
    return Counter(b[11] for b in BUGS)


def feature_status_counts():
    from collections import Counter
    return Counter(f[9] for f in FEATURES)


def write_table(ws, title: str, subtitle: str, headers: list, rows: list, header_row: int = 2):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, len(headers)))
    if subtitle:
        pass
    for i, h in enumerate(headers, 1):
        ws.cell(header_row, i, h)
    style_header(ws, header_row, len(headers))
    for ri, row in enumerate(rows, header_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    return header_row + len(rows)


# ═══════════════════════════════════════════════════════════════════════════
def build_lists(wb: Workbook):
    ws = wb.create_sheet("Lists")
    lists = {
        "A": ("Status", ["Not Started", "In Progress", "Completed", "Blocked", "Deferred", "Ready", "Complete", "Active"]),
        "B": ("PassFail", ["Pass", "Fail", "Blocked", "Not Run", "N/A", "In Progress"]),
        "C": ("Priority", ["Critical", "High", "Medium", "Low"]),
        "D": ("Severity", ["Critical", "High", "Medium", "Low"]),
        "E": ("Environment", ["Local", "Staging", "Production", "CI", "Electron Dev", "Electron Packaged"]),
        "F": ("Browser", ["Chrome", "Edge", "Firefox", "Safari", "Electron", "N/A"]),
        "G": ("Device", ["Desktop", "Laptop", "iPhone", "Android Phone", "iPad", "Android Tablet"]),
        "H": ("Platform", ["Web", "Windows", "macOS", "Linux", "iOS", "Android"]),
        "I": ("YesNo", ["Yes", "No", "Partial", "N/A"]),
        "J": ("QAStatus", ["Not Tested", "In Testing", "Passed", "Failed", "Blocked", "Deferred"]),
        "K": ("BugStatus", ["New", "Open", "In Progress", "Fixed", "Retest", "Closed", "Won't Fix", "Duplicate"]),
        "L": ("Risk", ["Critical", "High", "Medium", "Low"]),
        "M": ("Criticality", ["Critical", "High", "Medium", "Low"]),
        "N": ("Automation", ["Manual", "Automated", "Candidate", "Partial"]),
        "O": ("SecurityResult", ["Pass", "Fail", "Partial", "Not Tested", "N/A"]),
        "P": ("ReadyStatus", ["Ready", "Not Ready", "Conditional", "Blocked"]),
        "Q": ("Decision", ["Go", "No-Go", "Conditional Go"]),
        "R": ("Approval", ["Approved", "Pending", "Rejected", "Deferred"]),
        "S": ("PerfRating", ["Excellent", "Good", "Acceptable", "Poor", "Critical"]),
        "T": ("HttpMethod", ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"]),
        "U": ("StatusCode", ["200", "201", "204", "400", "401", "402", "403", "404", "409", "429", "500", "501"]),
        "V": ("AuthType", ["None", "Bearer JWT", "Anon Key", "Service Role", "Webhook HMAC", "API Key Header"]),
        "W": ("Sprint", ["Sprint 24", "Sprint 25", "Sprint 26", "Sprint 27", "Hotfix"]),
        "X": ("Release", ["1.0.0-beta", "1.0.0", "1.0.1", "1.1.0"]),
        "Y": ("RootCause", ["Logic Bug", "Missing Validation", "Config/Env", "Race Condition", "UI Regression", "API Contract", "DB/RLS", "Third-party", "Design Gap", "Unknown"]),
        "Z": ("Modules", [m[1] for m in MODULES] + ["Cross-Cutting", "All Portals"]),
        "AA": ("Portals", PORTALS),
        "AB": ("Roles", ROLES),
        "AC": ("Testers", TESTERS),
        "AD": ("Owners", OWNERS),
        "AE": ("ImplStatus", ["Fully Implemented", "Partial", "Stub", "Dead Code", "Not Built"]),
        "AF": ("Coverage", ["Yes", "No", "Partial", "N/A"]),
        "AG": ("ModuleStatus", ["Not Started", "In Progress", "Completed", "Blocked", "Deprecated"]),
        "AH": ("Developers", DEVELOPERS),
        "AI": ("TaskType", ["Smoke", "Functional", "Regression", "E2E", "UI/UX", "API", "Security", "Retest", "Exploratory"]),
        "AJ": ("TaskStatus", ["Not Started", "In Progress", "Blocked", "Done", "Deferred"]),
        "AK": ("DevTaskStatus", ["Queued", "In Progress", "Code Review", "Fixed — Ready for Retest", "Reopened", "Done", "Won't Fix"]),
        "AL": ("Assignees", DEVELOPERS + OWNERS),
    }
    wb._list_ranges = {}
    for col, (title, values) in lists.items():
        ws[f"{col}1"] = title
        ws[f"{col}1"].font = HEADER_FONT
        ws[f"{col}1"].fill = HEADER_FILL
        for i, v in enumerate(values, 2):
            ws[f"{col}{i}"] = v
        wb._list_ranges[title] = f"Lists!${col}$2:${col}${1 + len(values)}"
        ws.column_dimensions[col].width = max(14, max(len(str(x)) for x in [title] + values) + 2)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    return ws


def build_daily_log(wb: Workbook):
    """Helper sheet feeding Daily Progress chart on Dashboard."""
    ws = wb.create_sheet("Daily Log")
    headers = [
        "Date", "QA Person", "Cases Executed", "Passed", "Failed", "Blocked",
        "Bugs Opened", "Bugs Closed", "Hours Logged", "Remarks",
    ]
    people = ["Shreya Patil", "Raj Balani", "Shreya Patil", "Raj Balani", "Shreya Patil", "Raj Balani", "Shreya Patil"]
    rows = []
    for i in range(7):
        d = TODAY - timedelta(days=6 - i)
        rows.append([
            d, people[i], 8 + i * 2, 6 + i, 1 if i % 3 == 0 else 0, 1 if i == 4 else 0,
            1 if i % 2 == 0 else 0, 1 if i > 3 else 0, round(4.5 + i * 0.5, 1),
            "Sprint execution — production QA",
        ])
    write_table(ws, "Daily Progress Log (feeds Dashboard chart)", "", headers, rows, 2)
    add_dv(ws, lr(wb, "Testers"), "B3:B5000")
    # Totals
    t = 3 + len(rows)
    ws.cell(t, 1, "TOTAL")
    ws.cell(t, 1).font = Font(bold=True)
    for col, letter in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I")]:
        ws.cell(t, col, f"=SUM({letter}3:{letter}{t-1})")
        ws.cell(t, col).font = Font(bold=True)
    finish_sheet(ws, 2, len(rows) + 1, len(headers))
    return ws


def build_qa_tasks(wb: Workbook):
    """QA task board for Shreya Patil & Raj Balani — execution work queue."""
    ws = wb.create_sheet("QA Tasks")
    headers = [
        "Task ID", "QA Owner", "Module", "Task Type", "Title", "Linked TC / Flow",
        "Priority", "Status", "Due Date", "Est Hours", "Actual Hours",
        "Linked Bug ID", "Fails Linked (auto)", "Notes",
    ]
    # Seed one primary pack task per module + shared smoke/regression
    rows = []
    tid = 1
    for mid, name, _portal, desc, _owner, pri, _status in MODULES:
        owner = QA_MODULE_OWNER.get(name, "Shreya Patil")
        rows.append([
            f"QAT-{tid:03d}", owner, name, "Functional",
            f"Execute {name} pack — {desc[:48]}",
            f"See 04 / domain sheets for {mid}",
            pri if pri in ("Critical", "High", "Medium", "Low") else "High",
            "Not Started", TODAY + timedelta(days=7 + (tid % 5)),
            4.0, 0, "",
            None,  # auto formula filled below
            "Mark Fail on TC → create Bug in 06 → Bug ID appears here & on Dev Tasks",
        ])
        tid += 1
    # Shared packs
    for title, owner, ttype, pri, hours in [
        ("Smoke Pack — 25 Smoke Pack must-pass", "Shreya Patil", "Smoke", "Critical", 3.0),
        ("Smoke Pack — 25 Smoke Pack must-pass (pair)", "Raj Balani", "Smoke", "Critical", 3.0),
        ("Regression — 18 Regression Testing pack", "Raj Balani", "Regression", "High", 6.0),
        ("UI/UX pass — 07 UI-UX Testing", "Shreya Patil", "UI/UX", "High", 4.0),
        ("Mobile responsiveness — 08", "Shreya Patil", "UI/UX", "High", 3.0),
        ("Security pack — 12 Security Testing", "Raj Balani", "Security", "Critical", 4.0),
        ("API pack — 09 API Testing", "Raj Balani", "API", "High", 4.0),
        ("E2E golden paths — 05 E2E User Flows", "Shreya Patil", "E2E", "Critical", 5.0),
        ("Retest queue — bugs in Retest status", "Raj Balani", "Retest", "Critical", 3.0),
        ("Retest queue — bugs in Retest status (pair)", "Shreya Patil", "Retest", "Critical", 3.0),
    ]:
        rows.append([
            f"QAT-{tid:03d}", owner, "Cross-Cutting",
            ttype, title, title.split("—")[-1].strip() if "—" in title else "",
            pri, "Not Started", TODAY + timedelta(days=5), hours, 0, "",
            None,
            "Update Status daily; link Bug ID when filing defects",
        ])
        tid += 1

    write_table(
        ws,
        "QA Tasks — Shreya Patil & Raj Balani work queue",
        "When a TC Fails: set Pass/Fail=Fail on 04 → create row on 06 Bug Tracker (Reported By = you) → paste Bug ID here. Dev Tasks auto-mirrors open bugs.",
        headers,
        [[*r[:12], None, r[13]] for r in rows],
    )
    for i in range(len(rows)):
        r = 3 + i
        # Count open bugs reported by this task's QA owner (reflective KPI per row owner)
        ws.cell(r, 13, f"=IFERROR(COUNTIFS('06 Bug Tracker'!N$3:N$5000,B{r},'06 Bug Tracker'!L$3:L$5000,\"<>Closed\",'06 Bug Tracker'!L$3:L$5000,\"<>Won't Fix\",'06 Bug Tracker'!L$3:L$5000,\"<>Duplicate\"),0)")
        ws.cell(r, 13).protection = LOCKED
    add_dv(ws, lr(wb, "Testers"), "B3:B5000")
    add_dv(ws, lr(wb, "Modules"), "C3:C5000")
    add_dv(ws, lr(wb, "TaskType"), "D3:D5000")
    add_dv(ws, lr(wb, "Priority"), "G3:G5000")
    add_dv(ws, lr(wb, "TaskStatus"), "H3:H5000")
    cf_severity(ws, "G")
    cf_verdict(ws, "H")
    # Summary strip
    tot = 3 + len(rows)
    ws.cell(tot, 1, "TOTAL / KPIs")
    ws.cell(tot, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(tot, 10, f"=SUM(J3:J{tot-1})")
    ws.cell(tot, 11, f"=SUM(K3:K{tot-1})")
    ws.cell(tot, 13, f"=SUM(M3:M{tot-1})")
    for c in range(1, 15):
        ws.cell(tot, c).fill = KPI_FILL
        ws.cell(tot, c).font = Font(bold=True, color="FFFFFF")
    # Owner split legend
    ws.cell(tot + 2, 1, "Shreya open bugs (Reported By)")
    ws.cell(tot + 2, 2, "=IFERROR(COUNTIFS('06 Bug Tracker'!N3:N5000,\"Shreya Patil\",'06 Bug Tracker'!L3:L5000,\"<>Closed\",'06 Bug Tracker'!L3:L5000,\"<>Won't Fix\",'06 Bug Tracker'!L3:L5000,\"<>Duplicate\"),0)")
    ws.cell(tot + 3, 1, "Raj open bugs (Reported By)")
    ws.cell(tot + 3, 2, "=IFERROR(COUNTIFS('06 Bug Tracker'!N3:N5000,\"Raj Balani\",'06 Bug Tracker'!L3:L5000,\"<>Closed\",'06 Bug Tracker'!L3:L5000,\"<>Won't Fix\",'06 Bug Tracker'!L3:L5000,\"<>Duplicate\"),0)")
    ws.cell(tot + 4, 1, "Failed TCs (04 Pass/Fail=Fail)")
    ws.cell(tot + 4, 2, "=IFERROR(COUNTIF('04 Test Case Repository'!P3:P5000,\"Fail\"),0)")
    finish_sheet(ws, 2, len(rows) + 1, len(headers))
    return ws


def build_dev_team(wb: Workbook):
    ws = wb.create_sheet("Dev Team")
    headers = [
        "Developer", "Role", "Email", "Capacity Hrs/Week", "Focus Area",
        "Open Bugs Assigned", "In Progress", "Ready for Retest", "Utilization Hint", "Status", "Remarks",
    ]
    write_table(ws, "Dev Team Roster — Clarify AI", "Assigned bugs come from 06 Bug Tracker → Assigned To. Keep names in sync with Lists!Developers.", headers, [])
    for i, (name, role, email, cap, focus) in enumerate(DEV_TEAM):
        r = 3 + i
        ws.cell(r, 1, name)
        ws.cell(r, 2, role)
        ws.cell(r, 3, email)
        ws.cell(r, 4, cap)
        ws.cell(r, 5, focus)
        # Open = not Closed / Won't Fix / Duplicate
        ws.cell(r, 6, f"=IFERROR(COUNTIFS('06 Bug Tracker'!M$3:M$5000,A{r},'06 Bug Tracker'!L$3:L$5000,\"<>Closed\",'06 Bug Tracker'!L$3:L$5000,\"<>Won't Fix\",'06 Bug Tracker'!L$3:L$5000,\"<>Duplicate\"),0)")
        ws.cell(r, 7, f"=IFERROR(COUNTIFS('06 Bug Tracker'!M$3:M$5000,A{r},'06 Bug Tracker'!L$3:L$5000,\"In Progress\"),0)")
        ws.cell(r, 8, f"=IFERROR(COUNTIFS('06 Bug Tracker'!M$3:M$5000,A{r},'06 Bug Tracker'!L$3:L$5000,\"Retest\")+COUNTIFS('06 Bug Tracker'!M$3:M$5000,A{r},'06 Bug Tracker'!L$3:L$5000,\"Fixed\"),0)")
        ws.cell(r, 9, f'=IF(F{r}>8,"Overloaded",IF(F{r}>4,"Busy","OK"))')
        ws.cell(r, 10, "Active")
        ws.cell(r, 11, "")
        for c in (6, 7, 8, 9):
            ws.cell(r, c).protection = LOCKED
    add_dv(ws, lr(wb, "Status"), "J3:J50")
    n = len(DEV_TEAM)
    tot = 3 + n
    ws.cell(tot, 1, "TOTAL")
    ws.cell(tot, 4, f"=SUM(D3:D{tot-1})")
    ws.cell(tot, 6, f"=SUM(F3:F{tot-1})")
    ws.cell(tot, 7, f"=SUM(G3:G{tot-1})")
    ws.cell(tot, 8, f"=SUM(H3:H{tot-1})")
    for c in range(1, 12):
        ws.cell(tot, c).fill = KPI_FILL
        ws.cell(tot, c).font = Font(bold=True, color="FFFFFF")
    finish_sheet(ws, 2, n + 1, len(headers))
    return ws


def build_dev_tasks(wb: Workbook):
    """
    Developer work queue that MIRRORS 06 Bug Tracker.
    When QA files/updates a bug, the matching row here updates via formulas.
    Editable columns: Dev Status, Fix ETA, Build Fixed, Retest Owner, Dev Notes.
    """
    ws = wb.create_sheet("Dev Tasks")
    headers = [
        "Bug ID", "Title", "Module", "Severity", "Priority", "Bug Status",
        "Reported By (QA)", "Assigned Developer", "Needs Dev Action",
        "Dev Status", "Fix ETA", "Build Fixed", "Retest Owner (QA)",
        "Retest Result", "Dev Notes",
    ]
    # Mirror up to 150 bug rows (formulas) — empty source rows stay blank
    mirror_n = 150
    write_table(
        ws,
        "Dev Tasks — live mirror of 06 Bug Tracker",
        "SOURCE OF TRUTH = 06 Bug Tracker. Columns A–H are formula-linked. When QA sets Status=New/Open/In Progress/Fixed/Retest, Needs Dev Action=Yes. Dev updates Dev Status → QA retests → Bug Status=Closed.",
        headers,
        [],
    )
    for i in range(mirror_n):
        r = 3 + i
        src = 3 + i
        # A–H mirror Bug Tracker: A Title B Desc C Module … J Severity K Priority L Status M Assigned N Reported
        # Bug headers: A Bug ID, B Title, C Description, D Module, ... J Severity, K Priority, L Status, M Assigned To, N Reported By
        ws.cell(r, 1, f"='06 Bug Tracker'!A{src}")
        ws.cell(r, 2, f"='06 Bug Tracker'!B{src}")
        ws.cell(r, 3, f"='06 Bug Tracker'!D{src}")
        ws.cell(r, 4, f"='06 Bug Tracker'!J{src}")
        ws.cell(r, 5, f"='06 Bug Tracker'!K{src}")
        ws.cell(r, 6, f"='06 Bug Tracker'!L{src}")
        ws.cell(r, 7, f"='06 Bug Tracker'!N{src}")
        ws.cell(r, 8, f"='06 Bug Tracker'!M{src}")
        ws.cell(r, 9,
                f'=IF(A{r}="","",IF(OR(F{r}="New",F{r}="Open",F{r}="In Progress",F{r}="Fixed",F{r}="Retest"),"Yes","No"))')
        # Editable defaults
        ws.cell(r, 10, f'=IF(A{r}="","",IF(F{r}="Retest","Fixed — Ready for Retest",IF(F{r}="Fixed","Fixed — Ready for Retest",IF(F{r}="In Progress","In Progress",IF(OR(F{r}="New",F{r}="Open"),"Queued",IF(F{r}="Closed","Done",IF(F{r}="Won\'t Fix","Won\'t Fix","")))))))')
        ws.cell(r, 11, "")
        ws.cell(r, 12, f"='06 Bug Tracker'!R{src}")  # Build
        ws.cell(r, 13, f"='06 Bug Tracker'!N{src}")  # Retest owner defaults to reporter
        ws.cell(r, 14, "")
        ws.cell(r, 15, "")
        for c in range(1, 10):
            ws.cell(r, c).protection = LOCKED
    add_dv(ws, lr(wb, "DevTaskStatus"), "J3:J5000")
    add_dv(ws, lr(wb, "Testers"), "M3:M5000")
    add_dv(ws, lr(wb, "PassFail"), "N3:N5000")
    cf_severity(ws, "D")
    cf_severity(ws, "E")
    cf_verdict(ws, "N")
    # KPI block above filter end
    kpi_row = 3 + mirror_n + 1
    ws.cell(kpi_row, 1, "QUEUE KPIs (auto)")
    ws.cell(kpi_row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(kpi_row, 1).fill = KPI_FILL
    labels = [
        (kpi_row + 1, "Needs Dev Action = Yes", '=IFERROR(COUNTIF(I3:I152,"Yes"),0)'),
        (kpi_row + 2, "Reported by Shreya (open)", '=IFERROR(COUNTIFS(G3:G152,"Shreya Patil",I3:I152,"Yes"),0)'),
        (kpi_row + 3, "Reported by Raj (open)", '=IFERROR(COUNTIFS(G3:G152,"Raj Balani",I3:I152,"Yes"),0)'),
        (kpi_row + 4, "Ready for Retest", '=IFERROR(COUNTIF(J3:J152,"Fixed — Ready for Retest"),0)'),
        (kpi_row + 5, "HOW TO USE", "1) QA fails TC on 04  2) QA adds bug on 06 (Reported By=Shreya/Raj, Assigned To=dev)  3) Row appears here  4) Dev fixes → set Bug Status=Fixed/Retest on 06  5) QA retests → Closed"),
    ]
    for row, label, val in labels:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    finish_sheet(ws, 2, mirror_n, len(headers))
    # Note: auto_filter on 150 formula rows is OK — users filter Needs Dev Action=Yes
    return ws


# ═══════════════════════════════════════════════════════════════════════════
def build_readme(wb: Workbook):
    ws = wb.create_sheet("00 Read Me", 0)
    ws["A1"] = "Clarify AI — Enterprise QA Testing Workbook"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        ("", ""),
        ("Application", "Clarify AI"),
        ("Description", "AI-powered interview preparation & rehearsal platform with Indian gov-exam MCQ mock tests, live coaching overlay, and dual Stripe/Razorpay billing."),
        ("Tech stack", "React 18 + TypeScript + Vite + Tailwind/shadcn · Supabase (Auth, Postgres+RLS, Edge/Deno, Storage, Realtime) · Electron · Deepgram · Gemini/OpenAI/Anthropic · Stripe + Razorpay"),
        ("Portals", "Marketing · Auth · Candidate App · Admin · Electron Overlay"),
        ("Roles", "Guest · Free User · Pro User · Max User · Admin"),
        ("Version under test", "1.0.0 / 1.0.0-beta"),
        ("Generated", str(TODAY)),
        ("", ""),
        ("HOW TO USE", ""),
        ("1", "Start at NAV Hub or 01 Dashboard — KPIs + charts."),
        ("2", "QA opens QA Tasks — pick your rows (Shreya / Raj). Execute linked TCs on 04."),
        ("3", "On Fail: set Pass/Fail=Fail on 04 → add a row on 06 Bug Tracker (Reported By = you, Assigned To = a Developer)."),
        ("4", "Dev Tasks auto-mirrors 06 — developers see Needs Dev Action=Yes. Filter that column."),
        ("5", "Dev sets Bug Status=Fixed or Retest on 06 → QA retests → Closed. Fill Daily Log daily."),
        ("6", "Complete 19 Production Readiness and 20 Release Sign-Off for Go/No-Go."),
        ("7", "Slicer tip: AutoFilter on Bug Tracker / Dev Tasks / QA Tasks (Status, Owner, Priority)."),
        ("", ""),
        ("QA TEAM (production)", ""),
        ("", "Shreya Patil — Practice Coach, Overlay, Mock, Sessions/Debrief, Documents, UI/UX, Electron"),
        ("", "Raj Balani — Auth, Billing, Gov Exams, Admin, Prep Lab, Security, Regression"),
        ("DEV TEAM", ""),
        ("", "Dev Lead · Live Coach Dev · Platform Dev · GovExams Dev · Desktop Dev — see Dev Team sheet"),
        ("Green", "Pass / Yes / Closed / Done"),
        ("Red", "Fail / No / Critical / Blocked"),
        ("Yellow", "In Progress / Open / Medium / Queued"),
        ("Gray", "Not Started / Not Run"),
        ("Blue", "Completed / Passed / Ready"),
        ("", ""),
        ("SCALING", ""),
        ("", "One row per entity; filters + freeze panes. Dashboard / Dev Tasks formulas scan to row 5000 / 152."),
        ("Google Sheets", "Upload .xlsx to Drive → Open with Google Sheets. Hyperlinks transfer; confirm Lists ranges after import."),
        ("", ""),
        ("SHEET INDEX (click sheet tabs or use NAV Hub)", ""),
        ("NAV Hub", "Clickable map of every sheet — start here"),
        ("01 Dashboard", "Executive KPIs + charts (Power BI style)"),
        ("QA Tasks", "Shreya & Raj execution queue"),
        ("Dev Team", "Developer roster + open-bug counts"),
        ("Dev Tasks", "Live mirror of Bug Tracker for engineers"),
        ("02 Module Master", "20 Clarify modules"),
        ("03 Feature Inventory", "114 features + routes + how-it-works"),
        ("04 Test Case Repository", "Executable cases"),
        ("05 E2E User Flows", "Cross-module journeys"),
        ("06 Bug Tracker", "Defect lifecycle — source of truth"),
        ("07 UI-UX Testing", "Visual / a11y"),
        ("08 Mobile Responsiveness", "Breakpoints"),
        ("09 API Testing", "Edge functions"),
        ("10 Database Validation", "Schema / RLS"),
        ("11 Authentication Testing", "Auth matrix + MFA"),
        ("12 Security Testing", "OWASP-style"),
        ("13 Performance Testing", "Latency budgets"),
        ("14 Live Coaching Testing", "Domain: live coach + overlay"),
        ("15 Gov Exam Mock Testing", "Domain: MCQ engine"),
        ("16 Billing Credits Testing", "Domain: Stripe/Razorpay/credits"),
        ("17 Role Based Testing", "Plan + admin RBAC"),
        ("18 Regression Testing", "Release regression pack"),
        ("19 Production Readiness", "Go-live checklist"),
        ("20 Release Sign-Off", "Formal approval"),
        ("21 Test Credentials", "Placeholder QA accounts"),
        ("22 Environments", "URLs + secrets checklist"),
        ("23 Module Playbooks", "Happy-path runbooks"),
        ("24 Launch Status", "Fixed / Open / Blocked counts"),
        ("25 Smoke Pack", "~20 must-pass cases"),
        ("QA Team", "Roster & capacity (Shreya / Raj)"),
        ("Daily Log", "Feeds daily progress chart"),
        ("Lists", "Dropdown vocabulary — do not delete"),
    ]
    for i, (a, b) in enumerate(rows, 3):
        ca = ws.cell(i, 1, a)
        cb = ws.cell(i, 2, b)
        ca.font = Font(bold=True)
        cb.alignment = WRAP
        # Hyperlink sheet index entries
        if a and a not in ("", "HOW TO USE", "QA TEAM", "SCALING", "SHEET INDEX (click sheet tabs or use NAV Hub)", "Green", "Red", "Yellow", "Gray", "Blue") and not a[0].isdigit() and a not in ("Application", "Description", "Tech stack", "Portals", "Roles", "Version under test", "Generated", "Google Sheets"):
            # leave non-sheet labels
            pass
        sheet_targets = {
            "NAV Hub", "01 Dashboard", "QA Tasks", "Dev Team", "Dev Tasks",
            "02 Module Master", "03 Feature Inventory", "04 Test Case Repository",
            "05 E2E User Flows", "06 Bug Tracker", "07 UI-UX Testing", "08 Mobile Responsiveness",
            "09 API Testing", "10 Database Validation", "11 Authentication Testing", "12 Security Testing",
            "13 Performance Testing", "14 Live Coaching Testing", "15 Gov Exam Mock Testing",
            "16 Billing Credits Testing", "17 Role Based Testing", "18 Regression Testing",
            "19 Production Readiness", "20 Release Sign-Off", "21 Test Credentials", "22 Environments",
            "23 Module Playbooks", "24 Launch Status", "25 Smoke Pack", "QA Team", "Daily Log", "Lists",
        }
        if a in sheet_targets:
            ca.hyperlink = f"#'{a}'!A1"
            ca.font = LINK_FONT
        if a == "Green":
            ca.fill = PASS_F
        elif a == "Red":
            ca.fill = FAIL_F
        elif a == "Yellow":
            ca.fill = PROG_F
        elif a == "Gray":
            ca.fill = GRAY_F
        elif a == "Blue":
            ca.fill = BLUE_F
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A3"
    add_nav_bar(ws, "00 Read Me")


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("01 Dashboard", 1)
    ws["A1"] = "Clarify AI — QA Executive Dashboard (Power BI style)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    add_nav_bar(ws, "01 Dashboard")
    ws["A2"] = (
        f"Synced {TODAY} · Formula KPIs + inventory snapshot · "
        "Use AutoFilter on Features/Bugs as slicers · Click NAV Hub for sheet map"
    )
    ws["A2"].font = SUB_FONT

    bsc = bug_status_counts()
    fsc = feature_status_counts()
    # Snapshot tiles (inventory truth — complements formula KPIs)
    ws["I3"] = "INVENTORY SNAPSHOT"
    ws["I3"].font = SECTION_FONT
    snap = [
        (4, "Features total", len(FEATURES), OK_FILL),
        (5, "Features Completed", fsc.get("Completed", 0), OK_FILL),
        (6, "Bugs Closed", bsc.get("Closed", 0), OK_FILL),
        (7, "Bugs Open", bsc.get("Open", 0), WARN_FILL),
        (8, "Bugs Blocked (ops)", bsc.get("Blocked", 0), BLOCK_FILL),
        (9, "Won't Fix / Descope", bsc.get("Won't Fix", 0), GRAY_F),
    ]
    for row, label, val, fill in snap:
        ws.cell(row, 9, label).font = KPI_LABEL
        ws.cell(row, 9).fill = CARD
        ws.cell(row, 9).border = THIN
        c = ws.cell(row, 10, val)
        c.fill = fill
        c.font = Font(name="Calibri", bold=True, size=14)
        c.alignment = CENTER
        c.border = THIN

    ws["I11"] = "Quick links"
    ws["I11"].font = SECTION_FONT
    for r, (sheet, label) in enumerate([
        ("NAV Hub", "Open NAV Hub"),
        ("QA Tasks", "QA Tasks (Shreya/Raj)"),
        ("Dev Tasks", "Dev Tasks (bug mirror)"),
        ("06 Bug Tracker", "Bug Tracker"),
        ("25 Smoke Pack", "Smoke Pack"),
        ("24 Launch Status", "Launch Status"),
    ], 12):
        cell = ws.cell(r, 9, label)
        cell.hyperlink = f"#'{sheet}'!A1"
        cell.font = LINK_FONT

    # KPI grid — all IFERROR-wrapped; hours from Test Case cols T/U
    kpis = [
        (4, "Total Modules", "=IFERROR(COUNTA('02 Module Master'!A3:A5000)-COUNTIF('02 Module Master'!A3:A5000,\"TOTAL\"),0)"),
        (5, "Total Features", "=IFERROR(COUNTA('03 Feature Inventory'!A3:A5000)-COUNTIF('03 Feature Inventory'!A3:A5000,\"TOTAL\"),0)"),
        (6, "Total Test Cases", "=IFERROR(COUNTA('04 Test Case Repository'!A3:A5000)-COUNTIF('04 Test Case Repository'!A3:A5000,\"TOTAL\"),0)"),
        (7, "Passed", "=IFERROR(COUNTIF('04 Test Case Repository'!P3:P5000,\"Pass\"),0)"),
        (8, "Failed", "=IFERROR(COUNTIF('04 Test Case Repository'!P3:P5000,\"Fail\"),0)"),
        (9, "Blocked", "=IFERROR(COUNTIF('04 Test Case Repository'!P3:P5000,\"Blocked\"),0)"),
        (10, "Not Executed", "=IFERROR(COUNTIF('04 Test Case Repository'!P3:P5000,\"Not Run\"),0)"),
        (11, "In Progress", "=IFERROR(COUNTIF('04 Test Case Repository'!Q3:Q5000,\"In Progress\"),0)"),
        (12, "Critical Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!J3:J5000,\"Critical\",'06 Bug Tracker'!L3:L5000,\"<>Closed\"),0)"),
        (13, "High Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!J3:J5000,\"High\",'06 Bug Tracker'!L3:L5000,\"<>Closed\"),0)"),
        (14, "Medium Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!J3:J5000,\"Medium\",'06 Bug Tracker'!L3:L5000,\"<>Closed\"),0)"),
        (15, "Low Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!J3:J5000,\"Low\",'06 Bug Tracker'!L3:L5000,\"<>Closed\"),0)"),
        (16, "Overall Pass %", "=IFERROR(IF((B7+B8+B9)=0,0,B7/(B7+B8+B9)),0)"),
        (17, "Overall Fail %", "=IFERROR(IF((B7+B8+B9)=0,0,B8/(B7+B8+B9)),0)"),
        (18, "QA Progress %", "=IFERROR(IF(B6=0,0,(B7+B8+B9+B11)/B6),0)"),
        (19, "Production Readiness %", "=IFERROR(AVERAGEIF('19 Production Readiness'!A3:A5000,\"<>\",'19 Production Readiness'!R3:R5000),0)"),
        (20, "Est Hours (Total)", "=IFERROR(SUM('04 Test Case Repository'!T3:T5000),0)"),
        (21, "Actual Hours (Total)", "=IFERROR(SUM('04 Test Case Repository'!U3:U5000),0)"),
        (22, "Hours Variance", "=IFERROR(B21-B20,0)"),
        (23, "QA Team Size", "=IFERROR(COUNTA('QA Team'!A3:A50)-COUNTIF('QA Team'!A3:A50,\"TOTAL\"),0)"),
        (24, "QA Tasks Open", "=IFERROR(COUNTIFS('QA Tasks'!H3:H5000,\"<>Done\",'QA Tasks'!H3:H5000,\"<>\"),0)"),
        (25, "Dev Action Queue", "=IFERROR(COUNTIF('Dev Tasks'!I3:I152,\"Yes\"),0)"),
        (26, "Shreya Open Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!N3:N5000,\"Shreya Patil\",'06 Bug Tracker'!L3:L5000,\"<>Closed\",'06 Bug Tracker'!L3:L5000,\"<>Won't Fix\"),0)"),
        (27, "Raj Open Bugs", "=IFERROR(COUNTIFS('06 Bug Tracker'!N3:N5000,\"Raj Balani\",'06 Bug Tracker'!L3:L5000,\"<>Closed\",'06 Bug Tracker'!L3:L5000,\"<>Won't Fix\"),0)"),
    ]

    ws["A3"] = "KEY PERFORMANCE INDICATORS"
    ws["A3"].font = SECTION_FONT

    for row, label, formula in kpis:
        ws.cell(row, 1, label).font = KPI_LABEL
        ws.cell(row, 1).fill = CARD
        ws.cell(row, 1).border = THIN
        cell = ws.cell(row, 2, formula)
        cell.fill = KPI_FILL
        cell.font = KPI_FONT
        cell.alignment = CENTER
        cell.border = THIN
        cell.protection = LOCKED
        if "%" in label:
            cell.number_format = "0.0%"
        elif "Hours" in label or "Variance" in label:
            cell.number_format = "0.0"

    # Progress bars (text)
    ws["C4"] = "PROGRESS BARS"
    ws["C4"].font = SECTION_FONT
    ws["C5"] = "Pass Rate"
    ws["D5"] = progress_bar_formula("B16")
    ws["C6"] = "QA Progress"
    ws["D6"] = progress_bar_formula("B18")
    ws["C7"] = "Prod Ready"
    ws["D7"] = progress_bar_formula("B19")
    ws["C8"] = "Module Complete"
    ws["D8"] = (
        '=IFERROR(REPT("#",MAX(0,MIN(10,ROUND(IFERROR(COUNTIF(\'02 Module Master\'!H3:H5000,"Completed")/'
        'MAX(1,COUNTA(\'02 Module Master\'!A3:A5000)),0)*10,0))))&REPT("-",10-MAX(0,MIN(10,ROUND('
        'IFERROR(COUNTIF(\'02 Module Master\'!H3:H5000,"Completed")/MAX(1,COUNTA(\'02 Module Master\'!A3:A5000)),0)*10,0)))),REPT("-",10))'
    )
    ws["C9"] = "Hours Burn"
    ws["D9"] = (
        '=IFERROR(REPT("#",MAX(0,MIN(10,ROUND(IF(B20=0,0,MIN(1,B21/B20))*10,0))))'
        '&REPT("-",10-MAX(0,MIN(10,ROUND(IF(B20=0,0,MIN(1,B21/B20))*10,0)))),REPT("-",10))'
    )
    for r in range(5, 10):
        ws.cell(r, 3).fill = CARD
        ws.cell(r, 3).border = THIN
        ws.cell(r, 4).border = THIN
        ws.cell(r, 4).font = Font(name="Consolas", size=12, color="0F766E")
        ws.cell(r, 4).protection = LOCKED

    # Chart data blocks
    ws["F3"] = "Count"
    ws["F3"].font = SECTION_FONT
    ws["F4"] = "Passed"
    ws["G4"] = "=IFERROR(B7,0)"
    ws["F5"] = "Failed"
    ws["G5"] = "=IFERROR(B8,0)"
    ws["F6"] = "Blocked"
    ws["G6"] = "=IFERROR(B9,0)"
    ws["F7"] = "Not Executed"
    ws["G7"] = "=IFERROR(B10,0)"

    ws["F9"] = "Count"
    ws["F9"].font = SECTION_FONT
    ws["F10"] = "Critical"
    ws["G10"] = "=IFERROR(B12,0)"
    ws["F11"] = "High"
    ws["G11"] = "=IFERROR(B13,0)"
    ws["F12"] = "Medium"
    ws["G12"] = "=IFERROR(B14,0)"
    ws["F13"] = "Low"
    ws["G13"] = "=IFERROR(B15,0)"

    ws["F15"] = "Module"
    ws["F15"].font = SECTION_FONT
    ws["G15"] = "Completion %"
    ws["G15"].font = SECTION_FONT
    for i, (mid, name, *_rest) in enumerate(MODULES):
        total, done = FEATURE_COUNTS_BY_ID[mid]
        ws.cell(16 + i, 6, name)
        # Live formula from Module Master H/I columns (Total/Completed Features)
        ws.cell(16 + i, 7, f"=IFERROR(IF('02 Module Master'!I{3+i}=0,0,'02 Module Master'!J{3+i}/'02 Module Master'!I{3+i}),0)")
        ws.cell(16 + i, 7).number_format = "0%"
        ws.cell(16 + i, 7).protection = LOCKED

    # Daily progress mirror — Date / Executed / Passed / Hours (Daily Log cols A,C,D,I)
    ws["I3"] = "Daily Progress"
    ws["I3"].font = SECTION_FONT
    ws["I4"] = "Date"
    ws["J4"] = "Executed"
    ws["K4"] = "Passed"
    ws["L4"] = "Hours"
    for i in range(7):
        ws.cell(5 + i, 9, f"=IFERROR('Daily Log'!A{3+i},\"\")")
        ws.cell(5 + i, 10, f"=IFERROR('Daily Log'!C{3+i},0)")
        ws.cell(5 + i, 11, f"=IFERROR('Daily Log'!D{3+i},0)")
        ws.cell(5 + i, 12, f"=IFERROR('Daily Log'!I{3+i},0)")

    # QA hours by person (chart data)
    ws["N3"] = "QA Person"
    ws["N3"].font = SECTION_FONT
    ws["O3"] = "Actual Hours"
    for i, (name, *_rest) in enumerate(QA_TEAM):
        ws.cell(4 + i, 14, name)
        ws.cell(4 + i, 15, f"=IFERROR(SUMIF('04 Test Case Repository'!S$3:S$5000,N{4+i},'04 Test Case Repository'!U$3:U$5000),0)")
        ws.cell(4 + i, 15).protection = LOCKED

    # Pie — execution (G3 title "Count" for series)
    pie = PieChart()
    pie.title = "Test Execution Mix"
    labels = Reference(ws, min_col=6, min_row=4, max_row=7)
    data = Reference(ws, min_col=7, min_row=3, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.width = 12
    pie.height = 8
    ws.add_chart(pie, "A25")

    # Bar — bugs
    bar = BarChart()
    bar.type = "col"
    bar.title = "Open Bugs by Severity"
    bar.y_axis.title = "Count"
    data = Reference(ws, min_col=7, min_row=9, max_row=13)
    cats = Reference(ws, min_col=6, min_row=10, max_row=13)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 12
    bar.height = 8
    ws.add_chart(bar, "D25")

    # Bar — module completion
    bar2 = BarChart()
    bar2.type = "bar"
    bar2.title = "Module Completion %"
    data = Reference(ws, min_col=7, min_row=15, max_row=15 + len(MODULES))
    cats = Reference(ws, min_col=6, min_row=16, max_row=15 + len(MODULES))
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(cats)
    bar2.width = 15
    bar2.height = 12
    ws.add_chart(bar2, "H25")

    # Line — daily
    line = LineChart()
    line.title = "Daily Progress (Executed vs Passed)"
    line.y_axis.title = "Cases"
    line.x_axis.title = "Date"
    data = Reference(ws, min_col=10, min_row=4, max_row=11)
    cats = Reference(ws, min_col=9, min_row=5, max_row=11)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.width = 15
    line.height = 8
    ws.add_chart(line, "A42")

    # Bar — hours by QA
    bar3 = BarChart()
    bar3.type = "col"
    bar3.title = "Actual Hours by QA Person"
    bar3.y_axis.title = "Hours"
    data = Reference(ws, min_col=15, min_row=3, max_row=3 + len(QA_TEAM))
    cats = Reference(ws, min_col=14, min_row=4, max_row=3 + len(QA_TEAM))
    bar3.add_data(data, titles_from_data=True)
    bar3.set_categories(cats)
    bar3.width = 14
    bar3.height = 8
    ws.add_chart(bar3, "H42")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 12
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def build_modules(wb: Workbook):
    ws = wb.create_sheet("02 Module Master")
    headers = [
        "Module ID", "Module Name", "Portal", "Description", "Owner", "QA Owner", "Priority", "Status",
        "Total Features", "Completed Features", "Pending Features", "Est Hours", "Actual Hours",
        "Criticality", "Dependencies", "Remarks",
    ]
    deps = {
        "MOD-03": "Auth, Billing credits, Deepgram",
        "MOD-04": "Sessions, Credits",
        "MOD-08": "Credits (parse-resume)",
        "MOD-14": "Admin question bank",
        "MOD-18": "Stripe/Razorpay secrets",
        "MOD-19": "Live Practice Coach",
        "MOD-16": "RLS admin role",
    }
    qa_owners = {
        "MOD-01": "Raj Balani", "MOD-02": "Raj Balani", "MOD-03": "Shreya Patil", "MOD-04": "Shreya Patil",
        "MOD-05": "Shreya Patil", "MOD-06": "Raj Balani", "MOD-07": "Raj Balani", "MOD-08": "Shreya Patil",
        "MOD-09": "Raj Balani", "MOD-10": "Raj Balani", "MOD-11": "Raj Balani", "MOD-12": "Raj Balani",
        "MOD-13": "Raj Balani", "MOD-14": "Raj Balani", "MOD-15": "Raj Balani", "MOD-16": "Raj Balani",
        "MOD-17": "Shreya Patil", "MOD-18": "Raj Balani", "MOD-19": "Shreya Patil", "MOD-20": "Shreya Patil",
    }
    rows = []
    for mid, name, portal, desc, owner, pri, status in MODULES:
        total, done = FEATURE_COUNTS_BY_ID[mid]
        est = total * 2.5
        actual = round(done * 2.2, 1)
        rows.append([
            mid, name, portal, desc, owner, qa_owners[mid], pri, status,
            total, done, total - done, est, actual, pri, deps.get(mid, "—"),
            "Seeded from audit 28 Jul 2026",
        ])
    write_table(ws, "Module Master — Clarify AI (20 core modules)", "", headers, rows)
    add_dv(ws, lr(wb, "Portals"), "C3:C5000")
    add_dv(ws, lr(wb, "Owners"), "E3:E5000")
    add_dv(ws, lr(wb, "Testers"), "F3:F5000")
    add_dv(ws, lr(wb, "Priority"), "G3:G5000")
    add_dv(ws, lr(wb, "ModuleStatus"), "H3:H5000")
    add_dv(ws, lr(wb, "Criticality"), "N3:N5000")
    cf_severity(ws, "G")
    cf_severity(ws, "N")
    cf_verdict(ws, "H")
    # Totals
    t = 3 + len(rows)
    ws.cell(t, 1, "TOTAL")
    ws.cell(t, 1).font = Font(bold=True)
    for col, letter in [(9, "I"), (10, "J"), (11, "K"), (12, "L"), (13, "M")]:
        ws.cell(t, col, f"=SUM({letter}3:{letter}{t-1})")
        ws.cell(t, col).font = Font(bold=True)
    finish_sheet(ws, 2, len(rows) + 1, len(headers))


def build_features(wb: Workbook):
    ws = wb.create_sheet("03 Feature Inventory")
    headers = [
        "Feature ID", "Module", "Feature Name", "Description",
        "Frontend", "Backend", "Database", "API", "Edge Function",
        "Status", "Implementation Status", "QA Status", "Priority", "Risk",
        "Owner", "QA Owner", "Est Hours", "Actual Hours", "Comments",
        "Route", "Deep Link", "How it works", "Role required", "Credit cost", "Blocking bugs",
    ]
    samples = []
    for i, f in enumerate(FEATURES):
        route, how, role, credit, blocking = feature_qa_columns(f)
        excel_row = 3 + i
        deep = f"=IFERROR(Environments!$B$3&T{excel_row},T{excel_row})"
        samples.append(list(f) + [route, deep, how, role, credit, blocking])
    write_table(
        ws,
        "Feature Inventory — full Clarify AI catalogue (" + str(len(samples)) + " features)",
        "Deep Link = Environments!BaseURL & Route. Replace BaseURL after seeding staging.",
        headers,
        samples,
    )
    add_dv(ws, lr(wb, "Modules"), "B3:B5000")
    for col in "EFGHI":
        add_dv(ws, lr(wb, "Coverage"), f"{col}3:{col}5000")
    add_dv(ws, lr(wb, "Status"), "J3:J5000")
    add_dv(ws, lr(wb, "ImplStatus"), "K3:K5000")
    add_dv(ws, lr(wb, "QAStatus"), "L3:L5000")
    add_dv(ws, lr(wb, "Priority"), "M3:M5000")
    add_dv(ws, lr(wb, "Risk"), "N3:N5000")
    add_dv(ws, lr(wb, "Owners"), "O3:O5000")
    add_dv(ws, lr(wb, "Testers"), "P3:P5000")
    cf_verdict(ws, "J")
    cf_verdict(ws, "L")
    cf_severity(ws, "M")
    cf_severity(ws, "N")
    tot = 3 + len(samples)
    ws.cell(tot, 1, "TOTAL")
    ws.cell(tot, 17, f"=SUM(Q3:Q{tot-1})")
    ws.cell(tot, 18, f"=SUM(R3:R{tot-1})")
    for c in (1, 17, 18):
        ws.cell(tot, c).font = Font(bold=True)
    finish_sheet(ws, 2, len(samples) + 1, len(headers))


def build_credentials(wb: Workbook):
    ws = wb.create_sheet("21 Test Credentials")
    headers = [
        "Role", "Email", "Password location", "Plan", "Credits seed",
        "Notes", "Owner", "Last rotated",
    ]
    samples = [
        ("Free", "qa.free@clarify.ai.test", ".env.qa.local → QA_FREE_PASSWORD", "free", "50",
         "Seeded by npm run qa:seed-accounts — never commit passwords", "Raj Balani", ""),
        ("Pro", "qa.pro@clarify.ai.test", ".env.qa.local → QA_PRO_PASSWORD", "pro", "1400",
         "Stripe test checkout path", "Raj Balani", ""),
        ("Max", "qa.max@clarify.ai.test", ".env.qa.local → QA_MAX_PASSWORD", "enterprise", "4000",
         "Workbook Max/Elite maps to plan_id=enterprise", "Raj Balani", ""),
        ("Admin", "qa.admin@clarify.ai.test", ".env.qa.local → QA_ADMIN_PASSWORD", "enterprise", "4000",
         "user_roles.admin via seed script", "Shreya Patil", ""),
        ("Stripe test card", "N/A", "4242 4242 4242 4242", "N/A", "N/A",
         "Any future expiry + any CVC; Stripe test mode only", "Raj Balani", ""),
        ("Razorpay test", "N/A", "Dashboard test keys", "N/A", "N/A",
         "INR one-time Order — no auto-renew (BUG-OPEN-01); keys not in .env.local yet", "Raj Balani", ""),
    ]
    write_table(
        ws,
        "Test Credentials — seeded accounts",
        "Passwords live only in gitignored .env.qa.local (npm run qa:seed-accounts). Do not paste secrets into this workbook.",
        headers,
        samples,
    )
    finish_sheet(ws, 2, len(samples), len(headers))


def build_environments(wb: Workbook):
    ws = wb.create_sheet("22 Environments")
    headers = [
        "Environment", "Base URL", "Supabase project", "Owner", "Read/Write", "Notes",
    ]
    samples = [
        ("Local", "http://localhost:5173", "qzgvjrvtkwlzxpmlddkx (linked)", "Dev Lead", "Read-Write", "npm run dev"),
        ("Staging / closed beta", "https://clarify.ai.sltfinanceindia.com", "qzgvjrvtkwlzxpmlddkx", "QA Lead", "Read-Write",
         "Primary QA target — same project as prod until separate staging exists"),
        ("Prod (read-only QA)", "https://clarify.ai.sltfinanceindia.com", "qzgvjrvtkwlzxpmlddkx", "QA Lead", "Read-only",
         "Prefer seeded QA_* users; avoid destructive admin tests on shared data"),
    ]
    write_table(ws, "Environments", "B3 (Staging Base URL) powers Feature Inventory Deep Link formulas", headers, samples)

    # Secrets checklist section
    ws.cell(8, 1, "Environments & Secrets Checklist")
    ws.cell(8, 1).font = SECTION_FONT
    sec_headers = ["Secret / Config", "Where", "Launch impact", "Notes"]
    for i, h in enumerate(sec_headers, 1):
        cell = ws.cell(9, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(OPS_SECRETS_CHECKLIST):
        for j, val in enumerate(row, 1):
            ws.cell(10 + i, j, val)
            ws.cell(10 + i, j).alignment = WRAP
            ws.cell(10 + i, j).border = THIN
    finish_sheet(ws, 2, len(samples), len(headers), filter=True)


def build_playbooks(wb: Workbook):
    ws = wb.create_sheet("23 Module Playbooks")
    headers = [
        "Module ID", "Module", "Purpose", "Entry URL", "Happy path (steps)",
        "Plan gate", "Dependencies", "Done when",
    ]
    samples = []
    play = {
        "MOD-01": ("Sign up/login and finish onboarding", "/login", "1. Sign up 2. Verify email 3. Complete 5 steps", "Free", "Supabase Auth", "Land on dashboard with profile"),
        "MOD-02": ("See KPIs and navigate app", "/app/dashboard", "1. Open home 2. Check credits/sessions 3. Mobile More nav", "Free", "Auth session", "KPIs load; nav works"),
        "MOD-03": ("Run live coach with STT + hint", "/app/live", "1. Start session 2. Capture audio 3. Request hint", "Free/Pro overlay", "Deepgram + credits", "Transcript + hint without crash"),
        "MOD-04": ("Timed mock interview", "/app/mock", "1. Configure 2. Answer 3. Finish scoring", "Free", "AI edge", "Session saved; note no TTS"),
        "MOD-05": ("Review history + debrief", "/app/sessions", "1. Open session 2. Debrief 3. Share if enabled", "Free", "Credits for debrief", "Scores/debrief render"),
        "MOD-06": ("Analytics + usage ledger", "/app/analytics", "1. Open analytics 2. Open /app/usage", "Free", "Sessions data", "Charts/ledger load"),
        "MOD-07": ("Prep Lab tools", "/app/prep", "1. STAR 2. Rephrase 3. Confirm credit toast", "Free", "prep-tool edge", "Output appears; credits deduct"),
        "MOD-08": ("Documents + gap + versions", "/app/documents", "1. Upload resume 2. JD gap analysis 3. Version history", "Free", "parse/gap edges", "Parse + gap + versions visible"),
        "MOD-09": ("Answer bank CRUD", "/app/answers", "1. Create 2. Edit 3. Favourite", "Free", "DB", "Persists after refresh"),
        "MOD-10": ("Schedule + calendar", "/app/interviews", "1. Create interview 2. Interview Day 3. Calendar sync if secrets", "Pro for sync", "Resend/Google secrets", "CRUD works; sync Blocked without secrets"),
        "MOD-11": ("Company brief", "/app/companies", "1. Search company 2. Generate brief", "Pro+", "company-research", "Brief cached"),
        "MOD-12": ("Notifications + prefs", "/app/notifications", "1. See inbox 2. Settings notifications prefs", "Free", "send-email prefs", "Prefs save; email gated"),
        "MOD-13": ("Referrals + guide", "/app/referrals", "1. Copy code 2. Open guide", "Free", "DB", "Code visible"),
        "MOD-14": ("Gov exam mock", "/app/mock-test", "1. Pick exam 2. Create test 3. Submit", "Free (India)", "Expanded MCQ seeds", "Full mock completes"),
        "MOD-15": ("Settings honesty", "/app/settings", "1. Billing 2. Privacy analytics 3. Integrations", "Free", "PostHog optional", "Honesty banners accurate"),
        "MOD-16": ("Admin ops", "/app/admin", "1. Users 2. Support→Live Chat 3. Model costs read-only", "Admin", "admin role", "No stub Save; chat reachable"),
        "MOD-17": ("Marketing pages", "/", "1. Landing 2. Pricing 3. Legal", "Guest", "Static", "Pages render"),
        "MOD-18": ("Stripe/Razorpay + credits", "/app/settings/billing", "1. Checkout test 2. Webhook credit 3. Ledger", "Free→Pro", "Stripe/Razorpay secrets", "Credits grant; Razorpay one-time honesty"),
        "MOD-19": ("Electron overlay", "(desktop)", "1. Boot 2. Hotkeys 3. Overlay", "Pro overlay", "Unsigned build", "Works; SmartScreen expected"),
        "MOD-20": ("Security platform", "N/A", "1. RLS spot-check 2. Rate limit 3. Webhook signature", "Admin/ops", "Service role", "No open RLS holes in smoke"),
    }
    for mid, name, portal, desc, owner, pri, status in MODULES:
        purpose, entry, steps, gate, deps, done = play.get(
            mid, (desc, "/app", "Exercise module happy path", "Free", "Auth", "No crash")
        )
        samples.append([mid, name, purpose, entry, steps, gate, deps, done])
    write_table(ws, "Module Playbooks — 20 modules", "", headers, samples)
    finish_sheet(ws, 2, len(samples), len(headers))


def build_launch_status(wb: Workbook):
    ws = wb.create_sheet("24 Launch Status")
    closed = sum(1 for b in BUGS if b[11] == "Closed")
    opened = sum(1 for b in BUGS if b[11] == "Open")
    blocked = sum(1 for b in BUGS if b[11] == "Blocked")
    wont = sum(1 for b in BUGS if b[11] == "Won't Fix")
    headers = ["Metric", "Count", "Notes"]
    samples = [
        ("Bugs Fixed/Closed", closed, "Includes P0–P2 + launch-pass OPEN fixes"),
        ("Bugs Open", opened, "Product/roadmap still open"),
        ("Bugs Blocked (ops/secrets)", blocked, "Calendar, Resend, signing, key rotation, Gemini collect"),
        ("Bugs Won't Fix / Descope", wont, "Outlook etc."),
        ("Features total", len(FEATURES), "See Feature Inventory"),
        ("Features Completed status", sum(1 for f in FEATURES if f[9] == "Completed"), ""),
    ]
    write_table(ws, "Launch Status Dashboard", "Hire-ready summary — do not mark Blocked items Fixed", headers, samples)

    ws.cell(12, 1, "Blocking for launch (honesty)")
    ws.cell(12, 1).font = SECTION_FONT
    blockers = [
        "Razorpay is one-time Order only — copy must not promise renew (BUG-OPEN-01)",
        "Unsigned Windows installer → SmartScreen (BUG-OPEN-18/19) unless EV/OV cert",
        "Google Calendar sync 501 until OAuth secrets (BUG-OPEN-03) if calendar is promised",
        "Interview email reminders need RESEND_API_KEY (BUG-OPEN-26)",
        "Anon key rotation is Dashboard ops only (BUG-OPEN-30)",
        "Gov bank expanded in migration — apply 20260729120000 before claiming full mocks",
    ]
    for i, line in enumerate(blockers):
        ws.cell(13 + i, 1, line)
        ws.cell(13 + i, 1).alignment = WRAP
    finish_sheet(ws, 2, len(samples), len(headers))


def build_smoke_pack(wb: Workbook):
    ws = wb.create_sheet("25 Smoke Pack")
    headers = [
        "Smoke ID", "Title", "Credential role", "Route", "Deep Link", "Steps", "Expected", "Pass/Fail",
    ]
    cases = [
        ("SMK-01", "Login email", "Free", "/login", "Sign in with seeded free user", "Dashboard loads"),
        ("SMK-02", "Dashboard KPIs", "Free", "/app/dashboard", "Open home", "Credits/sessions visible"),
        ("SMK-03", "Mobile More nav", "Free", "/app/dashboard", "Open More; tap Notifications, Guide, Profile, Usage", "All routes open"),
        ("SMK-04", "Live hint", "Pro", "/app/live", "Start + request hint", "Hint returns; credits drop"),
        ("SMK-05", "Mock create", "Free", "/app/mock", "Start mock", "Session creates"),
        ("SMK-06", "Resume parse", "Free", "/app/documents", "Upload resume", "Parsed fields show"),
        ("SMK-07", "Gap analysis", "Pro", "/app/documents", "JD → Run gap analysis", "Match score UI; 402 if broke"),
        ("SMK-08", "Resume versions", "Free", "/app/documents", "Open resume detail", "Version history card"),
        ("SMK-09", "Gov mock", "Free", "/app/mock-test", "Create SSC/JEE-style test", "Enough questions load"),
        ("SMK-10", "Billing Stripe test", "Free", "/app/settings/billing", "Checkout with 4242…", "Webhook grants plan/credits"),
        ("SMK-11", "Razorpay honesty", "Free", "/app/settings/billing", "Read INR copy + button labels", "One-time / no auto-renew disclosed"),
        ("SMK-12", "Notification prefs", "Free", "/app/settings/notifications", "Read honesty + save prefs", "Save succeeds"),
        ("SMK-13", "Privacy analytics", "Free", "/app/settings/privacy", "Turn analytics off + save", "Opt-out applied"),
        ("SMK-14", "Admin model costs", "Admin", "/app/admin/model-costs", "Open page on narrow width", "Read-only; horizontal scroll works"),
        ("SMK-15", "Admin support→chat", "Admin", "/app/admin/support", "Click Live Chat", "Live Chat opens"),
        ("SMK-16", "Admin QA local", "Admin", "/app/admin/qa-checklist", "Read banner", "Says local browser only"),
        ("SMK-17", "Calendar blocked honesty", "Pro", "/app/settings/integrations", "Try sync without secrets", "501/blocked — not silently ok"),
        ("SMK-18", "Company research gate", "Free", "/app/companies", "Try generate", "Upgrade / plan gate"),
        ("SMK-19", "MFA challenge", "Free+MFA", "/login", "Sign in with TOTP-enrolled user", "Code screen before app"),
        ("SMK-20", "Electron boot", "Pro", "(desktop)", "Launch unsigned build", "App boots; SmartScreen may warn"),
    ]
    samples = []
    for i, (sid, title, role, route, steps, expected) in enumerate(cases):
        r = 3 + i
        deep = f"=IFERROR(Environments!$B$3&D{r},D{r})"
        samples.append([sid, title, role, route, deep, steps, expected, "Not Run"])
    write_table(ws, "Smoke Pack — ~20 must-pass", "Use Test Credentials + Staging BaseURL", headers, samples)
    add_dv(ws, lr(wb, "PassFail"), "H3:H5000")
    cf_verdict(ws, "H")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_test_cases(wb: Workbook):
    ws = wb.create_sheet("04 Test Case Repository")
    headers = [
        "Test Case ID", "Requirement ID", "Module", "Sub Module", "Feature", "Screen",
        "User Role", "Priority", "Severity", "Test Scenario", "Test Case", "Preconditions",
        "Test Steps", "Expected Result", "Actual Result", "Pass/Fail", "Status", "Bug ID",
        "Tester", "Est Hours", "Actual Hours", "Execution Date", "Build Version", "Environment",
        "Browser", "Device", "Platform", "Automation", "Regression", "Smoke", "Sanity",
        "Sprint", "Release", "Remarks",
    ]
    samples = [
        ("TC-001", "REQ-AUTH-01", "Auth & Onboarding", "Login", "Email login", "/login", "Free User", "Critical", "Critical",
         "Valid credentials", "User signs in with email/password", "Verified account",
         "1. Open /login\n2. Enter credentials\n3. Submit", "Redirect to /app or onboarding", "Landed on dashboard", "Pass", "Completed", "",
         "Raj Balani", 1.0, 0.75, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "Yes", "Yes", "Sprint 26", "1.0.0-beta", ""),
        ("TC-002", "REQ-AUTH-02", "Auth & Onboarding", "Login", "Lockout", "/login", "Free User", "High", "High",
         "Repeated failures", "Lockout after threshold", "Test account",
         "1. Fail login repeatedly\n2. Observe message", "Temporary lockout shown", "", "Not Run", "Ready", "",
         "Raj Balani", 1.5, 0, "", "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "No", "Yes", "Sprint 26", "1.0.0-beta", ""),
        ("TC-003", "REQ-LIVE-01", "Live Practice Coach", "STT", "Deepgram stream", "/app/live", "Pro User", "Critical", "Critical",
         "Mic open", "Near-real-time transcript", "Pro + mic permission",
         "1. Start live\n2. Speak\n3. Observe transcript", "Segments appear under 2s", "OK", "Pass", "Completed", "",
         "Shreya Patil", 2.0, 1.5, TODAY, "1.0.0", "Staging", "Electron", "Laptop", "Windows", "Manual", "Yes", "Yes", "No", "Sprint 26", "1.0.0-beta", ""),
        ("TC-004", "REQ-LIVE-02", "Live Practice Coach", "Overlay", "Pro gate", "/app/live/overlay", "Free User", "Critical", "High",
         "Free overlay blocked", "Server rejects overlay capture", "Free account",
         "1. Attempt overlay capture\n2. Check response", "403 / upgrade required", "403 returned", "Pass", "Completed", "BUG-001",
         "Shreya Patil", 1.5, 1.25, TODAY, "1.0.0", "Staging", "Electron", "Desktop", "Windows", "Candidate", "Yes", "Yes", "No", "Sprint 26", "1.0.0-beta", "Retest after P0-3"),
        ("TC-005", "REQ-BILL-01", "Billing & Credits", "Checkout", "Max Stripe price", "/pricing", "Free User", "Critical", "Critical",
         "Upgrade Max", "Enterprise price id used", "Stripe test keys",
         "1. Select Max\n2. Checkout\n3. Inspect price", "Non-empty enterprise price", "OK", "Pass", "Completed", "",
         "Raj Balani", 2.0, 1.5, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "Yes", "Yes", "Sprint 26", "1.0.0-beta", "P0-1"),
        ("TC-006", "REQ-BILL-02", "Billing & Credits", "Webhook", "payment_failed", "stripe-webhook", "Pro User", "Critical", "Critical",
         "Failed invoice", "Wallet not wiped", "credits=500",
         "1. Simulate invoice.payment_failed\n2. Read credits", "Credits remain 500", "Unchanged", "Pass", "Completed", "",
         "Raj Balani", 1.5, 1.0, TODAY, "1.0.0", "CI", "N/A", "Desktop", "Web", "Automated", "Yes", "Yes", "No", "Sprint 26", "1.0.0-beta", "P0-2"),
        ("TC-007", "REQ-DOC-01", "Documents & Resumes", "Parse", "Free onboarding parse", "/onboarding", "Free User", "High", "Medium",
         "First parse free", "No deduct on first onboarding parse", "onboarding_completed=false",
         "1. Upload resume\n2. Check ledger", "Free path / no deduct", "", "Not Run", "Ready", "",
         "Shreya Patil", 1.5, 0, "", "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "No", "Yes", "Sprint 26", "1.0.0-beta", ""),
        ("TC-008", "REQ-GOV-01", "Gov Exam Mock Tests", "Submit", "Atomic submit", "/app/mock-test/session/:id", "Free User", "Critical", "High",
         "Submit answers", "Score persisted once", "Active test",
         "1. Answer\n2. Submit\n3. Open results", "Score + breakdown", "OK", "Pass", "Completed", "",
         "Raj Balani", 2.0, 1.75, TODAY, "1.0.0", "Staging", "Chrome", "Android Phone", "Android", "Manual", "Yes", "Yes", "Yes", "Sprint 26", "1.0.0-beta", ""),
        ("TC-009", "REQ-ADM-01", "Admin Portal", "Revenue", "USD vs INR", "/app/admin/revenue", "Admin", "Critical", "High",
         "Admin revenue view", "Currencies separate", "Admin + sample data",
         "1. Open Revenue\n2. Compare hand calc", "USD MRR + INR separate", "OK", "Pass", "Completed", "",
         "Raj Balani", 2.5, 2.0, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "No", "No", "Sprint 26", "1.0.0-beta", "P0-6"),
        ("TC-010", "REQ-APP-01", "Dashboard", "Routing", "In-app 404", "/app/does-not-exist", "Pro User", "Medium", "Low",
         "Unknown route", "NotFound in AppShell", "Authenticated",
         "1. Navigate unknown /app path", "In-app NotFound", "OK", "Pass", "Completed", "",
         "Shreya Patil", 0.5, 0.25, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "Yes", "Yes", "Sprint 26", "1.0.0-beta", "P1-4"),
        ("TC-011", "REQ-LIVE-03", "Live Practice Coach", "Credits", "Hint deduct", "/app/live", "Pro User", "High", "High",
         "Insufficient credits", "402 when wallet low", "credits below cost",
         "1. Drain credits\n2. Request hint", "402 + message", "In progress", "In Progress", "In Progress", "",
         "Shreya Patil", 1.5, 0.5, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Candidate", "Yes", "No", "No", "Sprint 26", "1.0.0-beta", ""),
        ("TC-012", "REQ-SEC-01", "Security & Platform", "Authz", "Non-admin admin route", "/app/admin", "Free User", "Critical", "Critical",
         "Unauthorized admin", "Redirect / deny", "Non-admin user",
         "1. Open /app/admin", "Denied", "Denied", "Pass", "Completed", "",
         "Raj Balani", 1.0, 0.75, TODAY, "1.0.0", "Staging", "Chrome", "Desktop", "Web", "Manual", "Yes", "Yes", "Yes", "Sprint 26", "1.0.0-beta", ""),
    ]
    write_table(ws, "Test Case Repository — Clarify AI", "", headers, samples)
    add_dv(ws, lr(wb, "Modules"), "C3:C5000")
    add_dv(ws, lr(wb, "Roles"), "G3:G5000")
    add_dv(ws, lr(wb, "Priority"), "H3:H5000")
    add_dv(ws, lr(wb, "Severity"), "I3:I5000")
    add_dv(ws, lr(wb, "PassFail"), "P3:P5000")
    add_dv(ws, lr(wb, "Status"), "Q3:Q5000")
    add_dv(ws, lr(wb, "Testers"), "S3:S5000")
    add_dv(ws, lr(wb, "Environment"), "X3:X5000")
    add_dv(ws, lr(wb, "Browser"), "Y3:Y5000")
    add_dv(ws, lr(wb, "Device"), "Z3:Z5000")
    add_dv(ws, lr(wb, "Platform"), "AA3:AA5000")
    add_dv(ws, lr(wb, "Automation"), "AB3:AB5000")
    add_dv(ws, lr(wb, "YesNo"), "AC3:AE5000")
    add_dv(ws, lr(wb, "Sprint"), "AF3:AF5000")
    add_dv(ws, lr(wb, "Release"), "AG3:AG5000")
    cf_verdict(ws, "P")
    cf_verdict(ws, "Q")
    cf_severity(ws, "H")
    cf_severity(ws, "I")
    t = 3 + len(samples)
    ws.cell(t, 1, "TOTAL")
    ws.cell(t, 20, f"=SUM(T3:T{t-1})")
    ws.cell(t, 21, f"=SUM(U3:U{t-1})")
    for c in (1, 20, 21):
        ws.cell(t, c).fill = KPI_FILL
        ws.cell(t, c).font = Font(bold=True, color="FFFFFF")
    finish_sheet(ws, 2, len(samples) + 1, len(headers))


def build_e2e(wb: Workbook):
    ws = wb.create_sheet("05 E2E User Flows")
    headers = [
        "Flow ID", "Flow Name", "Portal", "Start Page", "End Page", "Steps",
        "Expected Outcome", "Actual Outcome", "Pass", "Risk", "Dependencies",
        "QA Owner", "Est Hours", "Actual Hours", "Comments",
    ]
    samples = [
        ("E2E-01", "Signup to Onboarding to Dashboard", "Auth Portal", "/signup", "/app/dashboard",
         "1. Sign up 2. Verify email 3. Complete 5 steps 4. Land dashboard",
         "Profile + onboarding_completed + starter credits", "OK on Staging", "Pass", "High", "Auth, Email",
         "Raj Balani", 3.0, 2.5, ""),
        ("E2E-02", "Free to Pro Stripe checkout", "Candidate App", "/pricing", "/app/settings/billing",
         "1. Choose Pro 2. Stripe Checkout 3. Webhook 4. Open Billing",
         "plan=pro, credits granted, ledger entry", "", "Not Run", "Critical", "Stripe test keys",
         "Raj Balani", 4.0, 0, "Money path"),
        ("E2E-03", "Live coach rehearsal", "Candidate App", "/app/live", "/app/sessions/:id",
         "1. Start 2. Audio 3. Generate answer 4. End 5. Debrief",
         "Session + transcript + debrief", "OK", "Pass", "Critical", "Deepgram, Gemini, Credits",
         "Shreya Patil", 5.0, 4.0, ""),
        ("E2E-04", "Gov exam create to results", "Candidate App", "/app/mock-test", "/app/mock-test/results/:id",
         "1. Select exam 2. create-test 3. Answer 4. submit 5. Results",
         "Atomic score, no double submit", "OK", "Pass", "High", "Question bank",
         "Raj Balani", 3.5, 3.0, "Thin content"),
        ("E2E-05", "Admin seed to candidate test", "Admin Portal", "/app/admin/seed-questions", "/app/mock-test",
         "1. Import questions 2. Candidate creates test",
         "Imported questions available", "Blocked", "Blocked", "High", "Admin, Gemini PDF",
         "Raj Balani", 4.0, 1.0, ""),
        ("E2E-06", "Electron overlay boot", "Electron Overlay", "Clarify AI.exe", "Overlay window",
         "1. Launch 2. Auth if needed 3. Overlay topmost 4. Hotkey",
         "Overlay-only shell", "", "Not Run", "Critical", "Code signing deferred",
         "Shreya Patil", 3.0, 0, "Unsigned beta OK"),
    ]
    write_table(ws, "End-to-End User Flows — golden paths", "", headers, samples)
    add_dv(ws, lr(wb, "Portals"), "C3:C5000")
    add_dv(ws, lr(wb, "PassFail"), "I3:I5000")
    add_dv(ws, lr(wb, "Risk"), "J3:J5000")
    add_dv(ws, lr(wb, "Testers"), "L3:L5000")
    cf_verdict(ws, "I")
    cf_severity(ws, "J")
    t = 3 + len(samples)
    ws.cell(t, 1, "TOTAL")
    ws.cell(t, 13, f"=SUM(M3:M{t-1})")
    ws.cell(t, 14, f"=SUM(N3:N{t-1})")
    for c in (1, 13, 14):
        ws.cell(t, c).font = Font(bold=True)
    finish_sheet(ws, 2, len(samples) + 1, len(headers))


def build_bugs(wb: Workbook):
    ws = wb.create_sheet("06 Bug Tracker")
    headers = [
        "Bug ID", "Title", "Description", "Module", "Feature", "Environment", "Browser",
        "Device", "OS", "Severity", "Priority", "Status", "Assigned To", "Reported By",
        "Found Date", "Fixed Date", "Retest Date", "Build", "Sprint", "Root Cause",
        "Resolution", "Screenshot Link", "Video Link", "Comments",
    ]
    samples = [list(b) for b in BUGS]
    write_table(
        ws,
        "Bug Tracker — audit defect register + open issues (" + str(len(samples)) + " bugs)",
        "",
        headers,
        samples,
    )
    add_dv(ws, lr(wb, "Modules"), "D3:D5000")
    add_dv(ws, lr(wb, "Environment"), "F3:F5000")
    add_dv(ws, lr(wb, "Browser"), "G3:G5000")
    add_dv(ws, lr(wb, "Device"), "H3:H5000")
    add_dv(ws, lr(wb, "Severity"), "J3:J5000")
    add_dv(ws, lr(wb, "Priority"), "K3:K5000")
    add_dv(ws, lr(wb, "BugStatus"), "L3:L5000")
    # Assigned To = developer or module owner role; Reported By = QA (Shreya/Raj)
    add_dv(ws, lr(wb, "Assignees"), "M3:M5000")
    add_dv(ws, lr(wb, "Testers"), "N3:N5000")
    add_dv(ws, lr(wb, "Sprint"), "S3:S5000")
    add_dv(ws, lr(wb, "RootCause"), "T3:T5000")
    cf_severity(ws, "J")
    cf_severity(ws, "K")
    # Workflow hint row under table
    tip = 3 + len(samples) + 2
    ws.cell(tip, 1, "WORKFLOW").font = Font(bold=True)
    ws.cell(tip, 2, "QA (Shreya/Raj) files bug → set Reported By + Assigned To (dev) → Dev Tasks sheet mirrors this row → Dev fixes → Status=Fixed/Retest → QA retests → Closed. Filter Status / Reported By like slicers.")
    ws.merge_cells(start_row=tip, start_column=2, end_row=tip, end_column=10)
    finish_sheet(ws, 2, len(samples), len(headers))


def build_uiux(wb: Workbook):
    ws = wb.create_sheet("07 UI-UX Testing")
    headers = [
        "Screen", "Component", "Alignment", "Spacing", "Responsive", "Typography",
        "Icons", "Accessibility", "Dark Mode", "Light Mode", "Animation", "Pass", "Remarks",
    ]
    samples = [
        ("/pricing", "Plan cards", "Pass", "Pass", "Pass", "Pass", "Pass", "Partial", "Pass", "Pass", "Pass", "Pass", "Max $79"),
        ("/app/live", "Transcript panel", "Pass", "Pass", "Partial", "Pass", "Pass", "Partial", "Pass", "Pass", "Pass", "Pass", "Electron primary"),
        ("/app/dashboard", "KPI cards", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", ""),
        ("/app/admin/revenue", "USD/INR charts", "Pass", "Pass", "Pass", "Pass", "Pass", "Partial", "Pass", "Pass", "N/A", "Pass", ""),
        ("/login", "Auth form", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", ""),
    ]
    write_table(ws, "UI/UX Testing", "", headers, samples)
    for col in list("CDEFGHIJKL"):
        add_dv(ws, lr(wb, "PassFail"), f"{col}3:{col}5000")
        cf_verdict(ws, col)
    finish_sheet(ws, 2, len(samples), len(headers))


def build_mobile(wb: Workbook):
    ws = wb.create_sheet("08 Mobile Responsiveness")
    headers = ["Screen", "Android", "iPhone", "Tablet", "Landscape", "Portrait", "Small Screen", "Medium Screen", "Large Screen", "Responsive", "Pass", "Comments"]
    samples = [
        ("/app/dashboard", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Yes", "Pass", "MobileNav partial"),
        ("/app/mock-test/session", "Pass", "Pass", "Pass", "Partial", "Pass", "Pass", "Pass", "Pass", "Yes", "Pass", ""),
        ("/pricing", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Yes", "Pass", ""),
        ("/app/live", "Fail", "Fail", "Partial", "N/A", "N/A", "Fail", "Partial", "Pass", "Partial", "Fail", "Desktop-first"),
        ("/app/admin/users", "Partial", "Partial", "Pass", "Pass", "Pass", "Partial", "Pass", "Pass", "Partial", "Partial", "Wide tables"),
    ]
    write_table(ws, "Mobile Responsiveness", "", headers, samples)
    for col in list("BCDEFGHI"):
        add_dv(ws, lr(wb, "PassFail"), f"{col}3:{col}5000")
        cf_verdict(ws, col)
    add_dv(ws, lr(wb, "YesNo"), "J3:J5000")
    add_dv(ws, lr(wb, "PassFail"), "K3:K5000")
    cf_verdict(ws, "K")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_api(wb: Workbook):
    ws = wb.create_sheet("09 API Testing")
    headers = [
        "API Name", "Method", "Endpoint", "Authentication", "Headers", "Request",
        "Expected Response", "Actual Response", "Status Code", "Performance", "Security", "Pass", "Comments",
    ]
    samples = [
        ("Create Checkout", "POST", "/functions/v1/create-checkout", "Bearer JWT", "Authorization, apikey",
         '{"plan":"enterprise"}', "Checkout URL", "URL ok", "200", "Good", "Pass", "Pass", "Enterprise price"),
        ("Parse Resume", "POST", "/functions/v1/parse-resume", "Bearer JWT", "x-clarify-onboarding-parse",
         "multipart", "Parsed JSON", "OK", "200", "Acceptable", "Pass", "Pass", "Free once"),
        ("Generate Answer", "POST", "/functions/v1/generate-answer", "Bearer JWT", "Authorization",
         '{"mode":"overlay"}', "403 free overlay", "403", "403", "Excellent", "Pass", "Pass", "Pro gate"),
        ("Razorpay Order", "POST", "/functions/v1/razorpay-create-order", "Bearer JWT", "Authorization",
         '{"pack":"credits_50"}', "order_id", "OK", "200", "Good", "Pass", "Pass", "Rate limited"),
        ("Stripe Webhook", "POST", "/functions/v1/stripe-webhook", "Webhook HMAC", "Stripe-Signature",
         "invoice.payment_failed", "200 credits intact", "200", "200", "Excellent", "Pass", "Pass", "P0-2"),
        ("Sync Calendar", "POST", "/functions/v1/sync-calendar", "Bearer JWT", "Authorization",
         "{}", "501 or events", "501", "501", "N/A", "Pass", "Blocked", "Secrets"),
    ]
    write_table(ws, "API Testing — Supabase Edge Functions", "", headers, samples)
    add_dv(ws, lr(wb, "HttpMethod"), "B3:B5000")
    add_dv(ws, lr(wb, "AuthType"), "D3:D5000")
    add_dv(ws, lr(wb, "StatusCode"), "I3:I5000")
    add_dv(ws, lr(wb, "PerfRating"), "J3:J5000")
    add_dv(ws, lr(wb, "SecurityResult"), "K3:K5000")
    add_dv(ws, lr(wb, "PassFail"), "L3:L5000")
    cf_verdict(ws, "L")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_db(wb: Workbook):
    ws = wb.create_sheet("10 Database Validation")
    headers = ["Table", "Columns", "Constraints", "Relationships", "Insert", "Update", "Delete", "Foreign Keys", "Indexes", "Triggers", "Views", "Validation Status", "Remarks"]
    samples = [
        ("profiles", "id, credits, plan, onboarding_completed, streak_days", "PK=auth.users", "1:1 auth.users", "Pass", "Pass", "Pass", "Pass", "Pass", "Partial", "N/A", "Pass", ""),
        ("subscriptions", "user_id, plan, status, stripe ids", "status checks", "profiles", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "N/A", "Pass", "MRR source"),
        ("credit_transactions", "user_id, amount, reason", "signed amount", "profiles", "Pass", "N/A", "N/A", "Pass", "Pass", "N/A", "N/A", "Pass", "Not money"),
        ("payment_orders", "amount_paise, status", "incl refunded", "profiles", "Pass", "Pass", "N/A", "Pass", "Pass", "N/A", "N/A", "Pass", "INR"),
        ("sessions", "nullable scores", "RLS user-owned", "answers/debriefs", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "N/A", "Pass", "Null≠0"),
        ("questions / mock_tests", "exam_type, options jsonb", "RLS", "exam bank", "Pass", "Pass", "Pass", "Pass", "Partial", "N/A", "N/A", "Partial", "Thin seed"),
    ]
    write_table(ws, "Database Validation", "", headers, samples)
    for col in list("EFGHIJKL"):
        add_dv(ws, lr(wb, "PassFail"), f"{col}3:{col}5000")
        cf_verdict(ws, col)
    finish_sheet(ws, 2, len(samples), len(headers))


def build_auth(wb: Workbook):
    ws = wb.create_sheet("11 Authentication Testing")
    headers = [
        "Portal/Role", "Login", "Signup", "Forgot Password", "Reset Password", "OTP",
        "Email Verification", "MFA", "Role Permissions", "Session", "Logout",
        "Unauthorized Access", "Status", "Remarks",
    ]
    samples = [
        ("Guest", "N/A", "Pass", "N/A", "N/A", "N/A", "N/A", "N/A", "Pass", "N/A", "N/A", "Pass", "Pass", "Public marketing"),
        ("Free User", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Limited live"),
        ("Pro User", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Overlay+calendar"),
        ("Max User", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Priority models"),
        ("Admin", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "requireAdmin"),
    ]
    write_table(ws, "Authentication Testing", "", headers, samples)
    for col in list("BCDEFGHIJKLM"):
        add_dv(ws, lr(wb, "PassFail"), f"{col}3:{col}5000")
        cf_verdict(ws, col)
    finish_sheet(ws, 2, len(samples), len(headers))


def build_security(wb: Workbook):
    ws = wb.create_sheet("12 Security Testing")
    headers = [
        "Target", "SQL Injection", "XSS", "CSRF", "Rate Limiting", "Authentication",
        "Authorization", "File Upload", "Input Validation", "Sensitive Data",
        "Headers", "Cookies", "JWT", "Encryption", "Status", "Remarks",
    ]
    samples = [
        ("Edge shared utils", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "BYOK removed"),
        ("parse-document", "N/A", "N/A", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "Key in header"),
        ("stripe-webhook", "N/A", "N/A", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Pass", "Pass", "N/A", "N/A", "Pass", "Pass", "HMAC"),
        ("Admin bulk_update_users", "Pass", "N/A", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Pass", "N/A", "N/A", "Pass", "Pass", "Pass", "M5"),
        ("Resume storage", "N/A", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Bucket RLS"),
        (".env.production history", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Fail", "N/A", "N/A", "Partial", "N/A", "Fail", "Rotate anon key"),
    ]
    write_table(ws, "Security Testing (OWASP-style)", "", headers, samples)
    for col in list("BCDEFGHIJKLMNO"):
        add_dv(ws, lr(wb, "SecurityResult"), f"{col}3:{col}5000")
    cf_verdict(ws, "O")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_perf(wb: Workbook):
    ws = wb.create_sheet("13 Performance Testing")
    headers = ["Screen", "API", "Load Time", "TTFB", "Database Query", "Memory", "CPU", "Network", "Optimization Needed", "Status"]
    samples = [
        ("/app/dashboard", "profile+sessions", "1200ms", "180ms", "90ms", "85MB", "12%", "220KB", "No", "Pass"),
        ("/app/live", "generate-answer", "900ms", "200ms", "40ms", "140MB", "28%", "80KB", "Partial", "Pass"),
        ("/app/admin/users", "paginated users", "2100ms", "320ms", "400ms", "110MB", "18%", "450KB", "Yes", "Partial"),
        ("/pricing", "billing_settings", "800ms", "150ms", "50ms", "70MB", "8%", "160KB", "No", "Pass"),
        ("/app/mock-test/session", "select-test-questions", "1500ms", "250ms", "180ms", "95MB", "15%", "300KB", "Partial", "Pass"),
    ]
    write_table(ws, "Performance Testing", "", headers, samples)
    add_dv(ws, lr(wb, "YesNo"), "I3:I5000")
    add_dv(ws, lr(wb, "PassFail"), "J3:J5000")
    cf_verdict(ws, "J")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_domain(wb, title, sheet_title, headers, samples, dv_map):
    ws = wb.create_sheet(title)
    write_table(ws, sheet_title, "", headers, samples)
    for col, list_name in dv_map.items():
        add_dv(ws, lr(wb, list_name), f"{col}3:{col}5000")
        if list_name in ("PassFail", "YesNo", "SecurityResult"):
            cf_verdict(ws, col)
        if list_name in ("Priority", "Severity", "Risk"):
            cf_severity(ws, col)
    finish_sheet(ws, 2, len(samples), len(headers))
    return ws


def build_live(wb: Workbook):
    headers = [
        "Case ID", "STT Stream", "Diarization", "AI Hint", "AI Answer", "Credit Deduct",
        "Overlay Window", "Hotkeys", "Pro Gate", "Session Persist", "Debrief Link",
        "Electron Boot", "Pass", "Status", "Comments",
    ]
    samples = [
        ("LC-01", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Completed", "Pro happy path"),
        ("LC-02", "Pass", "Pass", "Pass", "N/A", "Pass", "N/A", "N/A", "N/A", "Pass", "Pass", "N/A", "Pass", "Completed", "Web live"),
        ("LC-03", "Pass", "Partial", "Fail", "Fail", "N/A", "N/A", "N/A", "Pass", "N/A", "N/A", "N/A", "Fail", "In Progress", "402 path TC-011"),
        ("LC-04", "N/A", "N/A", "N/A", "N/A", "N/A", "Not Run", "Pass", "Pass", "N/A", "N/A", "Not Run", "Not Run", "Ready", "Packaged unsigned"),
        ("LC-05", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Partial", "Pass", "Completed", "Hotkey no dup"),
    ]
    build_domain(
        wb, "14 Live Coaching Testing", "Domain: Live Coaching & Electron Overlay (replaces Marketplace)",
        headers, samples,
        {c: "PassFail" for c in list("BCDEFGHIJKLM")} | {"N": "Status"},
    )


def build_gov(wb: Workbook):
    headers = [
        "Case ID", "Exam Select", "Create Test", "Question Load", "Timer", "Answer Save",
        "Submit Atomic", "Score Persist", "Results UI", "Analytics", "Admin Filter",
        "PDF Import", "Bank Coverage", "Pass", "Status", "Comments",
    ]
    samples = [
        ("GE-01", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "N/A", "Partial", "Pass", "Completed", "Happy path"),
        ("GE-02", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Partial", "N/A", "N/A", "Partial", "Pass", "Completed", "Idempotent submit"),
        ("GE-03", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "Pass", "N/A", "N/A", "Pass", "Pass", "Completed", "Expanded MCQ seeds"),
        ("GE-04", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Pass", "N/A", "N/A", "Pass", "Completed", "P1-5 filter"),
        ("GE-05", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Blocked", "N/A", "Blocked", "Blocked", "Provider quota"),
    ]
    build_domain(
        wb, "15 Gov Exam Mock Testing", "Domain: Gov Exam Mock Test Engine (replaces Checkout)",
        headers, samples,
        {c: "PassFail" for c in list("BCDEFGHIJKLMN")} | {"O": "Status"},
    )


def build_billing(wb: Workbook):
    headers = [
        "Case ID", "Gateway", "Success", "Failure", "Cancel", "Timeout", "Webhook",
        "Refund", "Partial Refund", "Duplicate Payment", "Credit Grant", "Credit Clawback",
        "Plan Gate", "Transaction History", "Settlement", "Pass", "Status", "Comments",
    ]
    samples = [
        ("BL-01", "Stripe", "Pass", "Pass", "Pass", "Not Run", "Pass", "Not Run", "Not Run", "Pass", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Completed", "Pro monthly"),
        ("BL-02", "Stripe", "Pass", "N/A", "N/A", "N/A", "Pass", "N/A", "N/A", "N/A", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Completed", "Max price IDs"),
        ("BL-03", "Stripe", "N/A", "Pass", "N/A", "N/A", "Pass", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Pass", "N/A", "Pass", "Completed", "payment_failed wallet"),
        ("BL-04", "Razorpay", "Pass", "Pass", "Pass", "Not Run", "Pass", "Not Run", "N/A", "Pass", "Pass", "N/A", "Partial", "Pass", "N/A", "Pass", "Completed", "One-time only"),
        ("BL-05", "Stripe", "N/A", "N/A", "N/A", "N/A", "Pass", "Not Run", "Not Run", "N/A", "N/A", "Not Run", "N/A", "Pass", "N/A", "Not Run", "Ready", "P4-2 refund"),
        ("BL-06", "Both", "Pass", "N/A", "N/A", "N/A", "Pass", "N/A", "N/A", "N/A", "Pass", "N/A", "Pass", "Pass", "N/A", "Pass", "Completed", "Settings ledger"),
    ]
    build_domain(
        wb, "16 Billing Credits Testing", "Domain: Billing, Credits & Subscriptions (replaces Payment)",
        headers, samples,
        {**{c: "PassFail" for c in list("CDEFGHIJKLMNO")}, "P": "PassFail", "Q": "Status"},
    )


def build_rbac(wb: Workbook):
    ws = wb.create_sheet("17 Role Based Testing")
    headers = [
        "Permission/Feature", "Module",
        "Guest", "Free User", "Pro User", "Max User", "Admin",
        "Permissions", "Status", "Expected Summary",
    ]
    samples = [
        ("Access /app", "Auth & Onboarding", "No", "Yes", "Yes", "Yes", "Yes", "Auth required", "Pass", "Guests → login"),
        ("Live answer limited", "Live Practice Coach", "No", "Yes", "Yes", "Yes", "Yes", "Free limited", "Pass", "live_rehearsal"),
        ("Desktop overlay", "Live Practice Coach", "No", "No", "Yes", "Yes", "Yes", "Pro+", "Pass", "requirePlan pro"),
        ("Company research", "Company Research", "No", "No", "Yes", "Yes", "Yes", "Pro+", "Pass", ""),
        ("Calendar sync", "Interviews & Calendar", "No", "No", "Yes", "Yes", "Yes", "Pro+", "Pass", ""),
        ("Priority models", "Billing & Credits", "No", "No", "No", "Yes", "Yes", "Max", "Pass", ""),
        ("Admin revenue", "Admin Portal", "No", "No", "No", "No", "Yes", "admin role", "Pass", "requireAdmin"),
        ("Promo CRUD", "Admin Portal", "No", "No", "No", "No", "Yes", "admin", "Pass", ""),
        ("View pricing", "Marketing Site", "Yes", "Yes", "Yes", "Yes", "Yes", "Public", "Pass", ""),
    ]
    write_table(ws, "Role Based Testing — Clarify plan entitlements", "", headers, samples)
    for col in list("CDEFG"):
        add_dv(ws, lr(wb, "YesNo"), f"{col}3:{col}5000")
    add_dv(ws, lr(wb, "Modules"), "B3:B5000")
    add_dv(ws, lr(wb, "PassFail"), "I3:I5000")
    cf_verdict(ws, "I")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_regression(wb: Workbook):
    ws = wb.create_sheet("18 Regression Testing")
    headers = [
        "Regression ID", "Module", "Feature", "QA Owner", "Executed",
        "Pass", "Fail", "Blocked", "Est Hours", "Actual Hours", "Remarks",
    ]
    samples = [
        ("REG-01", "Auth & Onboarding", "Login/OAuth/MFA", "Raj Balani", "Yes", 4, 0, 0, 4.0, 3.5, "Smoke green"),
        ("REG-02", "Live Practice Coach", "STT + answer + overlay", "Shreya Patil", "Yes", 4, 0, 0, 6.0, 5.0, "BUG-001 closed"),
        ("REG-03", "Billing & Credits", "Checkout + webhooks", "Raj Balani", "Yes", 5, 0, 0, 5.0, 4.5, "Money path"),
        ("REG-04", "Gov Exam Mock Tests", "Create/submit/results", "Raj Balani", "Yes", 3, 0, 1, 4.0, 3.0, "PDF blocked"),
        ("REG-05", "Admin Portal", "Users/revenue/questions", "Raj Balani", "Yes", 4, 0, 0, 4.0, 3.5, ""),
        ("REG-06", "Electron Desktop Overlay", "Boot + hotkeys", "Shreya Patil", "Partial", 2, 0, 1, 3.0, 1.5, "Unsigned"),
    ]
    write_table(ws, "Regression Testing", "", headers, samples)
    add_dv(ws, lr(wb, "Modules"), "B3:B5000")
    add_dv(ws, lr(wb, "Testers"), "D3:D5000")
    add_dv(ws, lr(wb, "YesNo"), "E3:E5000")
    t = 3 + len(samples)
    ws.cell(t, 1, "TOTAL")
    for col, letter in [(6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")]:
        ws.cell(t, col, f"=SUM({letter}3:{letter}{t-1})")
        ws.cell(t, col).font = Font(bold=True)
    finish_sheet(ws, 2, len(samples) + 1, len(headers))


def build_prod(wb: Workbook):
    ws = wb.create_sheet("19 Production Readiness")
    headers = [
        "Module", "Feature Complete", "API Complete", "Database Ready", "Security Passed",
        "Performance Passed", "Responsive", "Accessibility", "SEO", "Monitoring",
        "Logging", "Deployment Ready", "Backup", "Recovery", "Documentation",
        "Ready Count", "Total", "Readiness %", "Status", "Remarks",
    ]
    base = [
        ("Auth & Onboarding", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"),
        ("Live Practice Coach", "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "Partial", "N/A", "Yes", "Yes", "Yes", "Yes", "Partial", "Yes"),
        ("Billing & Credits", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "N/A", "Partial", "Yes", "Yes", "Yes", "Partial", "Yes"),
        ("Gov Exam Mock Tests", "Partial", "Yes", "Partial", "Yes", "Yes", "Yes", "Partial", "Partial", "Partial", "Yes", "Partial", "Yes", "Partial", "Partial"),
        ("Admin Portal", "Partial", "Yes", "Yes", "Yes", "Partial", "Partial", "Partial", "N/A", "Partial", "Yes", "Yes", "Yes", "Partial", "Partial"),
        ("Electron Desktop Overlay", "Partial", "N/A", "N/A", "Partial", "Yes", "N/A", "Partial", "N/A", "Partial", "Partial", "Partial", "Yes", "Partial", "Partial"),
        ("Marketing Site", "Yes", "N/A", "Yes", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"),
        ("Security & Platform", "Partial", "Yes", "Partial", "Partial", "Yes", "N/A", "N/A", "N/A", "Partial", "Yes", "Yes", "Yes", "Partial", "Yes"),
    ]
    rows = [list(m) + [None, 14, None, None, ""] for m in base]
    write_table(ws, "Production Readiness Checklist", "", headers, rows)
    for r in range(3, 3 + len(rows)):
        ws.cell(r, 16, f'=IFERROR(COUNTIF(B{r}:O{r},"Yes"),0)')
        ws.cell(r, 16).protection = LOCKED
        ws.cell(r, 17, 14)
        ws.cell(r, 17).protection = LOCKED
        ws.cell(r, 18, f"=IFERROR(IF(Q{r}=0,0,P{r}/Q{r}),0)")
        ws.cell(r, 18).number_format = "0.0%"
        ws.cell(r, 18).protection = LOCKED
        ws.cell(r, 19, f'=IFERROR(IF(R{r}>=0.9,"Ready",IF(R{r}>=0.7,"Conditional","Not Ready")),"Not Ready")')
        ws.cell(r, 19).protection = LOCKED
    for col in list("BCDEFGHIJKLMNO"):
        add_dv(ws, lr(wb, "YesNo"), f"{col}3:{col}5000")
    add_dv(ws, lr(wb, "Modules"), "A3:A5000")
    add_dv(ws, lr(wb, "ReadyStatus"), "S3:S5000")
    finish_sheet(ws, 2, len(rows), len(headers))


def build_signoff(wb: Workbook):
    ws = wb.create_sheet("20 Release Sign-Off")
    headers = [
        "Release Version", "Build", "Environment", "QA Lead", "Developer Lead", "Product Owner",
        "Passed", "Known Issues", "Blockers", "Risk Level", "Deployment Approval", "Go/No-Go", "Remarks",
    ]
    samples = [
        ("1.0.0-beta", "1.0.0", "Staging", "Shreya Patil", "Dev Lead", "Product",
         "Yes", "BUG-OPEN-01 Razorpay one-time; unsigned Electron if no cert",
         "Calendar/Resend/signing secrets if promising those features", "Medium", "Pending", "Conditional Go",
         "Launch-pass code + expanded MCQ seeds; apply migration; ops secrets still Blocked"),
        ("1.0.0", "1.0.0", "Production", "Shreya Patil", "Dev Lead", "Product",
         "No", "See CHANGELOG Launch gap matrix", "Code signing; anon key rotation; Calendar/Resend if promised", "High", "Pending", "No-Go",
         "Await cert + ops secrets"),
        ("1.0.1", "TBD", "Staging", "Shreya Patil", "Dev Lead", "Product",
         "No", "", "", "Low", "Deferred", "No-Go", "Placeholder"),
    ]
    write_table(ws, "Release Sign-Off", "", headers, samples)
    add_dv(ws, lr(wb, "Release"), "A3:A5000")
    add_dv(ws, lr(wb, "Environment"), "C3:C5000")
    add_dv(ws, lr(wb, "YesNo"), "G3:G5000")
    add_dv(ws, lr(wb, "Risk"), "J3:J5000")
    add_dv(ws, lr(wb, "Approval"), "K3:K5000")
    add_dv(ws, lr(wb, "Decision"), "L3:L5000")
    cf_severity(ws, "J")
    finish_sheet(ws, 2, len(samples), len(headers))


def build_nav_hub(wb: Workbook):
    ws = wb.create_sheet("NAV Hub")
    ws["A1"] = "NAV Hub — Clarify AI QA Workbook (Power BI style)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    add_nav_bar(ws, "NAV Hub")
    ws["A2"] = "Click any link to jump. Treat AutoFilter columns on Features/Bugs as slicers."
    ws["A2"].font = SUB_FONT

    groups = [
        ("Executive", [
            ("01 Dashboard", "KPIs, charts, inventory snapshot"),
            ("24 Launch Status", "Fixed / Open / Blocked rollup"),
            ("20 Release Sign-Off", "Go / No-Go"),
            ("19 Production Readiness", "Module readiness %"),
        ]),
        ("Scope & Coverage", [
            ("02 Module Master", "20 modules"),
            ("03 Feature Inventory", "Routes, deep links, how-it-works"),
            ("23 Module Playbooks", "Happy-path runbooks"),
            ("17 Role Based Testing", "Plan gates"),
        ]),
        ("Execution", [
            ("QA Tasks", "Shreya & Raj task board"),
            ("04 Test Case Repository", "Case library"),
            ("05 E2E User Flows", "Golden paths"),
            ("25 Smoke Pack", "Must-pass (~20)"),
            ("18 Regression Testing", "Release pack"),
            ("Daily Log", "Trend feed"),
        ]),
        ("Defects & Quality", [
            ("06 Bug Tracker", "Lifecycle (source of truth)"),
            ("Dev Tasks", "Live bug mirror for developers"),
            ("Dev Team", "Dev roster + load"),
            ("07 UI-UX Testing", "Visual / a11y"),
            ("08 Mobile Responsiveness", "Breakpoints"),
            ("13 Performance Testing", "Latency"),
        ]),
        ("Platform", [
            ("09 API Testing", "Edge functions"),
            ("10 Database Validation", "Schema / RLS"),
            ("11 Authentication Testing", "Auth + MFA"),
            ("12 Security Testing", "OWASP-style"),
        ]),
        ("Domains", [
            ("14 Live Coaching Testing", "Coach + overlay"),
            ("15 Gov Exam Mock Testing", "MCQ engine"),
            ("16 Billing Credits Testing", "Stripe / Razorpay"),
        ]),
        ("Ops & Hire Pack", [
            ("21 Test Credentials", "Placeholders only"),
            ("22 Environments", "URLs + secrets checklist"),
            ("QA Team", "Shreya Patil · Raj Balani"),
            ("00 Read Me", "How to use"),
            ("Lists", "Dropdown vocabulary"),
        ]),
    ]

    row = 4
    for group, items in groups:
        ws.cell(row, 1, group).font = SECTION_FONT
        ws.cell(row, 1).fill = NAV_FILL
        ws.cell(row, 1).font = NAV_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for sheet, desc in items:
            link = ws.cell(row, 1, sheet)
            link.hyperlink = f"#'{sheet}'!A1"
            link.font = LINK_FONT
            ws.cell(row, 2, desc).alignment = WRAP
            back = ws.cell(row, 3, "Open →")
            back.hyperlink = f"#'{sheet}'!A1"
            back.font = LINK_FONT
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_lists(wb)
    build_daily_log(wb)
    build_qa_team(wb)
    build_qa_tasks(wb)
    build_dev_team(wb)
    build_dev_tasks(wb)
    build_readme(wb)
    build_dashboard(wb)
    build_modules(wb)
    build_features(wb)
    build_test_cases(wb)
    build_e2e(wb)
    build_bugs(wb)
    build_uiux(wb)
    build_mobile(wb)
    build_api(wb)
    build_db(wb)
    build_auth(wb)
    build_security(wb)
    build_perf(wb)
    build_live(wb)
    build_gov(wb)
    build_billing(wb)
    build_rbac(wb)
    build_regression(wb)
    build_prod(wb)
    build_signoff(wb)
    build_credentials(wb)
    build_environments(wb)
    build_playbooks(wb)
    build_launch_status(wb)
    build_smoke_pack(wb)
    build_nav_hub(wb)

    # Apply navigable strip to every content sheet
    for name in wb.sheetnames:
        if name == "Lists":
            continue
        add_nav_bar(wb[name], name)

    # Order: Read Me, NAV, Dashboard, 02-20, hire-pack sheets, Daily Log, Lists at end
    desired = [
        "00 Read Me", "NAV Hub", "01 Dashboard",
        "QA Tasks", "QA Team", "Dev Team", "Dev Tasks",
        "02 Module Master", "03 Feature Inventory", "04 Test Case Repository",
        "05 E2E User Flows", "06 Bug Tracker", "07 UI-UX Testing", "08 Mobile Responsiveness",
        "09 API Testing", "10 Database Validation", "11 Authentication Testing",
        "12 Security Testing", "13 Performance Testing",
        "14 Live Coaching Testing", "15 Gov Exam Mock Testing", "16 Billing Credits Testing",
        "17 Role Based Testing", "18 Regression Testing", "19 Production Readiness",
        "20 Release Sign-Off",
        "21 Test Credentials", "22 Environments", "23 Module Playbooks",
        "24 Launch Status", "25 Smoke Pack",
        "Daily Log", "Lists",
    ]
    for i, name in enumerate(desired):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    for path in (OUT, OUT_FALLBACK, OUT_COMPLETE, OUT_COMPLETE_V2, OUT_FULL):
        try:
            wb.save(path)
            print(f"Wrote {path}")
        except PermissionError:
            print(f"Locked (skipped): {path}")
    print("Sheets:", wb.sheetnames)
    print(f"Features loaded: {len(FEATURES)} | Bugs loaded: {len(BUGS)}")
    print("Bug status:", dict(bug_status_counts()))


if __name__ == "__main__":
    main()
