"""One-off extractor for QA workbook — not a test file."""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = Path(r"C:\Users\TECH-GENIUSES\Downloads\clarity-assistant\Clarify_AI_BB_Manual_QA_Workbook (2).xlsx")
OUT = Path(r"C:\Users\TECH-GENIUSES\Downloads\clarity-assistant\_qa_cases_extract.json")


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v


def main():
    wb = load_workbook(PATH, data_only=False)
    ws = wb["05 All Test Cases"]

    # Find header row: look for TC-ID / Test Case / Status-like cells in first 15 rows
    header_row = None
    headers = []
    for r in range(1, 16):
        vals = [cell_str(ws.cell(r, c).value) for c in range(1, 30)]
        nonempty = sum(1 for v in vals if v)
        joined = " | ".join(str(v) for v in vals if v)
        if any(
            x in joined.lower()
            for x in ("tc-id", "test case id", "test case", "actual result", "pass/fail", "status")
        ):
            # Prefer row that has many non-empty headers
            if nonempty >= 8:
                header_row = r
                headers = vals
                break
        if header_row is None and nonempty >= 10:
            header_row = r
            headers = vals

    print("HEADER_ROW", header_row)
    print("HEADERS:")
    for i, h in enumerate(headers, 1):
        if h:
            print(f"  {i} ({get_column_letter(i)}): {h}")

    # Normalize keys (keep position)
    keys = []
    for i, h in enumerate(headers):
        if h:
            keys.append(str(h))
        else:
            keys.append(f"_col{i+1}")

    # Images / drawings
    images_info = []
    try:
        for img in getattr(ws, "_images", []) or []:
            anchor = getattr(img, "anchor", None)
            images_info.append(
                {
                    "type": type(img).__name__,
                    "anchor": str(anchor),
                    "width": getattr(img, "width", None),
                    "height": getattr(img, "height", None),
                }
            )
    except Exception as e:
        images_info.append({"error": str(e)})

    # Also check drawings XML for hyperlinks in media
    drawing_links = []

    rows_out = []
    start = (header_row or 1) + 1
    for r in range(start, ws.max_row + 1):
        item = {"excel_row": r}
        empty = True
        comments = []
        links = []
        for c, k in enumerate(keys, 1):
            cell = ws.cell(r, c)
            val = cell_str(cell.value)
            if val is not None and val != "":
                empty = False
                # truncate huge cells later
                item[k] = val
            if cell.hyperlink is not None:
                target = getattr(cell.hyperlink, "target", None) or getattr(cell.hyperlink, "ref", None)
                if target:
                    links.append({"col": k, "target": str(target), "display": val})
                    item[f"{k}__hyperlink"] = str(target)
            if cell.comment is not None:
                comments.append({"col": k, "text": cell.comment.text})
        # Also detect URLs embedded in any string cell
        url_hits = []
        for k, v in list(item.items()):
            if isinstance(v, str):
                for m in re.findall(r"https?://[^\s\)\]\"']+", v):
                    url_hits.append({"col": k, "url": m})
        if comments:
            item["_cell_comments"] = comments
        if links:
            item["_hyperlinks"] = links
        if url_hits:
            item["_urls_in_text"] = url_hits
        if empty:
            continue
        rows_out.append(item)

    # Status distribution
    status_cols = [
        k
        for k in keys
        if k
        and any(x in k.lower() for x in ("status", "pass", "result", "verdict"))
        and "expected" not in k.lower()
        and "actual" not in k.lower()
    ]
    print("CANDIDATE STATUS COLS", status_cols)
    for sk in status_cols:
        vals = [str(r.get(sk, "")).strip() for r in rows_out if r.get(sk) not in (None, "")]
        print("DIST", sk, Counter(vals).most_common(20))

    # Defect log + extra bugs
    extra = {}
    for sheet in ("39 Defect Log", "Anushka-Extra bug", "00b Section Completion Gate"):
        if sheet not in wb.sheetnames:
            continue
        s = wb[sheet]
        # header guess row 1
        h = [cell_str(s.cell(1, c).value) for c in range(1, min(s.max_column, 40) + 1)]
        srows = []
        for r in range(2, s.max_row + 1):
            d = {"excel_row": r}
            empty = True
            for c, hk in enumerate(h, 1):
                if not hk:
                    continue
                v = cell_str(s.cell(r, c).value)
                if v not in (None, ""):
                    empty = False
                    d[str(hk)] = v
                cell = s.cell(r, c)
                if cell.hyperlink is not None:
                    t = getattr(cell.hyperlink, "target", None)
                    if t:
                        d[f"{hk}__hyperlink"] = str(t)
            if not empty:
                srows.append(d)
        extra[sheet] = {"headers": h, "rows": srows}

    payload = {
        "header_row": header_row,
        "headers": keys,
        "image_count": len(getattr(ws, "_images", []) or []),
        "images_info": images_info,
        "test_cases": rows_out,
        "other_sheets": extra,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("WROTE", OUT, "cases=", len(rows_out))


if __name__ == "__main__":
    main()
